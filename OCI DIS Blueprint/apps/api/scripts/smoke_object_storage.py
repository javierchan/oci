"""Exercise the production import-to-export Object Storage lifecycle."""

from __future__ import annotations

import argparse
from io import BytesIO
import json
from pathlib import Path
import sys
import time
from typing import Any

import httpx
from openpyxl import load_workbook

API_ROOT = Path(__file__).resolve().parents[1]
if str(API_ROOT) not in sys.path:
    sys.path.insert(0, str(API_ROOT))

from app.services import storage_service


def _response_json(response: httpx.Response) -> dict[str, Any]:
    if response.is_error:
        raise RuntimeError(
            f"{response.request.method} {response.request.url} returned "
            f"{response.status_code}: {response.text}"
        )
    payload = response.json()
    if not isinstance(payload, dict):
        raise RuntimeError("Expected an object response")
    return payload


def _capture_workbook(client: httpx.Client) -> bytes:
    response = client.get("/exports/template/xlsx")
    response.raise_for_status()
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["Catálogo de Integraciones"]
    columns = {str(cell.value): cell.column for cell in sheet[1] if cell.value}
    values: dict[str, object] = {
        "#": 1,
        "ID de Interfaz": "INT-STORAGE-SMOKE-001",
        "Marca": "Object Storage Smoke",
        "Proceso de Negocio": "Order to Cash",
        "Interfaz": "Publish governed order event",
        "Frecuencia": "Cada 1 hora",
        "Tipo Trigger OIC": "Event Trigger",
        "Payload por Ejecución (KB)": 128,
        "Fan-out (Si/No)": "No",
        "# Destinos": 1,
        "Sistema de Origen": "Oracle ERP Cloud",
        "Sistema de Destino": "Order Fulfillment",
        "TBQ": "Y",
        "Patrón Seleccionado (Manual)": "#02",
        "Racional del Patrón (Manual)": "Governed asynchronous event delivery.",
        "Retry Policy": "3 attempts; exponential backoff; DLQ",
        "Herramientas Core Cuantificables / Volumétricas": "OCI Streaming | OIC Gen3",
    }
    for header, value in values.items():
        sheet.cell(2, columns[header], value)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _poll(
    client: httpx.Client,
    path: str,
    *,
    timeout_seconds: int,
) -> dict[str, Any]:
    deadline = time.monotonic() + timeout_seconds
    while time.monotonic() < deadline:
        payload = _response_json(client.get(path))
        status = str(payload.get("status", "")).lower()
        if status in {"completed", "failed"}:
            if status == "failed":
                raise RuntimeError(f"Job failed: {json.dumps(payload, default=str)}")
            return payload
        time.sleep(1)
    raise TimeoutError(f"Timed out waiting for {path}")


def run(base_url: str, timeout_seconds: int) -> dict[str, object]:
    project_id: str | None = None
    project_deleted = False
    evidence: dict[str, object] = {}
    headers = {"X-Actor-Id": "object-storage-smoke", "X-Actor-Role": "Admin"}
    with httpx.Client(base_url=base_url.rstrip("/"), headers=headers, timeout=60) as client:
        try:
            project = _response_json(
                client.post(
                    "/projects/",
                    json={
                        "name": f"Object Storage Smoke {int(time.time())}",
                        "owner_id": "object-storage-smoke",
                        "description": "Ephemeral Object Storage lifecycle validation.",
                    },
                )
            )
            project_id = str(project["id"])
            workbook_bytes = _capture_workbook(client)
            batch = _response_json(
                client.post(
                    f"/imports/{project_id}",
                    files={
                        "file": (
                            "object-storage-smoke.xlsx",
                            workbook_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    },
                )
            )
            batch = _poll(
                client,
                f"/imports/{project_id}/{batch['id']}",
                timeout_seconds=timeout_seconds,
            )
            if batch.get("loaded_count") != 1:
                raise RuntimeError(f"Expected one loaded row, received {batch.get('loaded_count')}")

            import_keys = storage_service.list_keys(f"imports/{project_id}/")
            if len(import_keys) != 1:
                raise RuntimeError(f"Expected one import object, received {import_keys}")

            recalc = _response_json(client.post(f"/recalculate/{project_id}"))
            recalc = _poll(
                client,
                f"/recalculate/{project_id}/jobs/{recalc['job_id']}",
                timeout_seconds=timeout_seconds,
            )
            snapshot_id = str(recalc.get("snapshot_id") or "")
            if not snapshot_id:
                raise RuntimeError("Recalculation completed without a snapshot")

            export = _response_json(
                client.post(
                    f"/exports/{project_id}/xlsx",
                    params={"snapshot_id": snapshot_id},
                )
            )
            download = client.get(
                f"/exports/{project_id}/jobs/{export['job_id']}/download"
            )
            download.raise_for_status()
            if not download.content.startswith(b"PK"):
                raise RuntimeError("Downloaded export is not an XLSX archive")
            export_keys = storage_service.list_keys(f"exports/{project_id}/")
            if len(export_keys) != 2:
                raise RuntimeError(f"Expected export artifact and manifest, received {export_keys}")

            evidence = {
                "project_id": project_id,
                "import_batch_id": batch["id"],
                "snapshot_id": snapshot_id,
                "export_job_id": export["job_id"],
                "import_keys_before_cleanup": import_keys,
                "export_keys_before_cleanup": export_keys,
                "download_bytes": len(download.content),
            }
            _response_json(client.post(f"/projects/{project_id}/archive"))
            _response_json(client.delete(f"/projects/{project_id}"))
            project_deleted = True
            remaining_keys = [
                *storage_service.list_keys(f"imports/{project_id}/"),
                *storage_service.list_keys(f"exports/{project_id}/"),
                *storage_service.list_keys(f"synthetic/{project_id}/"),
            ]
            if remaining_keys:
                raise RuntimeError(f"Project cleanup left Object Storage keys: {remaining_keys}")
            evidence["remaining_project_keys"] = remaining_keys
            evidence["status"] = "passed"
            return evidence
        finally:
            if project_id and not project_deleted:
                try:
                    client.post(f"/projects/{project_id}/archive")
                    client.delete(f"/projects/{project_id}")
                except httpx.HTTPError:
                    pass


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--base-url", default="http://127.0.0.1:8000/api/v1")
    parser.add_argument("--timeout-seconds", type=int, default=180)
    arguments = parser.parse_args()
    print(json.dumps(run(arguments.base_url, arguments.timeout_seconds), indent=2))


if __name__ == "__main__":
    main()
