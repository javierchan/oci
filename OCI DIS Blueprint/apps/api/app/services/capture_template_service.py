"""Governed offline-capture workbook generation and metadata assembly."""

from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    DictionaryOption,
    PatternDefinition,
    ServiceCapabilityProfile,
    ServiceEvidenceSource,
    ServiceInteroperabilityRule,
    ServiceLimit,
    ServiceProductVersion,
)
from app.schemas.export import CaptureTemplateMetadata, CaptureTemplateColumnMetadata
from app.services.pattern_support import get_pattern_support


TEMPLATE_VERSION = "3.0.0"
IMPORTER_MIN_VERSION = "3.0.0"
TEMPLATE_FILENAME = f"oci-dis-import-template-v{TEMPLATE_VERSION}.xlsx"
CAPTURE_SHEET_NAME = "Catálogo de Integraciones"
MANIFEST_SHEET_NAME = "_Listas"
CLIENT_CATALOGS_SHEET_NAME = "Catálogos del Cliente"
CAPTURE_ROW_LIMIT = 501


@dataclass(frozen=True)
class TemplateColumnSpec:
    """Single governed definition for one offline-capture column."""

    field: str
    header: str
    section: str
    requirement: str
    data_type: str
    description: str
    good_example: str
    bad_example: str
    app_usage: str
    validation_key: str | None = None
    unit: str | None = None


COLUMNS: tuple[TemplateColumnSpec, ...] = (
    TemplateColumnSpec("seq_number", "#", "Identidad", "Opcional", "Entero", "Orden visible de la integración.", "1", "Uno", "Conserva el orden de captura."),
    TemplateColumnSpec("interface_id", "ID de Interfaz", "Identidad", "Recomendado", "Texto", "Identificador único y estable de la integración.", "INT-001", "Integración nueva", "Permite trazabilidad, duplicados y cobertura."),
    TemplateColumnSpec("owner", "Owner", "Identidad", "Opcional", "Texto", "Persona responsable de completar o gobernar el registro.", "Finance Architect", "Equipo", "Facilita seguimiento operativo.", "CLIENT_OWNERS"),
    TemplateColumnSpec("brand", "Marca", "Negocio", "Requerido", "Texto", "Unidad, marca o entidad de negocio propietaria.", "Retail", "N/A", "Segmenta dashboard y catálogo.", "CLIENT_BRANDS"),
    TemplateColumnSpec("business_process", "Proceso de Negocio", "Negocio", "Requerido", "Texto", "Proceso de negocio que habilita la integración.", "Order to Cash", "Integración", "Agrupa riesgo y cobertura por proceso.", "CLIENT_PROCESSES"),
    TemplateColumnSpec("interface_name", "Interfaz", "Identidad", "Requerido", "Texto", "Nombre funcional que explica qué información se mueve.", "Publicar pedido confirmado", "INT-001", "Es el nombre principal mostrado por la App."),
    TemplateColumnSpec("description", "Descripción", "Negocio", "Recomendado", "Texto", "Descripción breve del propósito y resultado esperado.", "Publica pedidos confirmados al fulfillment.", "TBD", "Aporta contexto a QA y revisión de arquitectura."),
    TemplateColumnSpec("business_criticality", "Criticidad de Negocio", "Negocio", "Recomendado", "Lista", "Impacto operativo o financiero si la integración deja de funcionar.", "Alta", "Importante", "Prioriza riesgo, resiliencia y secuencia de remediación.", "BUSINESS_CRITICALITY"),
    TemplateColumnSpec("status", "Estado", "Gobernanza", "Opcional", "Texto", "Estado general del registro en el assessment.", "En Progreso", "Bien", "Conserva evidencia del levantamiento."),
    TemplateColumnSpec("mapping_status", "Estado de Mapeo", "Gobernanza", "Opcional", "Texto", "Avance del mapeo de campos origen-destino.", "Pendiente", "50", "Ayuda a priorizar trabajo de definición."),
    TemplateColumnSpec("initial_scope", "Alcance Inicial", "Negocio", "Opcional", "Texto", "Indica si la integración pertenece al alcance inicial.", "Sí", "Tal vez", "Preserva el alcance acordado."),
    TemplateColumnSpec("complexity", "Complejidad", "Arquitectura", "Recomendado", "Lista", "Complejidad estimada de implementación.", "Medio", "Normal", "Apoya priorización y revisión.", "COMPLEXITY"),
    TemplateColumnSpec("frequency", "Frecuencia", "Ejecución", "Requerido", "Lista", "Periodicidad gobernada; seleccione una opción de la lista.", "Cada 1 hora", "Frecuente", "Determina ejecuciones por día y volumetría.", "FREQUENCY"),
    TemplateColumnSpec("type", "Tipo", "Ejecución", "Opcional", "Texto", "Clasificación heredada del inventario fuente.", "Batch", "Proceso", "Mantiene fidelidad con workbooks existentes."),
    TemplateColumnSpec("interface_status", "Estado Interfaz", "Gobernanza", "Opcional", "Texto", "Estado específico de la interfaz; Duplicado 2 se excluye.", "Activo", "Duplicado", "Controla exclusiones heredadas."),
    TemplateColumnSpec("is_real_time", "Tiempo Real (Si/No)", "Ejecución", "Recomendado", "Lista", "Indica si el negocio requiere procesamiento en tiempo real.", "Sí", "Realtime", "Aporta semántica de latencia a QA.", "YES_NO"),
    TemplateColumnSpec("target_latency_sla", "SLA / Latencia Objetivo", "Ejecución", "Recomendado", "Texto", "Objetivo medible de latencia o tiempo máximo de procesamiento.", "p95 < 5 segundos", "Rápido", "Permite validar patrón, timeout y desacoplamiento."),
    TemplateColumnSpec("trigger_type", "Tipo Trigger OIC", "Ejecución", "Recomendado", "Lista", "Mecanismo que inicia la integración.", "REST Trigger", "API", "Activa reglas de QA y sizing.", "TRIGGER_TYPE"),
    TemplateColumnSpec("response_size_kb", "Response Size (KB)", "Volumetría", "Opcional", "Decimal", "Tamaño promedio de respuesta en KB; use 0 solo si realmente no hay respuesta.", "12.5", "12 MB", "Dimensiona respuestas síncronas.", unit="KB"),
    TemplateColumnSpec("payload_per_execution_kb", "Payload por Ejecución (KB)", "Volumetría", "Recomendado", "Decimal", "Tamaño promedio de la solicitud o mensaje en KB.", "150", "1.5 MB", "Alimenta OIC, Streaming, Queue y Functions.", unit="KB"),
    TemplateColumnSpec("is_fan_out", "Fan-out (Si/No)", "Ejecución", "Recomendado", "Lista", "Indica si una ejecución entrega a más de un destino.", "No", "N/A", "Activa validación de número de destinos.", "YES_NO"),
    TemplateColumnSpec("fan_out_targets", "# Destinos", "Ejecución", "Recomendado", "Entero", "Cantidad de destinos cuando Fan-out es Sí; use 1 en una ruta simple.", "3", "Varios", "Ajusta carga y QA de distribución."),
    TemplateColumnSpec("source_system", "Sistema de Origen", "Origen", "Requerido", "Texto", "Sistema que produce el dato o inicia la interacción.", "Oracle ERP Cloud", "Origen", "Construye linaje y topología.", "CLIENT_SYSTEMS"),
    TemplateColumnSpec("source_technology", "Tecnología de Origen", "Origen", "Recomendado", "Texto", "Producto, protocolo o tecnología del origen.", "REST API", "Oracle", "Ayuda a validar conectividad.", "CLIENT_TECHNOLOGIES"),
    TemplateColumnSpec("source_api_reference", "API Reference", "Origen", "Opcional", "Texto", "URL, operación o referencia técnica del contrato origen.", "/orders/{id}", "Ver documento", "Preserva evidencia técnica."),
    TemplateColumnSpec("source_owner", "Propietario de Origen", "Origen", "Opcional", "Texto", "Equipo responsable del sistema origen.", "ERP Platform Team", "Javier", "Facilita validación con responsables.", "CLIENT_OWNERS"),
    TemplateColumnSpec("destination_system", "Sistema de Destino", "Destino", "Requerido", "Texto", "Sistema que recibe o consume el dato.", "Order Fulfillment", "Destino", "Construye linaje y topología.", "CLIENT_SYSTEMS"),
    TemplateColumnSpec("destination_technology_1", "Tecnología de Destino #1", "Destino", "Recomendado", "Texto", "Tecnología principal del destino.", "Oracle ATP", "Base de datos", "Ayuda a validar adaptadores y servicios.", "CLIENT_TECHNOLOGIES"),
    TemplateColumnSpec("destination_technology_2", "Tecnología de Destino #2", "Destino", "Opcional", "Texto", "Tecnología secundaria cuando el destino combina dos componentes.", "OCI Streaming", "N/A", "Conserva una segunda tecnología sin inventarla.", "CLIENT_TECHNOLOGIES"),
    TemplateColumnSpec("destination_owner", "Propietario de Destino", "Destino", "Opcional", "Texto", "Equipo responsable del sistema destino.", "Fulfillment Team", "Usuario final", "Facilita validación y ownership.", "CLIENT_OWNERS"),
    TemplateColumnSpec("data_security_classification", "Clasificación de Datos / Seguridad", "Seguridad", "Recomendado", "Lista", "Clasificación que gobierna cifrado, acceso, retención y evidencia.", "Confidencial", "Segura", "Activa controles y revisión de seguridad proporcionales.", "DATA_CLASSIFICATION"),
    TemplateColumnSpec("calendarization", "Calendarización", "Ejecución", "Opcional", "Texto", "Horario, ventana o restricción de calendario.", "L-V 22:00 America/Mexico_City", "Noche", "Documenta ventanas operativas."),
    TemplateColumnSpec("retention_processing_window", "Retención / Ventana de Procesamiento", "Operación", "Recomendado", "Texto", "Retención de mensajes u objetos y ventana disponible para procesar o reprocesar.", "Retener 7 días; procesar 22:00-02:00", "Una semana", "Valida batch, replay, Claim Check y recuperación."),
    TemplateColumnSpec("tbq", "TBQ", "Gobernanza", "Requerido", "Lista", "Use Y para incluir el registro en el catálogo; N lo conserva como excluido.", "Y", "Sí", "Es la regla explícita de inclusión.", "TBQ"),
    TemplateColumnSpec("selected_pattern", "Patrón Seleccionado (Manual)", "Arquitectura", "Recomendado", "Lista", "Patrón gobernado propuesto por el arquitecto.", "#02", "Event Driven", "Activa QA, narrativas y agrupación.", "PATTERNS"),
    TemplateColumnSpec("pattern_rationale", "Racional del Patrón (Manual)", "Arquitectura", "Recomendado", "Texto", "Explica por qué el patrón satisface el escenario.", "Desacopla productor y tres consumidores.", "Es mejor", "Permite revisión y aprobación trazable."),
    TemplateColumnSpec("comments", "Comentarios / Observaciones", "Gobernanza", "Opcional", "Texto", "Restricciones, decisiones o información pendiente.", "Confirmar retención con Security.", "N/A", "Preserva incertidumbre y decisiones."),
    TemplateColumnSpec("retry_policy", "Retry Policy", "Arquitectura", "Recomendado", "Texto", "Política de reintentos, backoff y manejo final de error.", "3 intentos; backoff 1m/5m/15m; DLQ", "Automático", "Permite revisar resiliencia."),
    TemplateColumnSpec("idempotency", "Idempotencia", "Arquitectura", "Recomendado", "Texto", "Mecanismo que evita aplicar dos veces el mismo evento o comando.", "Clave orderId; conservar 7 días", "Sí", "Es requisito para retries, eventos, correlación y replay."),
    TemplateColumnSpec("core_tools", "Herramientas Core Cuantificables / Volumétricas", "Arquitectura", "Recomendado", "Texto", "Servicios que procesan la carga; separe múltiples valores con |.", "OCI Streaming | OIC Gen3", "Oracle", "Define la ruta que se dimensiona."),
    TemplateColumnSpec("additional_tools_overlays", "Herramientas Adicionales / Overlays (Complemento Manual)", "Arquitectura", "Opcional", "Texto", "Servicios de seguridad, observabilidad o contexto; separe con |.", "OCI API Gateway | OCI APM", "Monitoring", "Documenta overlays sin inflar volumetría."),
    TemplateColumnSpec("uncertainty", "Incertidumbre", "Gobernanza", "Recomendado", "Texto", "Dato faltante, hipótesis o nivel de confianza que aún debe validarse.", "Payload estimado; validar con ERP Team.", "Ninguna", "Evita convertir supuestos en hechos."),
    TemplateColumnSpec("identified_in", "Identificada en:", "Linaje", "Opcional", "Texto", "Documento, sesión o fuente donde se identificó.", "Workshop Finance 2026-07-10", "Reunión", "Refuerza linaje de captura."),
    TemplateColumnSpec("business_process_dd", "Proceso de Negocio DueDiligence", "Linaje", "Opcional", "Texto", "Proceso usado en la evidencia de due diligence cuando difiere del gobernado.", "Procure to Pay", "P2P?", "Preserva contexto fuente."),
    TemplateColumnSpec("slide", "Slide", "Linaje", "Opcional", "Texto", "Página o diapositiva de la evidencia fuente.", "Slide 14", "Presentación", "Permite volver a la evidencia original."),
)


@dataclass(frozen=True)
class TemplateContext:
    """Governed rows required to assemble one workbook snapshot."""

    dictionaries: dict[str, list[DictionaryOption]]
    patterns: list[PatternDefinition]
    profiles: list[ServiceCapabilityProfile]
    versions: dict[str, ServiceProductVersion]
    limits: list[ServiceLimit]
    rules: list[ServiceInteroperabilityRule]
    evidence: list[ServiceEvidenceSource]


NAVY = "1F2937"
BLUE = "2563EB"
TEAL = "0F766E"
AMBER = "D97706"
RED = "B91C1C"
LIGHT_BLUE = "DBEAFE"
LIGHT_TEAL = "D1FAE5"
LIGHT_AMBER = "FEF3C7"
LIGHT_GRAY = "F3F4F6"
MID_GRAY = "D1D5DB"
TEXT_GRAY = "4B5563"
WHITE = "FFFFFF"
THIN_GRAY = Side(style="thin", color=MID_GRAY)


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _safe_text(value: object | None) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        text = json.dumps(value, ensure_ascii=False, sort_keys=True)
    else:
        text = str(value)
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _readable_value(value: object | None) -> str:
    """Render governed JSON as compact human-readable workbook text."""

    if value is None:
        return ""
    if isinstance(value, dict):
        return "\n".join(
            f"• {str(key).replace('_', ' ').title()}: {_readable_value(item).removeprefix('• ')}"
            for key, item in value.items()
        )
    if isinstance(value, list):
        return "\n".join(f"• {_readable_value(item).removeprefix('• ')}" for item in value)
    return _safe_text(value)


def _lines(value: str | None) -> list[str]:
    if not value:
        return []
    return [line.lstrip("-• 0123456789.").strip() for line in value.splitlines() if line.strip()]


def _fallback_examples(pattern: PatternDefinition) -> list[str]:
    return _lines(pattern.when_to_use)[:3] or [f"Escenario candidato para {pattern.name}; validar con el arquitecto."]


def _fallback_questions(pattern: PatternDefinition) -> list[str]:
    first_use = (_lines(pattern.when_to_use) or ["el escenario coincide con la intención del patrón"])[0]
    return [
        f"¿Se confirmó que {first_use[:180].lower()}?",
        "¿Se documentaron fallos, reintentos, idempotencia y límites relevantes?",
        "¿Los componentes seleccionados coinciden con la ruta diseñada en Canvas?",
    ]


def _fallback_required_inputs(pattern: PatternDefinition) -> list[str]:
    category = pattern.category.upper()
    inputs = ["Origen y destino", "Frecuencia", "Payload por ejecución", "Tipo de trigger"]
    if "ASYN" in category or "DAT" in category:
        inputs.extend(["Retención o ventana de procesamiento", "Idempotencia y retry/DLQ"])
    if pattern.pattern_id in {"#02", "#07", "#17"}:
        inputs.append("Fan-out y número de destinos")
    return inputs


async def _load_context(db: AsyncSession) -> TemplateContext:
    dictionary_rows = list(
        (
            await db.scalars(
                select(DictionaryOption)
                .where(DictionaryOption.is_active.is_(True))
                .order_by(DictionaryOption.category, DictionaryOption.sort_order, DictionaryOption.value)
            )
        ).all()
    )
    dictionaries: dict[str, list[DictionaryOption]] = {}
    for option in dictionary_rows:
        dictionaries.setdefault(option.category, []).append(option)

    patterns = list(
        (
            await db.scalars(
                select(PatternDefinition)
                .where(PatternDefinition.is_active.is_(True))
                .order_by(PatternDefinition.pattern_id)
            )
        ).all()
    )
    profiles = list(
        (
            await db.scalars(
                select(ServiceCapabilityProfile)
                .where(ServiceCapabilityProfile.is_active.is_(True))
                .order_by(ServiceCapabilityProfile.category, ServiceCapabilityProfile.service_id)
            )
        ).all()
    )
    profile_ids = [profile.id for profile in profiles]
    versions: dict[str, ServiceProductVersion] = {}
    limits: list[ServiceLimit] = []
    rules: list[ServiceInteroperabilityRule] = []
    evidence: list[ServiceEvidenceSource] = []
    if profile_ids:
        version_rows = list(
            (
                await db.scalars(
                    select(ServiceProductVersion)
                    .where(ServiceProductVersion.service_profile_id.in_(profile_ids))
                    .order_by(ServiceProductVersion.created_at.desc())
                )
            ).all()
        )
        for version in version_rows:
            versions.setdefault(version.service_profile_id, version)
        limits = list(
            (
                await db.scalars(
                    select(ServiceLimit)
                    .where(ServiceLimit.service_profile_id.in_(profile_ids), ServiceLimit.is_active.is_(True))
                    .order_by(ServiceLimit.service_profile_id, ServiceLimit.limit_type, ServiceLimit.limit_key)
                )
            ).all()
        )
        rules = list(
            (
                await db.scalars(
                    select(ServiceInteroperabilityRule)
                    .where(ServiceInteroperabilityRule.is_active.is_(True))
                    .order_by(ServiceInteroperabilityRule.source_service_profile_id, ServiceInteroperabilityRule.target_service_profile_id)
                )
            ).all()
        )
        evidence = list(
            (
                await db.scalars(
                    select(ServiceEvidenceSource)
                    .where(ServiceEvidenceSource.service_profile_id.in_(profile_ids))
                    .order_by(ServiceEvidenceSource.service_profile_id, ServiceEvidenceSource.trust_tier, ServiceEvidenceSource.url)
                )
            ).all()
        )
    return TemplateContext(dictionaries, patterns, profiles, versions, limits, rules, evidence)


def _style_title(sheet: Any, title: str, subtitle: str, end_column: int) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    sheet.cell(1, 1, title)
    sheet.cell(1, 1).font = Font(name="Aptos Display", size=20, bold=True, color=WHITE)
    sheet.cell(1, 1).fill = PatternFill("solid", fgColor=NAVY)
    sheet.cell(1, 1).alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 34
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    sheet.cell(2, 1, subtitle)
    sheet.cell(2, 1).font = Font(name="Aptos", size=10, color=TEXT_GRAY)
    sheet.cell(2, 1).fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    sheet.cell(2, 1).alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 30
    sheet.sheet_view.showGridLines = False


def _style_table_header(cells: Iterable[Any], fill: str = NAVY) -> None:
    for cell in cells:
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GRAY)


def _set_widths(sheet: Any, widths: dict[int, float]) -> None:
    for index, width in widths.items():
        sheet.column_dimensions[get_column_letter(index)].width = width


def _add_list_name(workbook: Workbook, name: str, column_letter: str, final_row: int) -> None:
    target = f"'{MANIFEST_SHEET_NAME}'!${column_letter}$2:${column_letter}${max(2, final_row)}"
    workbook.defined_names.add(DefinedName(name, attr_text=target))


def _add_validation(sheet: Any, range_ref: str, formula: str, prompt: str) -> None:
    validation = DataValidation(type="list", formula1=f"={formula}", allow_blank=True)
    validation.error = "Seleccione un valor de la lista gobernada."
    validation.errorTitle = "Valor no gobernado"
    validation.prompt = prompt
    validation.promptTitle = "Ayuda de captura"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    validation.add(range_ref)
    sheet.add_data_validation(validation)


def _add_client_validation(sheet: Any, range_ref: str, formula: str, prompt: str) -> None:
    """Offer customer-maintained suggestions without rejecting a new valid value."""

    validation = DataValidation(type="list", formula1=f"={formula}", allow_blank=True)
    validation.prompt = f"{prompt} Add reusable values in {CLIENT_CATALOGS_SHEET_NAME}."
    validation.promptTitle = "Catálogo del cliente"
    validation.showErrorMessage = False
    validation.showInputMessage = True
    validation.add(range_ref)
    sheet.add_data_validation(validation)


def _add_numeric_validation(sheet: Any, range_ref: str, data_type: str, prompt: str) -> None:
    """Require non-negative numeric capture values while allowing blanks."""

    validation = DataValidation(
        type="whole" if data_type == "Entero" else "decimal",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=True,
    )
    validation.error = "Capture un número igual o mayor que cero, sin escribir la unidad."
    validation.errorTitle = "Valor numérico inválido"
    validation.prompt = prompt
    validation.promptTitle = "Ayuda de captura"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    validation.add(range_ref)
    sheet.add_data_validation(validation)


def _pattern_list_value(pattern: PatternDefinition) -> str:
    return pattern.pattern_id


def _create_lists_sheet(workbook: Workbook, context: TemplateContext, generated_at: datetime) -> dict[str, str]:
    sheet = workbook.create_sheet(MANIFEST_SHEET_NAME)
    lists: dict[str, list[str]] = {
        "FREQUENCY": [option.value for option in context.dictionaries.get("FREQUENCY", [])],
        "TRIGGER_TYPE": [option.value for option in context.dictionaries.get("TRIGGER_TYPE", [])],
        "COMPLEXITY": [option.value for option in context.dictionaries.get("COMPLEXITY", [])],
        "BUSINESS_CRITICALITY": ["Baja", "Media", "Alta", "Crítica"],
        "DATA_CLASSIFICATION": ["Pública", "Interna", "Confidencial", "Restringida"],
        "YES_NO": ["Sí", "No"],
        "TBQ": ["Y", "N"],
        "PATTERNS": [_pattern_list_value(pattern) for pattern in context.patterns],
    }
    name_map: dict[str, str] = {}
    for column_index, (key, values) in enumerate(lists.items(), start=1):
        defined_name = f"LIST_{key}"
        name_map[key] = defined_name
        sheet.cell(1, column_index, key)
        for row_index, value in enumerate(values, start=2):
            sheet.cell(row_index, column_index, _safe_text(value))
        _add_list_name(workbook, defined_name, get_column_letter(column_index), len(values) + 1)

    manifest_column = len(lists) + 2
    manifest = [
        ("template_version", TEMPLATE_VERSION),
        ("importer_min_version", IMPORTER_MIN_VERSION),
        ("generated_at_utc", generated_at.isoformat()),
        ("capture_sheet", CAPTURE_SHEET_NAME),
        ("capture_row_limit", str(CAPTURE_ROW_LIMIT - 1)),
        ("pattern_count", str(len(context.patterns))),
        ("service_product_count", str(len(context.profiles))),
        ("service_limit_count", str(len(context.limits))),
        ("interoperability_rule_count", str(len(context.rules))),
        ("evidence_source_count", str(len(context.evidence))),
    ]
    sheet.cell(1, manifest_column, "MANIFEST_KEY")
    sheet.cell(1, manifest_column + 1, "MANIFEST_VALUE")
    for row_index, (key, value) in enumerate(manifest, start=2):
        sheet.cell(row_index, manifest_column, key)
        sheet.cell(row_index, manifest_column + 1, value)
    sheet.sheet_state = "veryHidden"
    return name_map


def _create_client_catalogs_sheet(workbook: Workbook) -> dict[str, str]:
    """Create editable customer suggestions used by the capture dropdowns."""

    sheet = workbook.create_sheet(CLIENT_CATALOGS_SHEET_NAME)
    _style_title(
        sheet,
        "Catálogos del Cliente",
        "Reemplace o amplíe estas sugerencias. La captura también acepta valores nuevos; esta hoja sólo mejora consistencia y velocidad.",
        5,
    )
    catalogs = {
        "CLIENT_BRANDS": ("Marcas / unidades", ["Retail", "Finance", "Human Capital", "Supply Chain"]),
        "CLIENT_PROCESSES": ("Procesos de negocio", ["Order to Cash", "Procure to Pay", "Hire to Retire", "Record to Report", "Inventory Synchronization"]),
        "CLIENT_SYSTEMS": ("Sistemas", ["Oracle ERP Cloud", "Oracle HCM Cloud", "Salesforce", "SAP S/4HANA", "Custom Application"]),
        "CLIENT_TECHNOLOGIES": ("Tecnologías", ["REST API", "SOAP API", "SFTP", "JDBC", "Kafka", "Oracle ATP"]),
        "CLIENT_OWNERS": ("Owners / equipos", ["Enterprise Architecture", "Integration Platform Team", "Application Owner", "Data Platform Team"]),
    }
    name_map: dict[str, str] = {}
    for column, (defined_name, (header, values)) in enumerate(catalogs.items(), start=1):
        sheet.cell(4, column, header)
        for row, value in enumerate(values, start=5):
            sheet.cell(row, column, value)
        target = f"'{CLIENT_CATALOGS_SHEET_NAME}'!${get_column_letter(column)}$5:${get_column_letter(column)}$104"
        workbook.defined_names.add(DefinedName(defined_name, attr_text=target))
        name_map[defined_name] = defined_name
    _style_table_header(sheet[4], fill=TEAL)
    _set_widths(sheet, {1: 24, 2: 30, 3: 30, 4: 28, 5: 28})
    sheet.freeze_panes = "A5"
    sheet.sheet_view.showGridLines = False
    return name_map


def _create_start_sheet(workbook: Workbook, context: TemplateContext, generated_at: datetime) -> None:
    sheet = workbook.active
    sheet.title = "Inicio"
    _style_title(
        sheet,
        "OCI DIS Architect · Captura Offline",
        "Plantilla gobernada para documentar integraciones antes de importarlas a la App.",
        8,
    )
    sheet["A4"] = "Qué es este archivo"
    sheet["A4"].font = Font(size=14, bold=True, color=NAVY)
    sheet.merge_cells("A5:H6")
    sheet["A5"] = (
        "Este workbook permite capturar un inventario de integraciones sin conexión. La App usa cada fila "
        "para construir el catálogo, la topología, el QA, la volumetría y la revisión de arquitectura. "
        "No necesita conocer OCI para comenzar: use las listas, ejemplos y guías incluidas."
    )
    sheet["A5"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet["A5"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    steps = [
        ("1", "Lea Inicio", "Revise reglas, colores y el flujo de trabajo."),
        ("2", "Capture", "Agregue una integración por fila en Catálogo de Integraciones."),
        ("3", "Valide", "Revise Validación Previa; corrija filas No importable."),
        ("4", "Consulte", "Use Guía de Campos, Patrones y Servicios OCI cuando tenga dudas."),
        ("5", "Importe", "Suba este mismo .xlsx en Projects > Import de OCI DIS Architect."),
    ]
    sheet["A8"] = "Flujo recomendado"
    sheet["A8"].font = Font(size=14, bold=True, color=NAVY)
    for row, (number, title, detail) in enumerate(steps, start=9):
        sheet.cell(row, 1, number)
        sheet.cell(row, 1).font = Font(size=12, bold=True, color=WHITE)
        sheet.cell(row, 1).fill = PatternFill("solid", fgColor=BLUE)
        sheet.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        sheet.cell(row, 2, title)
        sheet.cell(row, 2).font = Font(bold=True, color=NAVY)
        sheet.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
        sheet.cell(row, 4, detail)
        sheet.cell(row, 4).alignment = Alignment(wrap_text=True)
        sheet.row_dimensions[row].height = 28

    sheet["A16"] = "Reglas que evitan errores"
    sheet["A16"].font = Font(size=14, bold=True, color=NAVY)
    rules = [
        "No renombre la hoja Catálogo de Integraciones ni sus encabezados.",
        "No pegue fórmulas en la hoja de captura; capture valores. La App rechazará fórmulas por seguridad y trazabilidad.",
        "Use TBQ = Y solo para filas que deben entrar al catálogo. TBQ = N conserva la fila como evidencia excluida.",
        "Deje una celda vacía cuando no tenga información. Use 0 únicamente cuando el valor real sea cero.",
        "No invente datos para completar el archivo. Documente lo desconocido en Incertidumbre.",
        "Los ejemplos viven en Ejemplos Guiados y nunca se importan.",
    ]
    for row, rule in enumerate(rules, start=17):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        sheet.cell(row, 1, f"• {rule}")
        sheet.cell(row, 1).alignment = Alignment(wrap_text=True)
        sheet.cell(row, 1).fill = PatternFill("solid", fgColor=LIGHT_AMBER if row < 20 else LIGHT_GRAY)
        sheet.row_dimensions[row].height = 26

    sheet["A25"] = "Leyenda de captura"
    sheet["A25"].font = Font(size=14, bold=True, color=NAVY)
    legend = [
        ("Requerido", BLUE, f"Sin este dato la fila no es importable desde la plantilla v{TEMPLATE_VERSION.split('.')[0]}."),
        ("Recomendado", TEAL, "La fila puede importar, pero QA, sizing o revisión tendrán menor confianza."),
        ("Opcional", "6B7280", "Contexto útil para linaje, ownership y operación."),
    ]
    for row, (label, color, detail) in enumerate(legend, start=26):
        sheet.cell(row, 1, label)
        sheet.cell(row, 1).fill = PatternFill("solid", fgColor=color)
        sheet.cell(row, 1).font = Font(bold=True, color=WHITE)
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        sheet.cell(row, 2, detail)

    sheet.merge_cells("A31:H32")
    sheet["A31"] = (
        f"Plantilla v{TEMPLATE_VERSION} · generada {generated_at:%Y-%m-%d %H:%M UTC} · "
        f"{len(context.patterns)} patrones · {len(context.profiles)} servicios OCI · "
        f"{len(context.limits)} límites · {len(context.rules)} reglas de interoperabilidad."
    )
    sheet["A31"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet["A31"].font = Font(color=TEXT_GRAY, italic=True)
    sheet["A31"].fill = PatternFill("solid", fgColor=LIGHT_TEAL)
    sheet["A34"] = "Navegación rápida"
    sheet["A34"].font = Font(size=14, bold=True, color=NAVY)
    navigation = [
        ("A35:B35", "Dashboard", "Dashboard"),
        ("C35:D35", "Capturar integraciones", CAPTURE_SHEET_NAME),
        ("E35:F35", "Validación previa", "Validación Previa"),
        ("G35:H35", "Catálogos del cliente", CLIENT_CATALOGS_SHEET_NAME),
        ("A36:B36", "Guía de campos", "Guía de Campos"),
        ("C36:D36", "Patrones", "Patrones"),
        ("E36:F36", "Servicios OCI", "Servicios OCI"),
        ("G36:H36", "Interoperabilidad", "Interoperabilidad"),
    ]
    for cell_range, label, target_sheet in navigation:
        sheet.merge_cells(cell_range)
        cell = sheet[cell_range.split(":", 1)[0]]
        cell.value = label
        cell.hyperlink = f"#'{target_sheet}'!A1"
        cell.font = Font(bold=True, color=BLUE, underline="single")
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    _set_widths(sheet, {1: 16, 2: 18, 3: 18, 4: 20, 5: 20, 6: 20, 7: 20, 8: 20})
    sheet.freeze_panes = "A4"


def _create_capture_sheet(workbook: Workbook, list_names: dict[str, str]) -> None:
    sheet = workbook.create_sheet(CAPTURE_SHEET_NAME)
    for index, column in enumerate(COLUMNS, start=1):
        cell = sheet.cell(1, index, column.header)
        fill = BLUE if column.requirement == "Requerido" else TEAL if column.requirement == "Recomendado" else "6B7280"
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.comment = Comment(
            f"{column.requirement}. {column.description}\nEjemplo correcto: {column.good_example}\nLa App lo usa para: {column.app_usage}",
            "OCI DIS Architect",
        )
    sheet.row_dimensions[1].height = 48
    sheet.freeze_panes = "F2"
    last_letter = get_column_letter(len(COLUMNS))
    sheet.auto_filter.ref = f"A1:{last_letter}{CAPTURE_ROW_LIMIT}"
    sheet.sheet_view.zoomScale = 75
    sheet.sheet_view.showGridLines = False

    for index, column in enumerate(COLUMNS, start=1):
        width = 14
        if column.data_type == "Texto":
            width = 24
        if column.header in {"Descripción", "Racional del Patrón (Manual)", "Comentarios / Observaciones", "Incertidumbre"}:
            width = 34
        sheet.column_dimensions[get_column_letter(index)].width = width
        if column.validation_key and column.validation_key in list_names:
            validation_writer = _add_client_validation if column.validation_key.startswith("CLIENT_") else _add_validation
            validation_writer(
                sheet,
                f"{get_column_letter(index)}2:{get_column_letter(index)}{CAPTURE_ROW_LIMIT}",
                list_names[column.validation_key],
                column.description,
            )
        elif column.data_type in {"Entero", "Decimal"}:
            _add_numeric_validation(
                sheet,
                f"{get_column_letter(index)}2:{get_column_letter(index)}{CAPTURE_ROW_LIMIT}",
                column.data_type,
                column.description,
            )
    # Blank rows remain truly blank; only conditional formatting communicates missing required input.
    required_letters = [get_column_letter(index) for index, column in enumerate(COLUMNS, start=1) if column.requirement == "Requerido"]
    data_range = f"A2:{last_letter}{CAPTURE_ROW_LIMIT}"
    first_required = required_letters[0]
    for letter in required_letters:
        sheet.conditional_formatting.add(
            data_range,
            FormulaRule(
                formula=[f'AND(COUNTA($A2:${last_letter}2)>0,${letter}2="")'],
                fill=PatternFill("solid", fgColor="FEE2E2"),
            ),
        )
    sheet.conditional_formatting.add(
        data_range,
        FormulaRule(
            formula=[f'AND(COUNTA($A2:${last_letter}2)>0,${first_required}2<>"")'],
            border=Border(bottom=Side(style="hair", color=MID_GRAY)),
        ),
    )


def _create_preflight_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Validación Previa")
    _style_title(sheet, "Validación Previa", "Revisión orientativa antes de importar. La QA final y autoritativa se ejecuta dentro de la App.", 7)
    headers = ["Fila", "ID", "Interfaz", "Estado", "Campos faltantes", "Observaciones", "Acción"]
    for column, header in enumerate(headers, start=1):
        sheet.cell(4, column, header)
    _style_table_header(sheet[4])
    sheet["I4"] = "Resumen"
    sheet["J4"] = "Valor"
    _style_table_header(sheet[4][8:10], fill=TEAL)
    summary_rows = [
        ("Filas capturadas", f'=COUNTIF(D5:D{CAPTURE_ROW_LIMIT + 3},"<>Sin captura")'),
        ("Listas para importar", f'=COUNTIF(D5:D{CAPTURE_ROW_LIMIT + 3},"Listo")'),
        ("No importables", f'=COUNTIF(D5:D{CAPTURE_ROW_LIMIT + 3},"No importable")'),
        ("Con observaciones", f'=COUNTIF(F5:F{CAPTURE_ROW_LIMIT + 3},"?*")'),
        ("Estado general", '=IF(J7>0,"CORREGIR",IF(J5=0,"SIN CAPTURA","LISTO"))'),
    ]
    for row, (label, formula) in enumerate(summary_rows, start=5):
        sheet.cell(row, 9, label)
        sheet.cell(row, 10, formula)
        sheet.cell(row, 10).data_type = "f"
    field_to_letter = {column.field: get_column_letter(index) for index, column in enumerate(COLUMNS, start=1)}
    last_letter = get_column_letter(len(COLUMNS))
    required = [column for column in COLUMNS if column.requirement == "Requerido"]
    for target_row, capture_row in enumerate(range(2, CAPTURE_ROW_LIMIT + 1), start=5):
        blank_check = f"COUNTA('{CAPTURE_SHEET_NAME}'!A{capture_row}:{last_letter}{capture_row})=0"
        sheet.cell(target_row, 1, capture_row)
        sheet.cell(
            target_row,
            2,
            f'=IF({blank_check},"—",\'{CAPTURE_SHEET_NAME}\'!{field_to_letter["interface_id"]}{capture_row})',
        )
        sheet.cell(target_row, 2).data_type = "f"
        sheet.cell(
            target_row,
            3,
            f'=IF({blank_check},"—",\'{CAPTURE_SHEET_NAME}\'!{field_to_letter["interface_name"]}{capture_row})',
        )
        sheet.cell(target_row, 3).data_type = "f"
        missing_terms = [
            f'IF(\'{CAPTURE_SHEET_NAME}\'!{field_to_letter[column.field]}{capture_row}="","{column.header}; ","")'
            for column in required
        ]
        missing_formula = "&".join(missing_terms)
        sheet.cell(
            target_row,
            4,
            f'=IF({blank_check},"Sin captura",IF({missing_formula}="","Listo","No importable"))',
        )
        sheet.cell(target_row, 4).data_type = "f"
        sheet.cell(target_row, 5, f'=IF({blank_check},"",{missing_formula})')
        sheet.cell(target_row, 5).data_type = "f"
        fanout_letter = field_to_letter["is_fan_out"]
        targets_letter = field_to_letter["fan_out_targets"]
        pattern_letter = field_to_letter["selected_pattern"]
        trigger_letter = field_to_letter["trigger_type"]
        interface_id_letter = field_to_letter["interface_id"]
        payload_letter = field_to_letter["payload_per_execution_kb"]
        frequency_letter = field_to_letter["frequency"]
        tbq_letter = field_to_letter["tbq"]
        sheet.cell(
            target_row,
            6,
            (
                f'=IF({blank_check},"",IF(AND(\'{CAPTURE_SHEET_NAME}\'!{fanout_letter}{capture_row}="Sí",'
                f'\'{CAPTURE_SHEET_NAME}\'!{targets_letter}{capture_row}<2),"Fan-out requiere al menos 2 destinos; ","")&'
                f'IF(\'{CAPTURE_SHEET_NAME}\'!{pattern_letter}{capture_row}="","Patrón no informado; ","")&'
                f'IF(\'{CAPTURE_SHEET_NAME}\'!{trigger_letter}{capture_row}="","Trigger no informado; ","")&'
                f'IF(AND(\'{CAPTURE_SHEET_NAME}\'!{interface_id_letter}{capture_row}<>"",COUNTIF(\'{CAPTURE_SHEET_NAME}\'!${interface_id_letter}$2:${interface_id_letter}${CAPTURE_ROW_LIMIT},\'{CAPTURE_SHEET_NAME}\'!{interface_id_letter}{capture_row})>1),"ID duplicado; ","")&'
                f'IF(AND(\'{CAPTURE_SHEET_NAME}\'!{payload_letter}{capture_row}<>"",NOT(ISNUMBER(\'{CAPTURE_SHEET_NAME}\'!{payload_letter}{capture_row}))),"Payload debe ser numérico; ","")&'
                f'IF(COUNTIF(LIST_FREQUENCY,\'{CAPTURE_SHEET_NAME}\'!{frequency_letter}{capture_row})=0,"Frecuencia fuera de catálogo; ","")&'
                f'IF(AND(\'{CAPTURE_SHEET_NAME}\'!{pattern_letter}{capture_row}<>"",COUNTIF(LIST_PATTERNS,\'{CAPTURE_SHEET_NAME}\'!{pattern_letter}{capture_row})=0),"Patrón fuera de catálogo; ","")&'
                f'IF(COUNTIF(LIST_TBQ,\'{CAPTURE_SHEET_NAME}\'!{tbq_letter}{capture_row})=0,"TBQ debe ser Y o N; ",""))'
            ),
        )
        sheet.cell(target_row, 6).data_type = "f"
        sheet.cell(
            target_row,
            7,
            f'=IF(D{target_row}="Sin captura","Sin acción",IF(D{target_row}="No importable","Corregir captura",IF(F{target_row}<>"","Revisar antes de importar","Importar")))',
        )
        sheet.cell(target_row, 7).data_type = "f"
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:G{CAPTURE_ROW_LIMIT + 3}"
    sheet.conditional_formatting.add(f"D5:D{CAPTURE_ROW_LIMIT + 3}", FormulaRule(formula=['D5="Listo"'], fill=PatternFill("solid", fgColor=LIGHT_TEAL)))
    sheet.conditional_formatting.add(f"D5:D{CAPTURE_ROW_LIMIT + 3}", FormulaRule(formula=['D5="No importable"'], fill=PatternFill("solid", fgColor="FEE2E2")))
    sheet.conditional_formatting.add(f"D5:D{CAPTURE_ROW_LIMIT + 3}", FormulaRule(formula=['D5="Sin captura"'], fill=PatternFill("solid", fgColor=LIGHT_GRAY)))
    _set_widths(sheet, {1: 8, 2: 16, 3: 30, 4: 18, 5: 42, 6: 52, 7: 24, 8: 3, 9: 24, 10: 18})
    sheet.sheet_view.zoomScale = 85


def _create_dashboard_sheet(workbook: Workbook) -> None:
    """Add a formula-driven offline summary without becoming an authority source."""

    sheet = workbook.create_sheet("Dashboard")
    _style_title(
        sheet,
        "Dashboard de Captura Offline",
        "Resumen orientativo. La App recalcula QA, volumetría, topología y arquitectura después de importar.",
        12,
    )
    fields = {column.field: get_column_letter(index) for index, column in enumerate(COLUMNS, start=1)}
    capture = f"'{CAPTURE_SHEET_NAME}'"
    metrics = [
        ("Filas capturadas", f'=COUNTIF({capture}!{fields["interface_name"]}2:{fields["interface_name"]}{CAPTURE_ROW_LIMIT},"?*")'),
        ("Incluidas (TBQ=Y)", f'=COUNTIF({capture}!{fields["tbq"]}2:{fields["tbq"]}{CAPTURE_ROW_LIMIT},"Y")'),
        ("Con patrón", f'=COUNTIF({capture}!{fields["selected_pattern"]}2:{fields["selected_pattern"]}{CAPTURE_ROW_LIMIT},"#??")'),
        ("Con payload", f'=COUNT({capture}!{fields["payload_per_execution_kb"]}2:{fields["payload_per_execution_kb"]}{CAPTURE_ROW_LIMIT})'),
    ]
    for index, (label, formula) in enumerate(metrics):
        start = 1 + index * 3
        sheet.merge_cells(start_row=4, start_column=start, end_row=4, end_column=start + 1)
        sheet.cell(4, start, label)
        sheet.cell(4, start).font = Font(bold=True, color=TEXT_GRAY)
        sheet.merge_cells(start_row=5, start_column=start, end_row=6, end_column=start + 1)
        sheet.cell(5, start, formula)
        sheet.cell(5, start).data_type = "f"
        sheet.cell(5, start).font = Font(size=24, bold=True, color=NAVY)
        sheet.cell(5, start).fill = PatternFill("solid", fgColor=LIGHT_BLUE if index < 2 else LIGHT_TEAL)
        sheet.cell(5, start).alignment = Alignment(horizontal="center", vertical="center")

    sheet["A9"] = "Cobertura de criticidad"
    sheet["A9"].font = Font(size=14, bold=True, color=NAVY)
    sheet["A10"] = "Nivel"
    sheet["B10"] = "Integraciones"
    _style_table_header(sheet[10][0:2], fill=TEAL)
    for row, level in enumerate(("Crítica", "Alta", "Media", "Baja"), start=11):
        sheet.cell(row, 1, level)
        sheet.cell(
            row,
            2,
            f'=COUNTIF({capture}!{fields["business_criticality"]}2:{fields["business_criticality"]}{CAPTURE_ROW_LIMIT},A{row})',
        )
        sheet.cell(row, 2).data_type = "f"

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Integraciones por criticidad"
    chart.legend = None
    chart.y_axis.title = "Nivel"
    chart.x_axis.title = "Integraciones"
    chart.height = 7
    chart.width = 13
    chart.add_data(Reference(sheet, min_col=2, min_row=10, max_row=14), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=11, max_row=14))
    sheet.add_chart(chart, "D9")

    sheet["A18"] = "Decisiones pendientes"
    sheet["A18"].font = Font(size=14, bold=True, color=NAVY)
    pending = [
        ("Sin patrón", f'=COUNTIFS({capture}!{fields["interface_name"]}2:{fields["interface_name"]}{CAPTURE_ROW_LIMIT},"?*",{capture}!{fields["selected_pattern"]}2:{fields["selected_pattern"]}{CAPTURE_ROW_LIMIT},"")'),
        ("Sin SLA / latencia", f'=COUNTIFS({capture}!{fields["interface_name"]}2:{fields["interface_name"]}{CAPTURE_ROW_LIMIT},"?*",{capture}!{fields["target_latency_sla"]}2:{fields["target_latency_sla"]}{CAPTURE_ROW_LIMIT},"")'),
        ("Sin clasificación", f'=COUNTIFS({capture}!{fields["interface_name"]}2:{fields["interface_name"]}{CAPTURE_ROW_LIMIT},"?*",{capture}!{fields["data_security_classification"]}2:{fields["data_security_classification"]}{CAPTURE_ROW_LIMIT},"")'),
        ("Sin idempotencia", f'=COUNTIFS({capture}!{fields["interface_name"]}2:{fields["interface_name"]}{CAPTURE_ROW_LIMIT},"?*",{capture}!{fields["idempotency"]}2:{fields["idempotency"]}{CAPTURE_ROW_LIMIT},"")'),
    ]
    for row, (label, formula) in enumerate(pending, start=19):
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, formula)
        sheet.cell(row, 2).data_type = "f"
    _set_widths(sheet, {1: 24, 2: 18, 3: 3, 4: 18, 5: 18, 6: 18, 7: 18, 8: 18, 9: 18, 10: 18, 11: 18, 12: 18})
    sheet.sheet_view.showGridLines = False


def _create_examples_sheet(workbook: Workbook, context: TemplateContext) -> None:
    sheet = workbook.create_sheet("Ejemplos Guiados")
    _style_title(sheet, "Ejemplos Guiados", "Estos ejemplos son educativos y nunca son leídos por el importador.", 9)
    headers = ["Patrón", "Escenario aplicable", "Origen", "Destino", "Trigger", "Frecuencia", "Core tools", "Inputs mínimos", "Evite"]
    for column, header in enumerate(headers, start=1):
        sheet.cell(4, column, header)
    _style_table_header(sheet[4])
    for row, pattern in enumerate(context.patterns, start=5):
        examples = pattern.applicability_examples or _fallback_examples(pattern)
        required_inputs = pattern.required_inputs or _fallback_required_inputs(pattern)
        sheet.append(
            [
                f"{pattern.pattern_id} · {pattern.name}",
                _safe_text(examples[0] if examples else pattern.description),
                "Sistema productor",
                "Sistema consumidor",
                "Event Trigger" if "ASYN" in pattern.category.upper() else "REST Trigger",
                "Tiempo Real" if "ASYN" in pattern.category.upper() else "Bajo demanda",
                _safe_text(pattern.oci_components),
                _safe_text("; ".join(required_inputs)),
                _safe_text(pattern.when_not_to_use),
            ]
        )
        for cell in sheet[row]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.row_dimensions[row].height = 105
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:I{max(4, sheet.max_row)}"
    _set_widths(sheet, {1: 28, 2: 42, 3: 22, 4: 22, 5: 18, 6: 18, 7: 42, 8: 42, 9: 54})


def _create_field_guide_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Guía de Campos")
    _style_title(sheet, "Guía de Campos", "Definiciones en lenguaje simple y consecuencias de cada dato.", 11)
    headers = ["Sección", "Encabezado", "Campo canónico", "Nivel", "Tipo", "Unidad", "Qué significa", "Ejemplo correcto", "Ejemplo incorrecto", "Uso en la App", "Si falta"]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(4, column_index, header)
    _style_table_header(sheet[4])
    for row, spec in enumerate(COLUMNS, start=5):
        consequence = (
            "La fila no es importable desde la plantilla v2."
            if spec.requirement == "Requerido"
            else "Reduce confianza de QA o sizing." if spec.requirement == "Recomendado" else "Se pierde contexto, pero no bloquea."
        )
        sheet.append([
            spec.section,
            spec.header,
            spec.field,
            spec.requirement,
            spec.data_type,
            spec.unit or "",
            spec.description,
            spec.good_example,
            spec.bad_example,
            spec.app_usage,
            consequence,
        ])
        for cell in sheet[row]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.row_dimensions[row].height = 50
    sheet.freeze_panes = "D5"
    sheet.auto_filter.ref = f"A4:K{sheet.max_row}"
    _set_widths(sheet, {1: 16, 2: 32, 3: 28, 4: 16, 5: 14, 6: 10, 7: 42, 8: 28, 9: 28, 10: 40, 11: 34})


def _create_patterns_sheet(workbook: Workbook, context: TemplateContext) -> None:
    sheet = workbook.create_sheet("Patrones")
    _style_title(
        sheet,
        "Patrones de Integración Certificados",
        "Cada patrón incluye un contrato versionado de evidencia, composición, sizing y validación; la integración concreta debe cumplirlo antes de quedar lista.",
        22,
    )
    headers = [
        "ID",
        "Nombre",
        "Categoría",
        "Certificación",
        "Versión certificación",
        "Estrategia de sizing",
        "Evidencia requerida",
        "Composiciones core certificadas",
        "Overlays requeridos",
        "Servicios comerciales",
        "Dependencias externas",
        "Controles de validación",
        "Descripción simple",
        "Cuándo usar",
        "Cuándo no usar",
        "Flujo técnico",
        "Componentes OCI",
        "Valor de negocio",
        "Ejemplos de aplicabilidad",
        "Preguntas de selección",
        "Inputs requeridos",
        "Versión biblioteca",
    ]
    for column, header in enumerate(headers, start=1):
        sheet.cell(4, column, header)
    _style_table_header(sheet[4])
    for row, pattern in enumerate(context.patterns, start=5):
        support = get_pattern_support(pattern.pattern_id)
        sheet.append([
            pattern.pattern_id,
            pattern.name,
            pattern.category,
            support.badge_label,
            support.certification_version,
            (support.sizing_strategy or "").replace("_", " "),
            "\n".join(item.replace("_", " ") for item in support.required_evidence),
            "\n".join(" + ".join(group) for group in support.approved_core_tool_groups),
            "\n".join(" + ".join(group) for group in support.approved_overlay_groups),
            "\n".join(support.commercial_service_ids),
            "\n".join(support.external_dependencies),
            "\n".join(support.validation_controls),
            _safe_text(pattern.description),
            _safe_text(pattern.when_to_use),
            _safe_text(pattern.when_not_to_use),
            _safe_text(pattern.technical_flow),
            _safe_text(pattern.oci_components),
            _safe_text(pattern.business_value),
            "\n".join(pattern.applicability_examples or _fallback_examples(pattern)),
            "\n".join(pattern.selection_questions or _fallback_questions(pattern)),
            "\n".join(pattern.required_inputs or _fallback_required_inputs(pattern)),
            pattern.version,
        ])
        for cell in sheet[row]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.row_dimensions[row].height = 180
    sheet.freeze_panes = "G5"
    sheet.auto_filter.ref = f"A4:V{sheet.max_row}"
    _set_widths(
        sheet,
        {
            1: 8,
            2: 26,
            3: 22,
            4: 18,
            5: 16,
            6: 22,
            7: 30,
            8: 34,
            9: 32,
            10: 28,
            11: 30,
            12: 34,
            13: 38,
            14: 44,
            15: 54,
            16: 54,
            17: 44,
            18: 44,
            19: 42,
            20: 42,
            21: 36,
            22: 14,
        },
    )


def _latest_evidence_by_profile(context: TemplateContext) -> dict[str, ServiceEvidenceSource]:
    result: dict[str, ServiceEvidenceSource] = {}
    for source in context.evidence:
        existing = result.get(source.service_profile_id)
        existing_date = existing.last_checked_at if existing else None
        if existing is None or (source.last_checked_at and (existing_date is None or source.last_checked_at > existing_date)):
            result[source.service_profile_id] = source
    return result


def _create_services_sheet(workbook: Workbook, context: TemplateContext) -> None:
    sheet = workbook.create_sheet("Servicios OCI")
    _style_title(sheet, "Biblioteca de Productos de Servicio", "Productos de Data Integration Services y servicios OCI relacionados, desde la fuente normalizada de la App.", 15)
    headers = ["Service ID", "Producto", "Categoría", "Rol arquitectónico", "Descripción", "Capacidades", "Casos de uso", "Anti-patrones", "Seguridad", "Disponibilidad regional", "Notas comerciales", "Deprecación", "Estado verificación", "Última verificación", "Fuente oficial"]
    for column, header in enumerate(headers, start=1):
        sheet.cell(4, column, header)
    _style_table_header(sheet[4])
    latest_evidence = _latest_evidence_by_profile(context)
    for row, profile in enumerate(context.profiles, start=5):
        version = context.versions.get(profile.id)
        evidence = latest_evidence.get(profile.id)
        sheet.append([
            profile.service_id,
            profile.name,
            profile.category,
            _readable_value((version.product_metadata if version else {}).get("architecture_role") or profile.architectural_fit),
            _safe_text(version.description if version else profile.architectural_fit),
            _readable_value(version.capabilities if version else profile.limits),
            _readable_value(version.use_cases if version else []),
            _readable_value(version.anti_patterns if version else profile.anti_patterns),
            _safe_text(version.security_notes if version else None),
            _safe_text(version.regional_availability if version else None),
            _safe_text(version.commercial_notes if version else profile.pricing_model),
            _safe_text(version.deprecation_notes if version else None),
            evidence.status if evidence else "Sin fuentes",
            evidence.last_checked_at.isoformat() if evidence and evidence.last_checked_at else "",
            evidence.url if evidence else _safe_text(profile.oracle_docs_urls),
        ])
        for cell in sheet[row]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.row_dimensions[row].height = 180
    sheet.freeze_panes = "D5"
    sheet.auto_filter.ref = f"A4:O{sheet.max_row}"
    _set_widths(sheet, {1: 16, 2: 30, 3: 22, 4: 36, 5: 46, 6: 42, 7: 42, 8: 42, 9: 38, 10: 30, 11: 38, 12: 36, 13: 20, 14: 24, 15: 58})


def _create_limits_sheet(workbook: Workbook, context: TemplateContext) -> None:
    sheet = workbook.create_sheet("Límites OCI")
    _style_title(sheet, "Límites OCI", "Restricciones normalizadas; valide alcance, región y posibilidad de aumento antes de diseñar producción.", 14)
    profile_map = {profile.id: profile for profile in context.profiles}
    headers = ["Servicio", "Limit key", "Nombre", "Scope", "Tipo", "Valor", "Unidad", "Default", "Aumento solicitable", "Confianza", "Notas", "Fuente", "Recuperado", "Actualizado"]
    for column, header in enumerate(headers, start=1):
        sheet.cell(4, column, header)
    _style_table_header(sheet[4])
    for row, limit in enumerate(context.limits, start=5):
        profile = profile_map.get(limit.service_profile_id)
        sheet.append([
            profile.name if profile else limit.service_profile_id,
            limit.limit_key,
            limit.label,
            limit.scope,
            limit.limit_type,
            _readable_value(limit.value),
            limit.unit or "",
            _readable_value(limit.default_value),
            "Sí" if limit.can_request_increase else "No",
            limit.confidence,
            _safe_text(limit.notes),
            limit.source_url or "",
            limit.source_retrieved_at.isoformat() if limit.source_retrieved_at else "",
            limit.updated_at.isoformat(),
        ])
        sheet.cell(row, 10).number_format = "0%"
        for cell in sheet[row]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.row_dimensions[row].height = 44
    sheet.freeze_panes = "D5"
    sheet.auto_filter.ref = f"A4:N{sheet.max_row}"
    _set_widths(sheet, {1: 30, 2: 28, 3: 38, 4: 18, 5: 18, 6: 18, 7: 12, 8: 18, 9: 18, 10: 12, 11: 46, 12: 58, 13: 24, 14: 24})


def _create_interoperability_sheet(workbook: Workbook, context: TemplateContext) -> None:
    sheet = workbook.create_sheet("Interoperabilidad")
    _style_title(sheet, "Matriz de Interoperabilidad", "Relaciones direccionales gobernadas entre servicios; Supported no reemplaza la validación de región, red y deployment.", 13)
    profile_map = {profile.id: profile for profile in context.profiles}
    headers = ["Servicio origen", "Servicio destino", "Relación", "Soportado", "Dirección", "Patrones", "Componentes requeridos", "Restricciones", "Riesgos", "Confianza", "Última verificación", "Fuente", "Actualizado"]
    for column, header in enumerate(headers, start=1):
        sheet.cell(4, column, header)
    _style_table_header(sheet[4])
    for row, rule in enumerate(context.rules, start=5):
        source = profile_map.get(rule.source_service_profile_id)
        target = profile_map.get(rule.target_service_profile_id)
        sheet.append([
            source.name if source else rule.source_service_profile_id,
            target.name if target else rule.target_service_profile_id,
            rule.relationship_type,
            "Sí" if rule.supported else "No",
            rule.directionality,
            _readable_value(rule.patterns),
            _readable_value(rule.required_components),
            _readable_value(rule.constraints),
            _safe_text(rule.risk_notes),
            rule.confidence,
            rule.last_verified_at.isoformat() if rule.last_verified_at else "",
            rule.source_url or "",
            rule.updated_at.isoformat(),
        ])
        sheet.cell(row, 10).number_format = "0%"
        for cell in sheet[row]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.row_dimensions[row].height = 90
    sheet.freeze_panes = "C5"
    sheet.auto_filter.ref = f"A4:M{sheet.max_row}"
    _set_widths(sheet, {1: 30, 2: 30, 3: 24, 4: 14, 5: 20, 6: 28, 7: 36, 8: 48, 9: 46, 10: 12, 11: 24, 12: 58, 13: 24})


def _protect_reference_sheets(workbook: Workbook) -> None:
    for sheet_name in ["Dashboard", "Ejemplos Guiados", "Guía de Campos", "Patrones", "Servicios OCI", "Límites OCI", "Interoperabilidad"]:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                cell.protection = Protection(locked=True)
        sheet.protection.sheet = True
        sheet.protection.autoFilter = False
        sheet.protection.sort = False


def _set_properties(workbook: Workbook, generated_at: datetime) -> None:
    workbook.properties.creator = "OCI DIS Architect"
    workbook.properties.lastModifiedBy = "OCI DIS Architect"
    workbook.properties.title = "OCI DIS Architect Governed Offline Capture Workbook"
    workbook.properties.subject = f"Template v{TEMPLATE_VERSION}"
    workbook.properties.description = "Offline integration capture with governed patterns and OCI Service Product Library references."
    workbook.properties.created = generated_at
    workbook.properties.modified = generated_at
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True


def _metadata(context: TemplateContext, generated_at: datetime) -> CaptureTemplateMetadata:
    evidence_dates = [source.last_checked_at for source in context.evidence if source.last_checked_at]
    last_verified_at = max(evidence_dates) if evidence_dates else None
    stale_count = sum(1 for source in context.evidence if source.status in {"stale", "pending_verification", "seeded_pending_verification", "failed", "source_unavailable"})
    return CaptureTemplateMetadata(
        template_version=TEMPLATE_VERSION,
        importer_min_version=IMPORTER_MIN_VERSION,
        filename=TEMPLATE_FILENAME,
        generated_at=generated_at,
        capture_sheet=CAPTURE_SHEET_NAME,
        capture_row_limit=CAPTURE_ROW_LIMIT - 1,
        pattern_count=len(context.patterns),
        service_product_count=len(context.profiles),
        service_limit_count=len(context.limits),
        interoperability_rule_count=len(context.rules),
        evidence_source_count=len(context.evidence),
        stale_evidence_count=stale_count,
        last_verified_at=last_verified_at,
        columns=[
            CaptureTemplateColumnMetadata(
                field=column.field,
                header=column.header,
                section=column.section,
                requirement=column.requirement,
                data_type=column.data_type,
                description=column.description,
            )
            for column in COLUMNS
        ],
    )


async def get_capture_template_metadata(db: AsyncSession) -> CaptureTemplateMetadata:
    """Return the current template contract without generating XLSX bytes."""

    context = await _load_context(db)
    return _metadata(context, _utc_now())


async def generate_capture_template(db: AsyncSession) -> tuple[bytes, CaptureTemplateMetadata]:
    """Generate the governed workbook and matching response metadata."""

    generated_at = _utc_now()
    context = await _load_context(db)
    workbook = Workbook()
    _set_properties(workbook, generated_at)
    _create_start_sheet(workbook, context, generated_at)
    _create_dashboard_sheet(workbook)
    list_names = _create_lists_sheet(workbook, context, generated_at)
    list_names.update(_create_client_catalogs_sheet(workbook))
    _create_capture_sheet(workbook, list_names)
    _create_preflight_sheet(workbook)
    _create_examples_sheet(workbook, context)
    _create_field_guide_sheet(workbook)
    _create_patterns_sheet(workbook, context)
    _create_services_sheet(workbook, context)
    _create_limits_sheet(workbook, context)
    _create_interoperability_sheet(workbook, context)
    _protect_reference_sheets(workbook)
    workbook.active = workbook.sheetnames.index("Inicio")
    output = BytesIO()
    workbook.save(output)
    return output.getvalue(), _metadata(context, generated_at)
