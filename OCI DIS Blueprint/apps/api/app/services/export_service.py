"""Synchronous file export helpers for catalog, snapshot, and dashboard artifacts."""

from __future__ import annotations

import json
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, cast
from uuid import uuid4

from fastapi import HTTPException
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    AiReviewBaseline,
    AiReviewJob,
    AiReviewJobStatus,
    DeploymentScenario,
    PriceCatalogSnapshot,
    PriceSource,
    Project,
    VolumetrySnapshot,
)
from app.schemas.export import ExportJobResponse
from app.services import bom_service, dashboard_service, justification_service, recalc_service, storage_service
from app.services.catalog_service import list_integrations
from app.services.pattern_support import get_pattern_support
from app.services.serializers import sanitize_for_json


EXPORT_ROOT = Path(tempfile.gettempdir()) / "oci-dis-exports"
FILES_DIR = EXPORT_ROOT / "files"
JOBS_DIR = EXPORT_ROOT / "jobs"


def _ensure_export_dirs() -> None:
    FILES_DIR.mkdir(parents=True, exist_ok=True)
    JOBS_DIR.mkdir(parents=True, exist_ok=True)


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _bom_export_value(status: str, value: object) -> object | None:
    """Suppress persistence sentinels that are not governed commercial prices."""

    return None if status == "rate_card_required" else value


async def _load_project(project_id: str, db: AsyncSession) -> Project:
    project = await db.get(Project, project_id)
    if project is None:
        raise HTTPException(
            status_code=404,
            detail={"detail": "Project not found", "error_code": "PROJECT_NOT_FOUND"},
        )
    return project


async def _load_snapshot(project_id: str, snapshot_id: str, db: AsyncSession) -> VolumetrySnapshot:
    snapshot = await db.scalar(
        select(VolumetrySnapshot).where(
            VolumetrySnapshot.project_id == project_id,
            VolumetrySnapshot.id == snapshot_id,
        )
    )
    if snapshot is None:
        raise HTTPException(
            status_code=404,
            detail={"detail": "Volumetry snapshot not found", "error_code": "VOLUMETRY_SNAPSHOT_NOT_FOUND"},
        )
    return snapshot


async def latest_snapshot_id(project_id: str, db: AsyncSession) -> str:
    """Return the latest volumetry snapshot ID for one project."""

    snapshot = await db.scalar(
        select(VolumetrySnapshot)
        .where(VolumetrySnapshot.project_id == project_id)
        .order_by(VolumetrySnapshot.created_at.desc())
    )
    if snapshot is None:
        raise HTTPException(
            status_code=404,
            detail={"detail": "Volumetry snapshot not found", "error_code": "VOLUMETRY_SNAPSHOT_NOT_FOUND"},
        )
    return snapshot.id


def _job_file(job_id: str) -> Path:
    return JOBS_DIR / f"{job_id}.json"


def _file_path(job_id: str, extension: str) -> Path:
    return FILES_DIR / f"{job_id}.{extension}"


def _artifact_key(project_id: str, job_id: str, extension: str) -> str:
    return f"exports/{project_id}/files/{job_id}.{extension}"


def _metadata_key(project_id: str, job_id: str) -> str:
    return f"exports/{project_id}/jobs/{job_id}.json"


def _write_job_metadata(job: ExportJobResponse, file_reference: str) -> str:
    payload = sanitize_for_json(
        {
            **job.model_dump(),
            "created_at": job.created_at,
            "file_reference": file_reference,
        }
    )
    return storage_service.put_json(
        _metadata_key(job.project_id, job.job_id),
        cast(dict[str, object], payload),
    )


def _job_from_payload(payload: dict[str, object]) -> ExportJobResponse:
    created_at = payload.get("created_at")
    normalized = {
        key: value
        for key, value in payload.items()
        if key not in {"file_path", "file_reference"}
    }
    if isinstance(created_at, str):
        normalized["created_at"] = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
    return ExportJobResponse(**cast(dict[str, Any], normalized))


async def _build_job(
    project_id: str,
    snapshot_id: str,
    export_format: str,
    filename: str,
    file_path: Path,
) -> ExportJobResponse:
    created_at = _utc_now()
    job_id = file_path.stem
    job = ExportJobResponse(
        job_id=job_id,
        project_id=project_id,
        snapshot_id=snapshot_id,
        format=export_format,
        status="completed",
        filename=filename,
        download_url=f"/api/v1/exports/{project_id}/jobs/{job_id}/download",
        created_at=created_at,
    )
    extension = file_path.suffix.lstrip(".")
    file_reference = storage_service.put_bytes(
        _artifact_key(project_id, job_id, extension),
        file_path.read_bytes(),
        metadata={"project-id": project_id, "snapshot-id": snapshot_id},
    )
    file_path.unlink(missing_ok=True)
    _write_job_metadata(job, file_reference)
    return job


def _pdf_escape(value: str) -> str:
    return value.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _excel_cell_value(value: object) -> object:
    serialized = sanitize_for_json(value)
    if isinstance(serialized, (list, dict)):
        return json.dumps(serialized, ensure_ascii=True)
    return serialized


def _pattern_support_payload(pattern_ids: list[str]) -> dict[str, object]:
    selected_pattern_ids = sorted({pattern_id for pattern_id in pattern_ids if pattern_id})
    return {
        "certified_pattern_ids": [
            pattern_id
            for pattern_id in selected_pattern_ids
            if get_pattern_support(pattern_id).certification_status == "certified"
        ],
        "uncertified_pattern_ids": [
            pattern_id
            for pattern_id in selected_pattern_ids
            if get_pattern_support(pattern_id).certification_status != "certified"
        ],
        "certification_versions": sorted(
            {
                version
                for pattern_id in selected_pattern_ids
                if (version := get_pattern_support(pattern_id).certification_version)
            }
        ),
    }


def _render_basic_pdf(lines: list[str], file_path: Path) -> None:
    sanitized = [line[:110] for line in lines]
    content_parts = ["BT", "/F1 12 Tf", "50 790 Td"]
    for index, line in enumerate(sanitized):
        if index > 0:
            content_parts.append("0 -16 Td")
        content_parts.append(f"({_pdf_escape(line)}) Tj")
    content_parts.append("ET")
    stream = "\n".join(content_parts).encode("latin-1", "replace")

    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        (
            "<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
            "/Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>"
        ).encode("ascii"),
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        f"<< /Length {len(stream)} >>\nstream\n".encode("ascii") + stream + b"\nendstream",
    ]

    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(len(pdf))
        pdf.extend(f"{index} 0 obj\n".encode("ascii"))
        pdf.extend(obj)
        pdf.extend(b"\nendobj\n")

    xref_offset = len(pdf)
    pdf.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    pdf.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    pdf.extend(
        (
            f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\n"
            f"startxref\n{xref_offset}\n%%EOF"
        ).encode("ascii")
    )
    file_path.write_bytes(pdf)


async def create_xlsx_export(
    project_id: str,
    snapshot_id: str,
    db: AsyncSession,
) -> ExportJobResponse:
    """Generate a workbook with catalog, row-level volumetry, and consolidated totals."""

    await _load_project(project_id, db)
    snapshot = await _load_snapshot(project_id, snapshot_id, db)
    catalog_page = await list_integrations(project_id, page=1, page_size=10_000, filters={}, db=db)

    workbook = Workbook()
    catalog_sheet = workbook.active
    catalog_sheet.title = "Catalog"
    catalog_rows = [item.model_dump() for item in catalog_page.integrations]
    catalog_headers = list(catalog_rows[0].keys()) if catalog_rows else ["id"]
    catalog_sheet.append(catalog_headers)
    for row in catalog_rows:
        catalog_sheet.append([_excel_cell_value(row.get(header)) for header in catalog_headers])

    volumetry_sheet = workbook.create_sheet("Volumetry")
    volumetry_headers = [
        "integration_id",
        "executions_per_day",
        "payload_per_hour_kb",
        "oic_billing_msgs_month",
        "functions_invocations_month",
        "functions_execution_units_gb_s",
        "data_integration_gb_month",
        "streaming_gb_month",
        "streaming_partition_count",
        "pattern_certification_status",
        "pattern_certification_version",
        "pattern_sizing_strategy",
        "pattern_composition_compliant",
        "pattern_composition_issues",
    ]
    volumetry_sheet.append(volumetry_headers)
    for integration_id, metrics in snapshot.row_results.items():
        raw_certification = metrics.get("pattern_certification", {})
        certification = (
            cast(dict[str, object], raw_certification)
            if isinstance(raw_certification, dict)
            else {}
        )
        volumetry_sheet.append(
            [
                integration_id,
                *[metrics.get(header) for header in volumetry_headers[1:9]],
                certification.get("status"),
                certification.get("version"),
                certification.get("sizing_strategy"),
                certification.get("composition_compliant"),
                _excel_cell_value(certification.get("composition_issues", [])),
            ]
        )

    consolidated_sheet = workbook.create_sheet("Consolidated")
    consolidated_sheet.append(["domain", "metric", "value"])
    for domain, metrics in snapshot.consolidated.items():
        for metric, value in metrics.items():
            consolidated_sheet.append([domain, metric, _excel_cell_value(value)])

    provenance_sheet = workbook.create_sheet("Rule Provenance")
    provenance_sheet.append(["key", "value"])
    raw_service_rules = (snapshot.snapshot_metadata or {}).get("service_rules", {})
    service_rules = cast(dict[str, object], raw_service_rules) if isinstance(raw_service_rules, dict) else {}
    for key, value in sorted(service_rules.items()):
        provenance_sheet.append([key, _excel_cell_value(value)])

    support_sheet = workbook.create_sheet("Pattern Certification")
    support_sheet.append(
        [
            "Pattern ID",
            "Certification",
            "Version",
            "Sizing Strategy",
            "Required Evidence",
            "Certified Core Compositions",
            "Required Overlays",
            "Commercial Services",
            "External Dependencies",
            "Validation Controls",
            "Summary",
        ]
    )
    selected_pattern_ids = sorted(
        {
            str(row.get("selected_pattern"))
            for row in catalog_rows
            if row.get("selected_pattern")
        }
    )
    if selected_pattern_ids:
        for pattern_id in selected_pattern_ids:
            support = get_pattern_support(pattern_id)
            support_sheet.append(
                [
                    pattern_id,
                    support.badge_label,
                    support.certification_version,
                    support.sizing_strategy,
                    ", ".join(support.required_evidence),
                    " | ".join(" + ".join(group) for group in support.approved_core_tool_groups),
                    " | ".join(" + ".join(group) for group in support.approved_overlay_groups),
                    ", ".join(support.commercial_service_ids),
                    ", ".join(support.external_dependencies),
                    ", ".join(support.validation_controls),
                    support.summary,
                ]
            )
    else:
        support_sheet.append(["—", "No assigned patterns", "—", "—", "—", "—", "—", "—", "—", "—", "This export does not include any selected pattern IDs."])

    _ensure_export_dirs()
    job_id = str(uuid4())
    file_path = _file_path(job_id, "xlsx")
    workbook.save(file_path)
    return await _build_job(
        project_id=project_id,
        snapshot_id=snapshot_id,
        export_format="xlsx",
        filename=f"{project_id}-{snapshot_id}.xlsx",
        file_path=file_path,
    )


async def create_json_export(
    project_id: str,
    snapshot_id: str,
    db: AsyncSession,
) -> ExportJobResponse:
    """Generate a JSON snapshot bundle for one project and volumetry snapshot."""

    project = await _load_project(project_id, db)
    snapshot = await _load_snapshot(project_id, snapshot_id, db)
    catalog_page = await list_integrations(project_id, page=1, page_size=10_000, filters={}, db=db)
    dashboard_snapshot = await dashboard_service.get_snapshot(project_id, snapshot_id, db)
    justifications = await justification_service.list_justifications(project_id, db)

    payload = sanitize_for_json(
        {
            "project": {
                "id": project.id,
                "name": project.name,
                "status": project.status,
                "owner_id": project.owner_id,
            },
            "snapshot": recalc_service.serialize_snapshot(snapshot).model_dump(),
            "dashboard": dashboard_snapshot.model_dump(),
            "catalog": catalog_page.model_dump(),
            "justifications": justifications.model_dump(),
            "pattern_certification": _pattern_support_payload(
                [item.selected_pattern for item in catalog_page.integrations if item.selected_pattern]
            ),
            "exported_at": _utc_now(),
        }
    )

    _ensure_export_dirs()
    job_id = str(uuid4())
    file_path = _file_path(job_id, "json")
    file_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return await _build_job(
        project_id=project_id,
        snapshot_id=snapshot_id,
        export_format="json",
        filename=f"{project_id}-{snapshot_id}.json",
        file_path=file_path,
    )


async def create_pdf_export(
    project_id: str,
    snapshot_id: str,
    db: AsyncSession,
) -> ExportJobResponse:
    """Generate a simple technical dashboard PDF summary."""

    project = await _load_project(project_id, db)
    dashboard_snapshot = await dashboard_service.get_snapshot(project_id, snapshot_id, db)
    catalog_page = await list_integrations(project_id, page=1, page_size=10_000, filters={}, db=db)
    support_boundary = _pattern_support_payload(
        [item.selected_pattern for item in catalog_page.integrations if item.selected_pattern]
    )

    lines = [
        f"OCI DIS Blueprint Dashboard Export - {project.name}",
        f"Project ID: {project_id}",
        f"Snapshot ID: {dashboard_snapshot.snapshot_id}",
        f"OIC msgs/month: {dashboard_snapshot.kpi_strip.oic_msgs_month:g}",
        f"OIC peak packs/hour: {dashboard_snapshot.kpi_strip.peak_packs_hour:g}",
        (
            "DI workspace active: "
            f"{'Yes' if dashboard_snapshot.kpi_strip.di_workspace_active else 'No'}"
        ),
        f"DI GB/month: {dashboard_snapshot.kpi_strip.di_data_processed_gb_month:g}",
        f"Functions GB-s: {dashboard_snapshot.kpi_strip.functions_execution_units_gb_s:g}",
        f"QA OK %: {dashboard_snapshot.maturity.qa_ok_pct:.2f}",
        f"Pattern assigned %: {dashboard_snapshot.maturity.pattern_assigned_pct:.2f}",
        f"Service rules version: {dashboard_snapshot.charts.service_rules.version}",
        f"Service rules freshness: {dashboard_snapshot.charts.service_rules.freshness_status}",
    ]
    if support_boundary["uncertified_pattern_ids"]:
        lines.append(
            "Uncertified patterns in use: "
            + ", ".join(cast(list[str], support_boundary["uncertified_pattern_ids"]))
        )
    if support_boundary["certified_pattern_ids"]:
        lines.append(
            "Certified patterns in use: "
            + ", ".join(cast(list[str], support_boundary["certified_pattern_ids"]))
        )
    for risk in dashboard_snapshot.risks[:5]:
        lines.append(f"Risk {risk.label}: {risk.count}")

    _ensure_export_dirs()
    job_id = str(uuid4())
    file_path = _file_path(job_id, "pdf")
    _render_basic_pdf(lines, file_path)
    return await _build_job(
        project_id=project_id,
        snapshot_id=snapshot_id,
        export_format="pdf",
        filename=f"{project_id}-{snapshot_id}.pdf",
        file_path=file_path,
    )


async def _latest_completed_ai_review(project_id: str, db: AsyncSession) -> AiReviewJob | None:
    """Return the latest completed governed AI review job for brief context."""

    return await db.scalar(
        select(AiReviewJob)
        .where(
            AiReviewJob.project_id == project_id,
            AiReviewJob.status == AiReviewJobStatus.COMPLETED,
            AiReviewJob.result_payload.is_not(None),
        )
        .order_by(AiReviewJob.finished_at.desc().nullslast(), AiReviewJob.created_at.desc())
    )


async def _active_project_baseline(project_id: str, db: AsyncSession) -> AiReviewBaseline | None:
    """Return the active project-level planned baseline when one has been approved."""

    return await db.scalar(
        select(AiReviewBaseline)
        .where(
            AiReviewBaseline.project_id == project_id,
            AiReviewBaseline.scope == "project",
            AiReviewBaseline.is_active.is_(True),
        )
        .order_by(AiReviewBaseline.created_at.desc())
    )


def _brief_text(value: object | None, fallback: str = "Not available") -> str:
    if value is None:
        return fallback
    text = str(value).strip()
    return text or fallback


def _brief_findings(review_payload: dict[str, object] | None) -> list[dict[str, object]]:
    if not review_payload:
        return []
    findings = review_payload.get("findings")
    if not isinstance(findings, list):
        return []
    return [cast(dict[str, object], item) for item in findings if isinstance(item, dict)]


def _brief_items(review_payload: dict[str, object] | None, key: str) -> list[dict[str, object]]:
    if not review_payload:
        return []
    items = review_payload.get(key)
    if not isinstance(items, list):
        return []
    return [cast(dict[str, object], item) for item in items if isinstance(item, dict)]


async def create_brief_export(project_id: str, db: AsyncSession) -> ExportJobResponse:
    """Generate an executive Markdown brief from dashboard, drift, and review evidence."""

    project = await _load_project(project_id, db)
    snapshot_id = await latest_snapshot_id(project_id, db)
    dashboard_snapshot = await dashboard_service.get_snapshot(project_id, snapshot_id, db)
    catalog_page = await list_integrations(project_id, page=1, page_size=10_000, filters={}, db=db)
    latest_review = await _latest_completed_ai_review(project_id, db)
    active_baseline = await _active_project_baseline(project_id, db)
    review_payload = cast(dict[str, object] | None, latest_review.result_payload if latest_review else None)
    drift = cast(dict[str, object], review_payload.get("drift", {})) if review_payload else {}
    support_boundary = _pattern_support_payload(
        [item.selected_pattern for item in catalog_page.integrations if item.selected_pattern]
    )

    lines = [
        f"# OCI DIS Blueprint Executive Brief - {project.name}",
        "",
        f"- Generated: {_utc_now().isoformat()}",
        f"- Project ID: `{project_id}`",
        f"- Snapshot ID: `{snapshot_id}`",
        f"- Catalog integrations: {catalog_page.total}",
        f"- Planned baseline: {_brief_text(active_baseline.label if active_baseline else None, 'No approved baseline')}",
        f"- Service rules: `{dashboard_snapshot.charts.service_rules.version}`",
        f"- Service-rule freshness: {dashboard_snapshot.charts.service_rules.freshness_status}",
        "",
        "## Executive Summary",
        "",
        _brief_text(review_payload.get("summary") if review_payload else None, "No completed AI review is available yet."),
        "",
        "## Core KPIs",
        "",
        f"- OIC billing messages/month: {dashboard_snapshot.kpi_strip.oic_msgs_month:g}",
        f"- OIC peak packs/hour: {dashboard_snapshot.kpi_strip.peak_packs_hour:g}",
        f"- Data Integration workspace active: {'Yes' if dashboard_snapshot.kpi_strip.di_workspace_active else 'No'}",
        f"- Data Integration GB/month: {dashboard_snapshot.kpi_strip.di_data_processed_gb_month:g}",
        f"- Functions GB-s/month: {dashboard_snapshot.kpi_strip.functions_execution_units_gb_s:g}",
        f"- QA OK: {dashboard_snapshot.charts.completeness.qa_ok}",
        f"- QA Review: {dashboard_snapshot.charts.completeness.qa_revisar}",
        f"- QA Pending: {dashboard_snapshot.charts.completeness.qa_pending}",
        "",
        "## Readiness and Drift",
        "",
        f"- Review readiness: {_brief_text(review_payload.get('readiness_label') if review_payload else None)}",
        f"- Readiness score: {_brief_text(review_payload.get('readiness_score') if review_payload else None)}",
        f"- Planned vs actual drift: {_brief_text(drift.get('status'), 'No drift result available')}",
        f"- Drift summary: {_brief_text(drift.get('summary'), 'No drift summary available')}",
        "",
        "## Decision Brief",
        "",
    ]
    decision_brief = cast(dict[str, object], review_payload.get("decision_brief", {})) if review_payload else {}
    if decision_brief:
        lines.extend(
            [
                f"- Sign-off status: {_brief_text(decision_brief.get('signoff_status'))}",
                f"- Primary risk: {_brief_text(decision_brief.get('primary_risk'))}",
                f"- Recommended next action: {_brief_text(decision_brief.get('recommended_next_action'))}",
            ]
        )
    else:
        lines.append("- No AI Review decision brief is available.")

    lines.extend(
        [
            "",
            "## Topology Intelligence",
            "",
        ]
    )
    topology_insights = _brief_items(review_payload, "topology_insights")
    if topology_insights:
        for insight in topology_insights[:5]:
            lines.append(
                "- "
                + f"{_brief_text(insight.get('title'))}: "
                + f"{_brief_text(insight.get('summary'))} "
                + f"({_brief_text(insight.get('metric'))})"
            )
    else:
        lines.append("- No topology insight is available.")

    lines.extend(["", "## Stress Scenarios", ""])
    stress_scenarios = _brief_items(review_payload, "stress_scenarios")
    if stress_scenarios:
        for scenario in stress_scenarios[:6]:
            lines.append(
                "- "
                + f"{_brief_text(scenario.get('name'))}: "
                + f"{_brief_text(scenario.get('summary'))}"
            )
    else:
        lines.append("- No stress scenario is available.")

    lines.extend(["", "## Remediation Plan", ""])
    remediation_plan = _brief_items(review_payload, "remediation_plan")
    if remediation_plan:
        for step in remediation_plan[:6]:
            lines.append(
                f"{_brief_text(step.get('priority'))}. "
                + f"{_brief_text(step.get('title'))} — "
                + f"{_brief_text(step.get('action'))} "
                + f"Owner: {_brief_text(step.get('owner'))}."
            )
    else:
        lines.append("- No remediation plan is available.")

    lines.extend(
        [
            "",
        "## Pattern Certification",
        "",
        "- Certified patterns in use: "
        + ", ".join(cast(list[str], support_boundary["certified_pattern_ids"]) or ["None"]),
        "- Uncertified patterns in use: "
        + ", ".join(cast(list[str], support_boundary["uncertified_pattern_ids"]) or ["None"]),
        "",
        "## Top Risks",
        "",
        ]
    )
    if dashboard_snapshot.risks:
        for risk in dashboard_snapshot.risks[:8]:
            lines.append(f"- {risk.label}: {risk.count}")
    else:
        lines.append("- No dashboard risks were detected in the latest snapshot.")

    findings = [
        finding
        for finding in _brief_findings(review_payload)
        if finding.get("severity") in {"critical", "high", "medium"}
    ][:8]
    lines.extend(["", "## Review Board Highlights", ""])
    if findings:
        for finding in findings:
            lines.append(
                "- "
                + f"[{_brief_text(finding.get('severity')).upper()}] "
                + f"{_brief_text(finding.get('title'))}: "
                + _brief_text(finding.get("recommendation"))
            )
    else:
        lines.append("- No critical, high, or medium AI review findings are available.")

    lines.extend(
        [
            "",
            "## Governance Notes",
            "",
            "- This brief is generated from governed dashboard, catalog, planned baseline, and AI Review evidence.",
            "- It does not mutate catalog data or accept recommendations automatically.",
        ]
    )

    _ensure_export_dirs()
    job_id = str(uuid4())
    file_path = _file_path(job_id, "md")
    file_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return await _build_job(
        project_id=project_id,
        snapshot_id=snapshot_id,
        export_format="md",
        filename=f"{project_id}-{snapshot_id}-executive-brief.md",
        file_path=file_path,
    )


async def create_bom_xlsx_export(
    project_id: str,
    bom_snapshot_id: str,
    db: AsyncSession,
) -> ExportJobResponse:
    """Generate a governed, auditable OCI BOM workbook for offline review."""

    project = await _load_project(project_id, db)
    bom = await bom_service.get_bom_snapshot(project_id, bom_snapshot_id, db)
    scenario = await db.get(DeploymentScenario, bom.scenario_id)
    price_snapshot = await db.get(PriceCatalogSnapshot, bom.price_catalog_snapshot_id)
    source = await db.get(PriceSource, price_snapshot.source_id) if price_snapshot else None

    workbook = Workbook()
    summary = workbook.active
    summary.title = "BOM Summary"
    summary.append(["OCI DIS Blueprint - Governed Bill of Materials"])
    summary.append(["Project", project.name])
    summary.append(["Scenario", scenario.name if scenario else bom.scenario_id])
    summary.append(["Status", bom.publication_status])
    summary.append(["Currency", bom.currency])
    summary.append(["Coverage", f"{bom.coverage_pct:.1f}%"])
    summary.append(["Monthly estimate", bom.monthly_total])
    summary.append(["Peak monthly estimate", bom.peak_monthly_total])
    summary.append(["Steady-state month", bom.steady_state_period])
    summary.append(["First active month", bom.first_active_period])
    summary.append(["Deferred spend from phased activation", bom.ramp_deferred_amount])
    summary.append(["Annual estimate", bom.annual_total])
    summary.append(["Contract estimate", bom.contract_total])
    summary.append(["Notice", "Planning estimate only; not an Oracle quote."])

    lines = workbook.create_sheet("Line Items")
    lines.append(
        [
            "Environment", "Service", "Commercial Variant", "Part Number", "Description", "Metric", "Quantity", "Unit",
            "Unit Price", "Monthly", "Annual", "Contract", "Status", "Formula", "Warnings",
        ]
    )
    for line in bom.line_items:
        requires_rate_card = line.status == "rate_card_required"
        lines.append(
            [
                line.environment, line.service_id, line.provenance.get("commercial_variant"), line.part_number, line.description, line.metric_name,
                line.quantity, line.unit,
                _bom_export_value(line.status, line.unit_price),
                _bom_export_value(line.status, line.monthly_amount),
                _bom_export_value(line.status, line.annual_amount),
                _bom_export_value(line.status, line.contract_amount),
                line.status,
                "Approved customer rate card required" if requires_rate_card else line.formula,
                _excel_cell_value(line.warnings),
            ]
        )

    schedule = workbook.create_sheet("Monthly Schedule")
    schedule.append(
        [
            "Month", "Period Start", "Monthly Total", "Cumulative Total",
            "Environment Mix", "Service Mix",
        ]
    )
    for summary_period in bom.monthly_series:
        schedule.append(
            [
                summary_period.period_index,
                summary_period.period_start,
                summary_period.total,
                summary_period.cumulative_total,
                _excel_cell_value(summary_period.by_environment),
                _excel_cell_value(summary_period.by_service),
            ]
        )

    line_periods = workbook.create_sheet("Line Periods")
    line_periods.append(
        [
            "Line ID", "Environment", "Service", "Part Number", "Month", "Period Start",
            "Normalized Quantity Ratio", "Billed Quantity", "Active Hours", "Unit Price", "Amount", "Status",
        ]
    )
    for line in bom.line_items:
        for line_period in line.periods:
            line_periods.append(
                [
                    line.id, line.environment, line.service_id, line.part_number,
                    line_period.period_index, line_period.period_start,
                    line_period.multiplier, line_period.quantity,
                    line_period.active_hours,
                    _bom_export_value(line_period.status, line_period.unit_price),
                    _bom_export_value(line_period.status, line_period.amount),
                    line_period.status,
                ]
            )

    provenance = workbook.create_sheet("Provenance")
    provenance.append(["Key", "Value"])
    provenance_rows = {
        "BOM snapshot": bom.id,
        "Technical snapshot": bom.technical_snapshot_id,
        "Price catalog snapshot": bom.price_catalog_snapshot_id,
        "Price source": source.name if source else None,
        "Price source type": source.source_type if source else None,
        "Price content hash": price_snapshot.content_hash if price_snapshot else None,
        "Price retrieved at": price_snapshot.retrieved_at if price_snapshot else None,
        "Mapping version": bom.mapping_version,
        "Engine version": bom.engine_version,
        "Scenario assumptions": scenario.scenario_assumptions if scenario else {},
        "Service configuration": scenario.service_config if scenario else {},
        "Start date": scenario.start_date if scenario else None,
        "Proration policy": scenario.proration_policy if scenario else None,
        "Environments": (
            [
                item.model_dump(mode="json")
                for item in (await bom_service.serialize_scenario(scenario, db)).environments
            ]
            if scenario else []
        ),
    }
    for key, value in provenance_rows.items():
        provenance.append([key, _excel_cell_value(value)])

    warnings = workbook.create_sheet("Warnings")
    warnings.append(["Type", "Detail"])
    if bom.warnings:
        for warning in bom.warnings:
            warnings.append(["BOM", _excel_cell_value(warning)])
    else:
        warnings.append(["BOM", "No blocking warnings."])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 60)
            sheet.column_dimensions[column_cells[0].column_letter].width = width

    _ensure_export_dirs()
    job_id = str(uuid4())
    file_path = _file_path(job_id, "xlsx")
    workbook.save(file_path)
    return await _build_job(
        project_id=project_id,
        snapshot_id=bom_snapshot_id,
        export_format="xlsx",
        filename=f"{project_id}-{bom_snapshot_id}-oci-bom.xlsx",
        file_path=file_path,
    )


async def create_bom_json_export(
    project_id: str,
    bom_snapshot_id: str,
    db: AsyncSession,
) -> ExportJobResponse:
    """Generate a machine-readable BOM bundle with immutable provenance."""

    project = await _load_project(project_id, db)
    bom = await bom_service.get_bom_snapshot(project_id, bom_snapshot_id, db)
    scenario = await db.get(DeploymentScenario, bom.scenario_id)
    price_snapshot = await db.get(PriceCatalogSnapshot, bom.price_catalog_snapshot_id)
    source = await db.get(PriceSource, price_snapshot.source_id) if price_snapshot else None
    payload = sanitize_for_json(
        {
            "schema_version": "oci-dis-bom-2.0",
            "project": {"id": project.id, "name": project.name},
            "bom": bom.model_dump(),
            "scenario": (await bom_service.serialize_scenario(scenario, db)).model_dump() if scenario else None,
            "price_provenance": {
                "snapshot_id": price_snapshot.id if price_snapshot else None,
                "source_id": source.id if source else None,
                "source_name": source.name if source else None,
                "source_type": source.source_type if source else None,
                "content_hash": price_snapshot.content_hash if price_snapshot else None,
                "retrieved_at": price_snapshot.retrieved_at if price_snapshot else None,
            },
            "estimate_notice": "Planning estimate only; not an Oracle quote.",
            "exported_at": _utc_now(),
        }
    )
    _ensure_export_dirs()
    job_id = str(uuid4())
    file_path = _file_path(job_id, "json")
    file_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return await _build_job(
        project_id=project_id,
        snapshot_id=bom_snapshot_id,
        export_format="json",
        filename=f"{project_id}-{bom_snapshot_id}-oci-bom.json",
        file_path=file_path,
    )


async def create_bom_pdf_export(
    project_id: str,
    bom_snapshot_id: str,
    db: AsyncSession,
) -> ExportJobResponse:
    """Generate a concise review PDF for one governed BOM snapshot."""

    project = await _load_project(project_id, db)
    bom = await bom_service.get_bom_snapshot(project_id, bom_snapshot_id, db)
    scenario = await db.get(DeploymentScenario, bom.scenario_id)
    lines = [
        "OCI DIS Blueprint - Governed Bill of Materials",
        f"Project: {project.name}",
        f"Scenario: {scenario.name if scenario else bom.scenario_id}",
        f"Status: {bom.publication_status} | Coverage: {bom.coverage_pct:.1f}%",
        f"Monthly: {bom.currency} {bom.monthly_total:,.2f}",
        f"Peak month: {bom.currency} {bom.peak_monthly_total:,.2f}",
        f"First active month: {bom.first_active_period or 'N/A'}",
        f"Steady-state month: {bom.steady_state_period or 'N/A'}",
        f"Phased activation timing effect: {bom.currency} {bom.ramp_deferred_amount:,.2f}",
        f"Annual: {bom.currency} {bom.annual_total:,.2f}",
        f"Contract: {bom.currency} {bom.contract_total:,.2f}",
        "",
        "Top line items:",
    ]
    for item in sorted(bom.line_items, key=lambda line: line.monthly_amount, reverse=True)[:20]:
        amount_label = (
            "Approved customer rate card required"
            if item.status == "rate_card_required"
            else f"{bom.currency} {item.monthly_amount:,.2f}/month"
        )
        lines.append(
            f"{item.environment} | {item.service_id} | {item.provenance.get('commercial_variant', 'Governed default')} | {item.part_number or 'N/A'} | "
            f"{item.quantity:,.3f} {item.unit} | "
            f"{amount_label} | {item.status}"
        )
    lines.extend(["", "Planning estimate only; not an Oracle quote."])
    _ensure_export_dirs()
    job_id = str(uuid4())
    file_path = _file_path(job_id, "pdf")
    _render_basic_pdf(lines, file_path)
    return await _build_job(
        project_id=project_id,
        snapshot_id=bom_snapshot_id,
        export_format="pdf",
        filename=f"{project_id}-{bom_snapshot_id}-oci-bom.pdf",
        file_path=file_path,
    )


async def get_export_job(project_id: str, job_id: str) -> ExportJobResponse:
    """Load export metadata from authoritative Object Storage."""

    payload = _load_export_payload(project_id, job_id)
    if payload.get("project_id") != project_id:
        raise HTTPException(
            status_code=404,
            detail={"detail": "Export job not found", "error_code": "EXPORT_JOB_NOT_FOUND"},
        )
    return _job_from_payload(payload)


def _load_export_payload(project_id: str, job_id: str) -> dict[str, object]:
    reference = storage_service.object_reference(_metadata_key(project_id, job_id))
    try:
        return storage_service.read_json(reference)
    except FileNotFoundError:
        try:
            return storage_service.read_json(
                storage_service.object_reference(f"exports/jobs/{job_id}.json")
            )
        except FileNotFoundError:
            pass
    metadata_path = _job_file(job_id)
    if metadata_path.is_file():
        payload = json.loads(metadata_path.read_text(encoding="utf-8"))
        if isinstance(payload, dict):
            return cast(dict[str, object], payload)
    raise HTTPException(
        status_code=404,
        detail={"detail": "Export job not found", "error_code": "EXPORT_JOB_NOT_FOUND"},
    )


def get_export_content(project_id: str, job_id: str) -> tuple[bytes, ExportJobResponse]:
    """Resolve a generated artifact without materializing it on persistent disk."""

    payload = _load_export_payload(project_id, job_id)
    if payload.get("project_id") != project_id:
        raise HTTPException(
            status_code=404,
            detail={"detail": "Export job not found", "error_code": "EXPORT_JOB_NOT_FOUND"},
        )
    reference = str(payload.get("file_reference") or payload.get("file_path") or "")
    try:
        contents = storage_service.read_bytes(reference)
    except (FileNotFoundError, ValueError):
        raise HTTPException(
            status_code=404,
            detail={"detail": "Export file not found", "error_code": "EXPORT_FILE_NOT_FOUND"},
        )
    job = _job_from_payload(payload)
    return contents, job


def get_export_references(project_id: str, job_id: str) -> tuple[str, str]:
    """Return artifact and metadata references for synthetic-run provenance."""

    payload = _load_export_payload(project_id, job_id)
    artifact = str(payload.get("file_reference") or payload.get("file_path") or "")
    return artifact, storage_service.object_reference(_metadata_key(project_id, job_id))
