"""Pure parser for Oracle commercial price-list workbooks.

The parser deliberately owns no persistence, network, or approval behavior. It
turns the two official workbook sheets into immutable, source-addressable
records that later M51 stages can validate and persist.
"""

from __future__ import annotations

from dataclasses import asdict, dataclass, replace
from decimal import Decimal, InvalidOperation
from io import BytesIO
from os import PathLike
import re
from collections.abc import Callable
from typing import BinaryIO, TypeAlias

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


PRICE_LIST_SHEET = "Oracle PaaS and IaaS Price List"
SUPPLEMENT_SHEET = "Oracle PaaS and IaaS Supplement"

CommercialScalar: TypeAlias = Decimal | bool | str | None
WorkbookSource: TypeAlias = str | PathLike[str] | BinaryIO | BytesIO

_PART_NUMBER_PATTERN = re.compile(r"^[A-Z]\d{4,}[A-Z0-9-]*$")
_NUMERIC_PATTERN = re.compile(r"^[+-]?(?:\d+(?:\.\d*)?|\.\d+)$")


@dataclass(frozen=True, slots=True)
class SourceCell:
    """One non-empty source cell retained for audit and reconciliation."""

    coordinate: str
    value: object
    state: str

    def to_dict(self) -> dict[str, object]:
        return asdict(self)


@dataclass(frozen=True, slots=True)
class SourceEvidence:
    """The exact workbook row used as commercial evidence."""

    sheet: str
    row: int
    cells: tuple[SourceCell, ...]

    def to_dict(self) -> dict[str, object]:
        return asdict(self)


@dataclass(frozen=True, slots=True)
class CommercialPriceTerm:
    """One source-labelled public, commitment, or flex price value."""

    term_type: str
    value: CommercialScalar
    source_cell: str
    source_label: str | None
    value_state: str


@dataclass(frozen=True, slots=True)
class CommercialWorkbookRecord:
    """Canonical commercial terms and supplement evidence for one OCI SKU."""

    part_number: str
    service_name: str
    service_category: str | None
    catalog_section: str | None
    product_family: str | None
    product_group: str | None
    pay_as_you_go: CommercialScalar
    annual_commitment: CommercialScalar
    commercial_price_terms: tuple[CommercialPriceTerm, ...]
    metric: str | None
    metric_minimum: CommercialScalar
    additional_information: str | None
    notes: str | None
    included_entitlements: str | None
    prerequisites: str | None
    product_paths: tuple[tuple[str, ...], ...]
    source_evidence: tuple[SourceEvidence, ...]
    source_conflicts: tuple[dict[str, object], ...] = ()

    @property
    def product_hierarchy(self) -> tuple[str, ...]:
        """Return the workbook product path without repeated adjacent labels."""

        values: list[str] = []
        for value in (
            self.catalog_section,
            self.product_family,
            self.product_group,
            self.service_category,
        ):
            if value and value not in values:
                values.append(value)
        return tuple(values)

    @property
    def source_sheets(self) -> tuple[str, ...]:
        return tuple(dict.fromkeys(item.sheet for item in self.source_evidence))

    @property
    def source_rows(self) -> tuple[tuple[str, int], ...]:
        return tuple((item.sheet, item.row) for item in self.source_evidence)

    def to_dict(self) -> dict[str, object]:
        return asdict(self)


@dataclass(frozen=True, slots=True)
class ParsedCommercialWorkbook:
    """Deterministically ordered canonical records from one workbook."""

    records: tuple[CommercialWorkbookRecord, ...]
    price_record_count: int
    supplement_record_count: int

    def by_part_number(self) -> dict[str, CommercialWorkbookRecord]:
        return {record.part_number: record for record in self.records}

    def to_dicts(self) -> list[dict[str, object]]:
        return [record.to_dict() for record in self.records]


_PRICE_HEADER_ALIASES: dict[str, frozenset[str]] = {
    "part_number": frozenset({"part number", "part no", "part no.", "partnumber"}),
    "service_name": frozenset({"service name", "subscription service", "support service"}),
    "service_category": frozenset({"service category", "category"}),
    "pay_as_you_go": frozenset({"pay as you go", "payg", "price"}),
    "annual_commitment": frozenset({"annual commitment"}),
    "annual_flex": frozenset({"annual flex"}),
    "monthly_flex": frozenset({"monthly flex"}),
    "metric": frozenset({"metric"}),
    "metric_minimum": frozenset({"metric minimum"}),
    "additional_information": frozenset({"additional information"}),
    "notes": frozenset({"notes"}),
    "annual_commitment_minimum": frozenset({"universal credits minimums per year"}),
    "commitment_details": frozenset({"minimum universal credits details"}),
    "monthly_commitment_minimum": frozenset({"universal credits minimum per month"}),
}

_SUPPLEMENT_HEADER_ALIASES: dict[str, frozenset[str]] = {
    "part_number": _PRICE_HEADER_ALIASES["part_number"],
    "service_name": frozenset({"subscription service", "support service", "service name"}),
    "service_category": _PRICE_HEADER_ALIASES["service_category"],
    "metric": frozenset({"metric"}),
    "included_entitlements": frozenset(
        {
            "included with subscription service",
            "included with support service",
            "included entitlements",
        }
    ),
    "prerequisites": frozenset(
        {
            "additional requirements and prerequisites",
            "additional information",
            "prerequisites",
        }
    ),
}


def parse_oci_commercial_workbook(source: WorkbookSource) -> ParsedCommercialWorkbook:
    """Parse official OCI price-list and supplement sheets from ``source``.

    Headers are discovered independently for every repeated table. The parser
    never relies on a fixed row or column number and always closes workbooks it
    opens.
    """

    workbook = load_workbook(source, read_only=False, data_only=True)
    try:
        return parse_oci_commercial_workbook_object(workbook)
    finally:
        workbook.close()


def parse_oci_commercial_workbook_object(workbook: Workbook) -> ParsedCommercialWorkbook:
    """Parse an already-open workbook without closing the caller-owned object."""

    missing = [name for name in (PRICE_LIST_SHEET, SUPPLEMENT_SHEET) if name not in workbook.sheetnames]
    if missing:
        raise ValueError(f"Commercial workbook is missing required sheet(s): {', '.join(missing)}")

    price_sheet = workbook[PRICE_LIST_SHEET]
    supplement_sheet = workbook[SUPPLEMENT_SHEET]
    if not _sheet_contains_header(price_sheet, _PRICE_HEADER_ALIASES, _is_price_header):
        raise ValueError(
            "Commercial workbook Price List layout is not recognized; "
            "no governed Part Number and price header was found"
        )
    if not _sheet_contains_header(
        supplement_sheet,
        _SUPPLEMENT_HEADER_ALIASES,
        _is_supplement_header,
    ):
        raise ValueError(
            "Commercial workbook Supplement layout is not recognized; "
            "no governed Part Number, service, metric, and evidence header was found"
        )

    price_records = _parse_price_sheet(price_sheet)
    supplement_records = _parse_supplement_sheet(supplement_sheet)
    if not price_records:
        raise ValueError(
            "Commercial workbook Price List contains no recognized OCI SKU records"
        )
    if not supplement_records:
        raise ValueError(
            "Commercial workbook Supplement contains no recognized OCI SKU records"
        )
    part_numbers = sorted(set(price_records) | set(supplement_records))
    records = tuple(
        _merge_records(price_records.get(part_number), supplement_records.get(part_number))
        for part_number in part_numbers
    )
    return ParsedCommercialWorkbook(
        records=records,
        price_record_count=len(price_records),
        supplement_record_count=len(supplement_records),
    )


def _sheet_contains_header(
    sheet: Worksheet,
    aliases: dict[str, frozenset[str]],
    predicate: Callable[[dict[str, int]], bool],
) -> bool:
    """Confirm source layout before accepting any extracted commercial evidence."""

    return any(predicate(_detect_header(row, aliases)) for row in sheet.iter_rows())


def _parse_price_sheet(sheet: Worksheet) -> dict[str, CommercialWorkbookRecord]:
    candidates: dict[str, list[CommercialWorkbookRecord]] = {}
    header: dict[str, int] | None = None
    category: str | None = None
    catalog_section: str | None = None
    family_lines: list[str] = []
    last_family_row: int | None = None
    product_group: str | None = None
    current_part_number: str | None = None
    current_service_name: str | None = None
    current_service_category: str | None = None

    for row in sheet.iter_rows():
        section = _catalog_section_from_row(row)
        if section:
            catalog_section = section
            family_lines = []
            last_family_row = None
        family_line = _product_family_line(row)
        if family_line:
            row_number = row[0].row
            if last_family_row is None or row_number > last_family_row + 2:
                family_lines = [family_line]
            elif family_line not in family_lines:
                family_lines.append(family_line)
            last_family_row = row_number
        detected = _detect_header(row, _PRICE_HEADER_ALIASES)
        if _is_price_header(detected):
            header = detected
            category = None
            product_group = None
            current_part_number = None
            current_service_name = None
            current_service_category = None
            continue
        if header is None:
            continue

        part_number = _part_number(_cell_value(row, header["part_number"]))
        if part_number is None:
            if current_part_number and _has_price_detail(row, header):
                candidates.setdefault(current_part_number, []).append(
                    _price_record(
                        sheet,
                        row,
                        header,
                        part_number=current_part_number,
                        service_name=current_service_name or current_part_number,
                        service_category=current_service_category,
                        catalog_section=catalog_section,
                        product_family=_joined_family(family_lines),
                        product_group=product_group,
                    )
                )
            elif _row_has_content(row):
                detected_category = _category_from_row(row, header)
                if detected_category:
                    if product_group is None:
                        product_group = detected_category
                    category = detected_category
                current_part_number = None
                current_service_name = None
                current_service_category = None
            continue

        service_name = _service_name(row, header)
        if not service_name:
            continue
        current_part_number = part_number
        current_service_name = service_name
        current_service_category = _text_at(row, header.get("service_category")) or category
        record = _price_record(
            sheet,
            row,
            header,
            part_number=part_number,
            service_name=service_name,
            service_category=current_service_category,
            catalog_section=catalog_section,
            product_family=_joined_family(family_lines),
            product_group=product_group,
        )
        candidates.setdefault(part_number, []).append(record)

    return {part_number: _canonical_price(records) for part_number, records in candidates.items()}


def _parse_supplement_sheet(sheet: Worksheet) -> dict[str, CommercialWorkbookRecord]:
    candidates: dict[str, list[CommercialWorkbookRecord]] = {}
    header: dict[str, int] | None = None
    category: str | None = None
    catalog_section: str | None = None
    family_lines: list[str] = []
    last_family_row: int | None = None
    product_group: str | None = None
    current_part_number: str | None = None
    current_service_name: str | None = None
    current_service_category: str | None = None

    for row in sheet.iter_rows():
        section = _catalog_section_from_row(row)
        if section:
            catalog_section = section
            family_lines = []
            last_family_row = None
        family_line = _product_family_line(row)
        if family_line:
            row_number = row[0].row
            if last_family_row is None or row_number > last_family_row + 2:
                family_lines = [family_line]
            elif family_line not in family_lines:
                family_lines.append(family_line)
            last_family_row = row_number
        detected = _detect_header(row, _SUPPLEMENT_HEADER_ALIASES)
        if _is_supplement_header(detected):
            header = detected
            category = None
            product_group = None
            current_part_number = None
            current_service_name = None
            current_service_category = None
            continue
        if header is None:
            continue

        part_number = _part_number(_cell_value(row, header["part_number"]))
        if part_number is None:
            if current_part_number and _has_supplement_detail(row, header):
                candidates.setdefault(current_part_number, []).append(
                    _supplement_record(
                        sheet,
                        row,
                        header,
                        part_number=current_part_number,
                        service_name=current_service_name or current_part_number,
                        service_category=current_service_category,
                        catalog_section=catalog_section,
                        product_family=_joined_family(family_lines),
                        product_group=product_group,
                    )
                )
            elif _row_has_content(row):
                detected_category = _category_from_row(row, header)
                if detected_category:
                    if product_group is None:
                        product_group = detected_category
                    category = detected_category
                current_part_number = None
                current_service_name = None
                current_service_category = None
            continue

        service_name = _service_name(row, header)
        if not service_name:
            continue
        current_part_number = part_number
        current_service_name = service_name
        current_service_category = _text_at(row, header.get("service_category")) or category
        record = _supplement_record(
            sheet,
            row,
            header,
            part_number=part_number,
            service_name=service_name,
            service_category=current_service_category,
            catalog_section=catalog_section,
            product_family=_joined_family(family_lines),
            product_group=product_group,
        )
        candidates.setdefault(part_number, []).append(record)

    return {part_number: _canonical_supplement(records) for part_number, records in candidates.items()}


def _price_record(
    sheet: Worksheet,
    row: tuple[Cell, ...],
    header: dict[str, int],
    *,
    part_number: str,
    service_name: str,
    service_category: str | None,
    catalog_section: str | None,
    product_family: str | None,
    product_group: str | None,
) -> CommercialWorkbookRecord:
    return CommercialWorkbookRecord(
        part_number=part_number,
        service_name=service_name,
        service_category=service_category,
        catalog_section=catalog_section,
        product_family=product_family,
        product_group=product_group,
        pay_as_you_go=_commercial_scalar(_cell_value(row, header.get("pay_as_you_go"))),
        annual_commitment=_commercial_scalar(_cell_value(row, header.get("annual_commitment"))),
        commercial_price_terms=_commercial_price_terms(row, header),
        metric=_text_at(row, header.get("metric")),
        metric_minimum=_commercial_scalar(_cell_value(row, header.get("metric_minimum"))),
        additional_information=(
            _text_at(row, header.get("additional_information"))
            or _text_at(row, header.get("commitment_details"))
        ),
        notes=_text_at(row, header.get("notes")),
        included_entitlements=None,
        prerequisites=None,
        product_paths=(
            _product_path(catalog_section, product_family, product_group, service_category),
        ),
        source_evidence=(_evidence(sheet.title, row, header),),
    )


def _supplement_record(
    sheet: Worksheet,
    row: tuple[Cell, ...],
    header: dict[str, int],
    *,
    part_number: str,
    service_name: str,
    service_category: str | None,
    catalog_section: str | None,
    product_family: str | None,
    product_group: str | None,
) -> CommercialWorkbookRecord:
    return CommercialWorkbookRecord(
        part_number=part_number,
        service_name=service_name,
        service_category=service_category,
        catalog_section=catalog_section,
        product_family=product_family,
        product_group=product_group,
        pay_as_you_go=None,
        annual_commitment=None,
        commercial_price_terms=(),
        metric=_text_at(row, header.get("metric")),
        metric_minimum=None,
        additional_information=None,
        notes=None,
        included_entitlements=_text_at(row, header.get("included_entitlements")),
        prerequisites=_text_at(row, header.get("prerequisites")),
        product_paths=(
            _product_path(catalog_section, product_family, product_group, service_category),
        ),
        source_evidence=(_evidence(sheet.title, row, header),),
    )


def _has_price_detail(row: tuple[Cell, ...], header: dict[str, int]) -> bool:
    return any(
        not _is_blank(_cell_value(row, header.get(field)))
        for field in (
            "pay_as_you_go",
            "annual_commitment",
            "annual_flex",
            "monthly_flex",
            "metric",
            "metric_minimum",
            "additional_information",
            "notes",
            "annual_commitment_minimum",
            "commitment_details",
            "monthly_commitment_minimum",
        )
    )


def _has_supplement_detail(row: tuple[Cell, ...], header: dict[str, int]) -> bool:
    return any(
        not _is_blank(_cell_value(row, header.get(field)))
        for field in ("metric", "included_entitlements", "prerequisites")
    )


def _row_has_content(row: tuple[Cell, ...]) -> bool:
    return any(not _is_blank(cell.value) for cell in row)


def _catalog_section_from_row(row: tuple[Cell, ...]) -> str | None:
    for cell in row:
        value = _clean_text(cell.value)
        if value and value.casefold().startswith("section "):
            return value
    return None


def _product_family_line(row: tuple[Cell, ...]) -> str | None:
    """Identify the centered page-family title used by the official workbook.

    Product-family titles are visually distinct from table group headings: they
    are bold, centered, and use 14pt type. The rule is style-based rather than
    tied to workbook row numbers or style IDs, which change between releases.
    """

    populated = [cell for cell in row if not _is_blank(cell.value)]
    if len(populated) != 1:
        return None
    cell = populated[0]
    value = _clean_text(cell.value)
    if not value or value.casefold().startswith("section "):
        return None
    if not cell.font.bold or float(cell.font.sz or 0) < 13.5:
        return None
    if cell.alignment.horizontal != "center":
        return None
    return value.rstrip(":").strip()


def _joined_family(lines: list[str]) -> str | None:
    unique = list(dict.fromkeys(line.strip() for line in lines if line.strip()))
    return ": ".join(unique) or None


def _product_path(*values: str | None) -> tuple[str, ...]:
    path: list[str] = []
    for value in values:
        if value and value not in path:
            path.append(value)
    return tuple(path)


def _detect_header(row: tuple[Cell, ...], aliases: dict[str, frozenset[str]]) -> dict[str, int]:
    detected: dict[str, int] = {}
    for index, cell in enumerate(row):
        normalized = _header_text(cell.value)
        if not normalized:
            continue
        for field, accepted in aliases.items():
            if field not in detected and normalized in accepted:
                detected[field] = index
    return detected


def _is_price_header(header: dict[str, int]) -> bool:
    price_fields = {"pay_as_you_go", "annual_commitment", "annual_flex", "monthly_flex"}
    special_fields = {
        "annual_commitment_minimum",
        "commitment_details",
        "monthly_commitment_minimum",
    }
    # Legacy Pre-Paid / Metered tables expose Price, Minimum, Includes, Notes,
    # and Part Number without a Metric column. They are still authoritative
    # commercial tables and must reset the parser's hierarchy state.
    return "part_number" in header and (
        bool(price_fields & header.keys()) or bool(special_fields & header.keys())
    )


def _is_supplement_header(header: dict[str, int]) -> bool:
    return {"part_number", "service_name", "metric"}.issubset(header) and bool(
        {"included_entitlements", "prerequisites"} & header.keys()
    )


def _service_name(row: tuple[Cell, ...], header: dict[str, int]) -> str | None:
    explicit = _text_at(row, header.get("service_name"))
    if explicit:
        return explicit

    # Some official Price List tables place a second ``Price`` column after
    # Notes and Part Number. Using the first commercial column as the only
    # boundary can therefore expose numeric note cells as a product name. The
    # workbook contract is more stable: the product description is textual and
    # appears before Part Number. Prefer that region and ignore numeric labels.
    part_number_column = header.get("part_number")
    if part_number_column is not None:
        for index in range(part_number_column):
            value = _cell_value(row, index)
            if not isinstance(value, str):
                continue
            text = _clean_text(value)
            if (
                text
                and text != "-"
                and _part_number(text) is None
                and _NUMERIC_PATTERN.fullmatch(text) is None
            ):
                return text

    commercial_columns = [
        header[field]
        for field in (
            "pay_as_you_go", "annual_commitment", "annual_flex", "monthly_flex",
            "metric", "metric_minimum",
            "annual_commitment_minimum", "monthly_commitment_minimum",
        )
        if field in header
    ]
    if not commercial_columns:
        return None
    for index in range(min(commercial_columns) - 1, -1, -1):
        value = _cell_value(row, index)
        if not isinstance(value, str):
            continue
        text = _clean_text(value)
        if (
            text
            and text != "-"
            and _part_number(text) is None
            and _NUMERIC_PATTERN.fullmatch(text) is None
        ):
            return text
    return None


def _category_from_row(row: tuple[Cell, ...], header: dict[str, int]) -> str | None:
    explicit = _text_at(row, header.get("service_category"))
    if explicit:
        return explicit
    service_column = header.get("service_name")
    if service_column is not None:
        return _text_at(row, service_column)

    commercial_columns = [
        header[field]
        for field in ("pay_as_you_go", "annual_commitment", "metric")
        if field in header
    ]
    if not commercial_columns:
        return None
    values = [
        _clean_text(cell.value)
        for cell in row[: min(commercial_columns)]
        if _clean_text(cell.value) is not None
    ]
    return values[-1] if len(values) == 1 else None


def _canonical_price(records: list[CommercialWorkbookRecord]) -> CommercialWorkbookRecord:
    ordered = _deduplicated_records(records)
    canonical = sorted(ordered, key=lambda record: (-_price_score(record), _evidence_key(record)))[0]
    identity = sorted(ordered, key=_identity_rank)[0]
    terms = _deduplicated_price_terms(ordered)
    evidence = tuple(item for record in ordered for item in record.source_evidence)
    return replace(
        canonical,
        service_name=identity.service_name,
        service_category=identity.service_category,
        catalog_section=identity.catalog_section,
        product_family=identity.product_family,
        product_group=identity.product_group,
        pay_as_you_go=_first_scalar(ordered, "pay_as_you_go"),
        annual_commitment=_first_scalar(ordered, "annual_commitment"),
        commercial_price_terms=terms,
        metric=_first_text(ordered, "metric"),
        metric_minimum=_first_scalar(ordered, "metric_minimum"),
        additional_information=_joined_text(ordered, "additional_information"),
        notes=_joined_text(ordered, "notes"),
        product_paths=_deduplicated_product_paths(ordered),
        source_evidence=evidence,
        source_conflicts=_repeated_record_conflicts(ordered),
    )


def _canonical_supplement(records: list[CommercialWorkbookRecord]) -> CommercialWorkbookRecord:
    ordered = _deduplicated_records(records)
    canonical = sorted(ordered, key=lambda record: (-_supplement_score(record), _evidence_key(record)))[0]
    identity = sorted(ordered, key=_identity_rank)[0]
    evidence = tuple(item for record in ordered for item in record.source_evidence)
    return replace(
        canonical,
        service_name=identity.service_name,
        service_category=identity.service_category,
        catalog_section=identity.catalog_section,
        product_family=identity.product_family,
        product_group=identity.product_group,
        metric=_first_text(ordered, "metric"),
        included_entitlements=_joined_text(ordered, "included_entitlements"),
        prerequisites=_joined_text(ordered, "prerequisites"),
        product_paths=_deduplicated_product_paths(ordered),
        source_evidence=evidence,
        source_conflicts=_repeated_record_conflicts(ordered),
    )


def _deduplicated_records(
    records: list[CommercialWorkbookRecord],
) -> list[CommercialWorkbookRecord]:
    unique: dict[tuple[object, ...], CommercialWorkbookRecord] = {}
    for record in sorted(records, key=_evidence_key):
        source_values = tuple(
            cell.value
            for evidence in record.source_evidence
            for cell in evidence.cells
            if cell.state != "blank"
        )
        price_terms = tuple(
            (term.term_type, term.value, term.source_label, term.value_state)
            for term in record.commercial_price_terms
        )
        signature = (
            record.part_number,
            record.service_name,
            record.service_category,
            record.catalog_section,
            record.product_family,
            record.product_group,
            record.pay_as_you_go,
            record.annual_commitment,
            price_terms,
            record.metric,
            record.metric_minimum,
            record.additional_information,
            record.notes,
            record.included_entitlements,
            record.prerequisites,
            record.product_paths,
            source_values,
        )
        unique.setdefault(signature, record)
    return list(unique.values())


def _deduplicated_price_terms(
    records: list[CommercialWorkbookRecord],
) -> tuple[CommercialPriceTerm, ...]:
    unique: dict[tuple[object, ...], CommercialPriceTerm] = {}
    for record in records:
        for term in record.commercial_price_terms:
            signature = (
                term.term_type,
                term.value,
                term.source_label,
                term.value_state,
            )
            unique.setdefault(signature, term)
    return tuple(unique.values())


def _deduplicated_product_paths(
    records: list[CommercialWorkbookRecord],
) -> tuple[tuple[str, ...], ...]:
    paths = {
        path
        for record in records
        for path in record.product_paths
        if path
    }
    return tuple(sorted(paths, key=lambda path: (len(path), path)))


def _semantic_text(value: str | None) -> str | None:
    """Normalize source prose for conflict comparison without changing evidence."""

    if not value or not value.strip() or value.strip() == "-":
        return None
    normalized = re.sub(r"\s+", " ", value).strip().casefold()
    normalized = re.sub(
        r"\s*\((?:priced in advance of availability|continued)\)\s*$",
        "",
        normalized,
    )
    return re.sub(r"(?:\s*\(?(?:note|footnote)\s*\d+\)?\s*)+$", "", normalized)


def _valid_metric_identity(value: str) -> bool:
    normalized = _semantic_text(value)
    if not normalized or _NUMERIC_PATTERN.fullmatch(normalized):
        return False
    if normalized in {"metric", "not applicable.", "not applicable", "n/a"}:
        return False
    if normalized.startswith("country zone") or re.fullmatch(r"[a-z]{3}-[a-z0-9-]+", normalized):
        return False
    return True


def _repeated_record_conflicts(
    records: list[CommercialWorkbookRecord],
) -> tuple[dict[str, object], ...]:
    """Report only material disagreements across repeated official SKU rows.

    Product placement is intentionally excluded: a SKU may be listed under many
    workbook products or sections. Blank values do not conflict with populated
    values, and price tiers are compared only when their typed term and source
    label overlap.
    """

    conflicts: list[dict[str, object]] = []
    for field in ("service_name", "metric"):
        observed: dict[str, set[str]] = {}
        for record in records:
            raw = getattr(record, field)
            if not isinstance(raw, str):
                continue
            normalized = _semantic_text(raw)
            if normalized and (field != "metric" or _valid_metric_identity(raw)):
                observed.setdefault(normalized, set()).add(raw.strip())
        if len(observed) > 1:
            conflicts.append(
                {
                    "field": field,
                    "values": sorted({value for values in observed.values() for value in values}),
                }
            )

    term_scope = tuple[tuple[str, ...], str | None, str | None, str, str | None]
    scoped_terms: dict[term_scope, dict[str, set[str]]] = {}
    for record in records:
        product_path = tuple(_header_text(value) for value in record.product_hierarchy)
        metric = _semantic_text(record.metric)
        minimum = _semantic_scalar(record.metric_minimum)
        for term in record.commercial_price_terms:
            value = _semantic_scalar(term.value)
            if value is None:
                continue
            scope: term_scope = (
                product_path,
                metric,
                minimum,
                term.term_type,
                _semantic_text(term.source_label),
            )
            scoped_terms.setdefault(scope, {}).setdefault(value, set()).add(
                str(term.value)
            )
    for scope, observed in sorted(scoped_terms.items(), key=lambda item: repr(item[0])):
        if len(observed) > 1:
            conflicts.append(
                {
                    "field": "commercial_price_term",
                    "scope": [list(scope[0]), *scope[1:]],
                    "values": sorted(
                        {value for values in observed.values() for value in values}
                    ),
                }
            )

    return tuple(conflicts)


def _semantic_scalar(value: CommercialScalar) -> str | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return format(value.normalize(), "f")
    if isinstance(value, bool):
        return "true" if value else "false"
    return _semantic_text(str(value))


def _deduplicated_conflicts(
    *groups: tuple[dict[str, object], ...],
) -> tuple[dict[str, object], ...]:
    unique: dict[str, dict[str, object]] = {}
    for conflict in (item for group in groups for item in group):
        signature = repr(sorted(conflict.items()))
        unique.setdefault(signature, conflict)
    return tuple(unique[key] for key in sorted(unique))


def _first_scalar(
    records: list[CommercialWorkbookRecord], field: str
) -> CommercialScalar:
    for record in records:
        value = getattr(record, field)
        if value is not None:
            return value
    return None


def _first_text(records: list[CommercialWorkbookRecord], field: str) -> str | None:
    for record in records:
        value = getattr(record, field)
        if isinstance(value, str) and value.strip() and value != "-":
            return value
    return None


def _joined_text(records: list[CommercialWorkbookRecord], field: str) -> str | None:
    values: list[str] = []
    for record in records:
        value = getattr(record, field)
        if isinstance(value, str) and value.strip() and value not in values:
            values.append(value)
    return "\n".join(values) or None


def _merge_records(
    price: CommercialWorkbookRecord | None,
    supplement: CommercialWorkbookRecord | None,
) -> CommercialWorkbookRecord:
    if price is None and supplement is None:
        raise ValueError("Cannot merge empty commercial records")
    base = price or supplement
    assert base is not None
    if supplement is None:
        return base
    if price is None:
        return supplement
    return replace(
        price,
        service_name=price.service_name or supplement.service_name,
        service_category=price.service_category or supplement.service_category,
        catalog_section=price.catalog_section or supplement.catalog_section,
        product_family=price.product_family or supplement.product_family,
        product_group=price.product_group or supplement.product_group,
        product_paths=_deduplicated_product_paths([price, supplement]),
        metric=price.metric or supplement.metric,
        included_entitlements=supplement.included_entitlements,
        prerequisites=supplement.prerequisites,
        source_evidence=price.source_evidence + supplement.source_evidence,
        source_conflicts=_deduplicated_conflicts(
            price.source_conflicts,
            supplement.source_conflicts,
        ),
    )


def _commercial_price_terms(
    row: tuple[Cell, ...], header: dict[str, int]
) -> tuple[CommercialPriceTerm, ...]:
    terms: list[CommercialPriceTerm] = []
    for field, term_type in (
        ("pay_as_you_go", "pay_as_you_go"),
        ("annual_commitment", "annual_commitment"),
        ("annual_flex", "annual_flex"),
        ("monthly_flex", "monthly_flex"),
    ):
        index = header.get(field)
        if index is None or index >= len(row):
            continue
        value = _commercial_scalar(row[index].value)
        if value is None:
            continue
        terms.append(
            CommercialPriceTerm(
                term_type=term_type,
                value=value,
                source_cell=row[index].coordinate,
                source_label=_service_name(row, header),
                value_state=_value_state(value),
            )
        )
    for field, term_type in (
        ("annual_commitment_minimum", "annual_commitment_minimum"),
        ("monthly_commitment_minimum", "monthly_commitment_minimum"),
    ):
        index = header.get(field)
        if index is None or index >= len(row):
            continue
        value = _commercial_scalar(row[index].value)
        if value is None:
            continue
        label = _text_at(row, header.get("commitment_details"))
        resolved_term_type = term_type
        if label and "annual flex" in label.casefold():
            resolved_term_type = term_type.replace("commitment", "flex")
        terms.append(
            CommercialPriceTerm(
                term_type=resolved_term_type,
                value=value,
                source_cell=row[index].coordinate,
                source_label=label,
                value_state=_value_state(value),
            )
        )
    return tuple(terms)


def _price_score(record: CommercialWorkbookRecord) -> int:
    fields = (
        record.pay_as_you_go,
        record.annual_commitment,
        record.commercial_price_terms,
        record.metric,
        record.metric_minimum,
        record.additional_information,
        record.notes,
        record.service_category,
    )
    return sum(value is not None for value in fields)


def _supplement_score(record: CommercialWorkbookRecord) -> int:
    fields = (
        record.metric,
        record.included_entitlements,
        record.prerequisites,
        record.service_category,
    )
    return sum(value is not None for value in fields)


def _identity_rank(record: CommercialWorkbookRecord) -> tuple[object, ...]:
    """Choose a stable semantic placement independent of workbook page order."""

    hierarchy = record.product_hierarchy
    normalized_hierarchy = tuple(_header_text(value) for value in hierarchy)
    service_tokens = _identity_tokens(record.service_name)
    hierarchy_tokens = set().union(*(_identity_tokens(value) for value in hierarchy)) if hierarchy else set()
    semantic_overlap = len(service_tokens & hierarchy_tokens)
    return (
        -semantic_overlap,
        -len(hierarchy),
        normalized_hierarchy,
        _header_text(record.service_name),
    )


def _identity_tokens(value: str | None) -> set[str]:
    """Return stable product-identity tokens without workbook footnote noise."""

    ignored = {
        "and",
        "as",
        "cloud",
        "for",
        "in",
        "of",
        "on",
        "oracle",
        "service",
        "services",
        "the",
    }
    return {
        token
        for token in re.findall(r"[a-z0-9]+", _header_text(value))
        if len(token) > 1 and token not in ignored
    }


def _evidence_key(record: CommercialWorkbookRecord) -> tuple[str, int]:
    evidence = record.source_evidence[0]
    return evidence.sheet, evidence.row


def _evidence(sheet: str, row: tuple[Cell, ...], header: dict[str, int]) -> SourceEvidence:
    governed_columns = set(header.values())
    cells = tuple(
        SourceCell(cell.coordinate, cell.value, _value_state(cell.value))
        for index, cell in enumerate(row)
        if index in governed_columns or not _is_blank(cell.value)
    )
    row_number = next((cell.row for cell in row if isinstance(cell.row, int)), 0)
    return SourceEvidence(sheet=sheet, row=row_number, cells=cells)


def _part_number(value: object) -> str | None:
    text = _clean_text(value)
    if text is None:
        return None
    candidate = text.replace(" ", "").upper()
    return candidate if _PART_NUMBER_PATTERN.fullmatch(candidate) else None


def _commercial_scalar(value: object) -> CommercialScalar:
    if _is_blank(value):
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))

    text = _clean_text(value)
    if text is None:
        return None
    if text == "-":
        return "-"
    normalized = text.casefold()
    if normalized == "always free":
        return "Always Free"
    if normalized in {"true", "yes"}:
        return True
    if normalized in {"false", "no"}:
        return False
    numeric = text.replace(",", "").removeprefix("$")
    if _NUMERIC_PATTERN.fullmatch(numeric):
        try:
            return Decimal(numeric)
        except InvalidOperation:
            pass
    return text


def _text_at(row: tuple[Cell, ...], index: int | None) -> str | None:
    return _clean_text(_cell_value(row, index))


def _cell_value(row: tuple[Cell, ...], index: int | None) -> object:
    if index is None or index < 0 or index >= len(row):
        return None
    return row[index].value


def _header_text(value: object) -> str:
    text = _clean_text(value)
    if text is None:
        return ""
    return re.sub(r"[^a-z0-9]+", " ", text.casefold()).strip()


def _clean_text(value: object) -> str | None:
    if _is_blank(value):
        return None
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    lines = [" ".join(line.split()) for line in text.split("\n")]
    cleaned = "\n".join(line for line in lines if line).strip()
    return cleaned or None


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _value_state(value: object) -> str:
    if _is_blank(value):
        return "blank"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (Decimal, int, float)):
        return "numeric"
    text = _clean_text(value)
    if text == "-":
        return "not_applicable"
    if text and text.casefold() == "free tier":
        return "free_tier"
    if text and text.casefold() == "always free":
        return "always_free"
    return "text"
