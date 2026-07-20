"""Workbook import service that persists source and catalog rows."""

from __future__ import annotations

import json
import re
import uuid
from datetime import UTC, datetime
from io import BytesIO
from typing import Any, Literal, cast
from unicodedata import normalize

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from app.core.calc_engine import (
    HEADER_ALIASES,
    detect_header_row,
    evaluate_qa,
    executions_per_day,
    normalize_payload_to_kb,
    parse_rows,
    payload_per_hour_kb,
)
from app.models import (
    CatalogIntegration,
    ImportBatch,
    ImportMappingProfile,
    JustificationRecord,
    Project,
    SourceIntegrationRow,
)
from app.models.project import ImportStatus
from app.schemas.imports import (
    ImportBatchDeleteResponse,
    ImportBatchListResponse,
    ImportBatchResponse,
    ImportMappingProfileListResponse,
    ImportMappingProfileResponse,
    ImportMappingReviewApproveRequest,
    ImportMappingReviewUpdateRequest,
    ImportQualityAssistantResponse,
    ImportQualityFinding,
    ImportQualityMetric,
    NormalizationEventResponse,
    SourceRowListResponse,
    SourceRowResponse,
)
from app.services import audit_service, import_mapping_service, recalc_service, storage_service
from app.services.capture_template_service import (
    CAPTURE_SHEET_NAME,
    COLUMNS as TEMPLATE_COLUMNS,
    IMPORTER_MIN_VERSION,
    MANIFEST_SHEET_NAME,
    TEMPLATE_VERSION,
)
from app.services.pattern_support import support_reason_code
from app.services.serializers import parse_bool, parse_float, parse_int, parse_text, sanitize_for_json

SOURCE_SHEET_NAME = CAPTURE_SHEET_NAME
LEGACY_SOURCE_SHEET_NAMES = ("Catálogo de Integraciones", "TPL - Catálogo")
RAW_HEADERS_METADATA_KEY = "__raw_headers__"
TEMPLATE_VERSION_METADATA_KEY = "__template_version__"
TEMPLATE_COMPATIBILITY_METADATA_KEY = "__template_compatibility__"
TEMPLATE_GENERATED_AT_METADATA_KEY = "__template_generated_at__"
FALLBACK_COLUMN_INDEXES: dict[str, int] = {
    "seq_number": 1,
    "interface_id": 2,
    "owner": 3,
    "brand": 4,
    "business_process": 5,
    "interface_name": 6,
    "description": 7,
    "status": 8,
    "mapping_status": 9,
    "initial_scope": 10,
    "complexity": 11,
    "frequency": 12,
    "type": 13,
    "base": 14,
    "interface_status": 15,
    "is_real_time": 16,
    "trigger_type": 17,
    "response_size_kb": 18,
    "payload_per_execution_kb": 19,
    "is_fan_out": 20,
    "fan_out_targets": 21,
    "source_system": 22,
    "source_technology": 23,
    "source_api_reference": 24,
    "source_owner": 25,
    "destination_system": 26,
    "destination_technology_1": 27,
    "destination_technology_2": 28,
    "destination_owner": 29,
    "calendarization": 30,
    "selected_pattern": 33,
    "pattern_rationale": 34,
    "comments": 35,
    "retry_policy": 36,
    "core_tools": 37,
    "additional_tools_overlays": 38,
    "qa_status": 39,
}
_MISSING = object()
_PAYLOAD_WITH_UNIT_PATTERN = re.compile(
    r"^\s*(?P<value>-?\d+(?:[.,]\d+)?)\s*(?P<unit>kb|mb)\s*$",
    re.IGNORECASE,
)


def _normalize_header_key(value: str) -> str:
    collapsed = " ".join(value.strip().lower().replace("\n", " ").split())
    return normalize("NFKD", collapsed).encode("ascii", "ignore").decode("ascii")


def _header_alias_matches(alias: str, header: str) -> bool:
    if header == alias:
        return True
    return (
        header.startswith(f"{alias} ")
        or header.startswith(f"{alias}(")
        or header.startswith(f"{alias}:")
        or header.startswith(f"{alias}-")
    )


def _bad_request(detail: str, error_code: str) -> HTTPException:
    return HTTPException(status_code=400, detail={"detail": detail, "error_code": error_code})


def _read_template_manifest(workbook: Any) -> dict[str, str]:
    """Read the hidden template manifest without depending on fixed cell coordinates."""

    if MANIFEST_SHEET_NAME not in workbook.sheetnames:
        return {}
    sheet = workbook[MANIFEST_SHEET_NAME]
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value != "MANIFEST_KEY":
                continue
            result: dict[str, str] = {}
            for manifest_row in range(cell.row + 1, sheet.max_row + 1):
                key = sheet.cell(manifest_row, cell.column).value
                value = sheet.cell(manifest_row, cell.column + 1).value
                if key in (None, ""):
                    break
                result[str(key)] = "" if value is None else str(value)
            return result
    return {}


def _validate_template_version(manifest: dict[str, str]) -> str:
    """Return a compatibility label or reject a workbook from a future major contract."""

    version = manifest.get("template_version")
    if not version:
        return "legacy_v1_accepted"
    try:
        major = int(version.split(".", 1)[0])
        current_major = int(TEMPLATE_VERSION.split(".", 1)[0])
    except ValueError as exc:
        raise _bad_request(
            f"Template version '{version}' is not valid. Download a fresh template from the App.",
            "IMPORT_TEMPLATE_VERSION_INVALID",
        ) from exc
    if major > current_major:
        raise _bad_request(
            f"Template v{version} requires a newer importer. This environment supports through v{TEMPLATE_VERSION}.",
            "IMPORT_TEMPLATE_VERSION_UNSUPPORTED",
        )
    return "current" if version == TEMPLATE_VERSION else "older_supported"


def _assert_no_capture_formulas(sheet: Any) -> None:
    """Reject formulas in capture cells to prevent hidden logic and formula injection."""

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
                raise _bad_request(
                    f"Formula found in {SOURCE_SHEET_NAME}!{cell.coordinate}. Replace it with its final value before upload.",
                    "IMPORT_FORMULA_NOT_ALLOWED",
                )


def _formula_classification(header: str) -> tuple[str, str]:
    """Classify external formula evidence without executing or trusting it."""

    normalized = _normalize_header_key(header)
    if "cost" in normalized or "costo" in normalized or "price" in normalized or "precio" in normalized:
        return (
            "commercial_evidence",
            "Client cost logic is retained for comparison only; governed App pricing remains authoritative.",
        )
    if "mensaje" in normalized or "message" in normalized or "execution" in normalized or "ejecucion" in normalized:
        return (
            "derived_demand",
            "The result is derived demand. Map the underlying payload, frequency, and fan-out inputs instead.",
        )
    return (
        "needs_review",
        "The formula is preserved as source evidence, but its result cannot become an operational App value.",
    )


def _collect_external_formula_evidence(
    formula_sheet: Any,
    value_sheet: Any,
    header_row_index: int,
    raw_headers: dict[str, str],
) -> tuple[dict[int, list[dict[str, object]]], list[dict[str, object]]]:
    """Capture formula text and cached display values without evaluating formulas."""

    by_row: dict[int, list[dict[str, object]]] = {}
    by_column: dict[str, dict[str, object]] = {}
    first_data_row = header_row_index + 2
    formula_rows = formula_sheet.iter_rows(min_row=first_data_row)
    value_rows = value_sheet.iter_rows(min_row=first_data_row)
    for formula_row, value_row in zip(formula_rows, value_rows, strict=False):
        for formula_cell, value_cell in zip(formula_row, value_row, strict=False):
            formula = formula_cell.value
            if formula_cell.data_type != "f" and not (
                isinstance(formula, str) and formula.startswith("=")
            ):
                continue
            source_index = str(formula_cell.column - 1)
            source_header = raw_headers.get(source_index, f"Column {formula_cell.column}")
            cached_value = sanitize_for_json(value_cell.value)
            classification, rationale = _formula_classification(source_header)
            is_error = isinstance(cached_value, str) and cached_value.startswith("#")
            evidence = {
                "coordinate": formula_cell.coordinate,
                "source_index": source_index,
                "source_header": source_header,
                "formula": str(formula),
                "cached_value": cached_value,
                "cached_value_status": "error" if is_error else ("missing" if cached_value is None else "available"),
                "classification": classification,
                "rationale": rationale,
            }
            by_row.setdefault(formula_cell.row, []).append(evidence)
            column_policy = (
                "evidence_only"
                if classification in {"commercial_evidence", "derived_demand"}
                else "formula_rows_only"
            )
            summary = by_column.setdefault(
                source_header,
                {
                    "source_index": source_index,
                    "source_header": source_header,
                    "classification": classification,
                    "rationale": rationale,
                    "formula_count": 0,
                    "cached_value_count": 0,
                    "cached_error_count": 0,
                    "sample_formulas": [],
                    "sample_cached_values": [],
                    "operational_policy": column_policy,
                },
            )
            summary["formula_count"] = cast(int, summary["formula_count"]) + 1
            if is_error:
                summary["cached_error_count"] = cast(int, summary["cached_error_count"]) + 1
            elif cached_value is not None:
                summary["cached_value_count"] = cast(int, summary["cached_value_count"]) + 1
            sample_formulas = cast(list[str], summary["sample_formulas"])
            if str(formula) not in sample_formulas and len(sample_formulas) < 2:
                sample_formulas.append(str(formula)[:240])
            sample_values = cast(list[object], summary["sample_cached_values"])
            if cached_value not in (None, "") and cached_value not in sample_values and len(sample_values) < 3:
                sample_values.append(cached_value)
    return by_row, sorted(by_column.values(), key=lambda item: int(str(item["source_index"])))


def _validate_current_headers(all_rows: list[list], manifest: dict[str, str]) -> None:
    """Keep the current capture contract exact while preserving older workbooks."""

    if manifest.get("template_version") != TEMPLATE_VERSION:
        return
    header_index = detect_header_row(all_rows)
    actual_headers = [str(value).strip() if value is not None else "" for value in all_rows[header_index]]
    expected_headers = [column.header for column in TEMPLATE_COLUMNS]
    if actual_headers[: len(expected_headers)] != expected_headers:
        raise _bad_request(
            f"The governed v{TEMPLATE_VERSION} capture headers were renamed, removed, or reordered. Download a fresh template and paste values only into its capture rows.",
            "IMPORT_TEMPLATE_HEADERS_CHANGED",
        )


def _select_import_sheet(workbook: Any, manifest: dict[str, str]) -> tuple[str, list[list]]:
    """Return the governed capture sheet or the strongest non-template evidence sheet.

    A manifest-bearing workbook is a governed contract and must retain its declared
    capture sheet.  For external evidence, a client is not required to know our
    sheet name: select the non-manifest sheet with the most recognized headers,
    then the most non-empty rows.  This only identifies a source tab; it never
    accepts a semantic field mapping.
    """

    official_names = (SOURCE_SHEET_NAME, *LEGACY_SOURCE_SHEET_NAMES)
    if manifest:
        source_sheet_name = next((name for name in official_names if name in workbook.sheetnames), None)
        if source_sheet_name is None:
            raise _bad_request(
                f"Required sheet '{SOURCE_SHEET_NAME}' not found in governed workbook.",
                "IMPORT_SHEET_NOT_FOUND",
            )
        sheet = workbook[source_sheet_name]
        return source_sheet_name, [list(row) for row in sheet.iter_rows(values_only=True)]

    candidates: list[tuple[int, int, int, str, list[list]]] = []
    for position, name in enumerate(workbook.sheetnames):
        if name == MANIFEST_SHEET_NAME:
            continue
        sheet = workbook[name]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        if not any(any(value not in (None, "") for value in row) for row in rows):
            continue
        result = parse_rows(rows)
        candidates.append((len(result.header_map), result.source_row_count, -position, name, rows))

    if not candidates:
        raise _bad_request(
            "No populated worksheet was found for import.",
            "IMPORT_SHEET_NOT_FOUND",
        )
    _, _, _, source_sheet_name, rows = max(candidates)
    return source_sheet_name, rows


def serialize_batch(batch: ImportBatch) -> ImportBatchResponse:
    """Convert an import batch model into a response schema."""

    return ImportBatchResponse(
        id=batch.id,
        project_id=batch.project_id,
        filename=batch.filename,
        parser_version=batch.parser_version,
        status=batch.status.value,
        source_row_count=batch.source_row_count,
        candidate_count=batch.candidate_count,
        tbq_y_count=batch.tbq_y_count,
        tbq_n_count=batch.tbq_n_count,
        excluded_count=batch.excluded_count,
        loaded_count=batch.loaded_count,
        header_map=batch.header_map,
        error_details=batch.error_details,
        intake_mode=cast(Literal["official_template", "external_mapping"], batch.intake_mode),
        mapping_contract=batch.mapping_contract,
        mapping_profile_id=batch.mapping_profile_id,
        mapping_reviewed_by=batch.mapping_reviewed_by,
        mapping_reviewed_at=batch.mapping_reviewed_at,
        created_at=batch.created_at,
        updated_at=batch.updated_at,
    )


def serialize_source_row(row: SourceIntegrationRow) -> SourceRowResponse:
    """Convert a source row model into a response schema."""

    return SourceRowResponse(
        id=row.id,
        source_row_number=row.source_row_number,
        included=row.included,
        exclusion_reason=row.exclusion_reason,
        raw_data=row.raw_data,
        normalization_events=[
            NormalizationEventResponse(**event) for event in (row.normalization_events or [])
        ],
    )


def _extract_raw_headers(header_map: dict[str, str] | None) -> dict[str, str]:
    if not header_map:
        return {}
    encoded_headers = header_map.get(RAW_HEADERS_METADATA_KEY)
    if not encoded_headers:
        return {}
    try:
        payload = json.loads(encoded_headers)
    except json.JSONDecodeError:
        return {}
    if not isinstance(payload, dict):
        return {}
    return {str(key): str(value) for key, value in payload.items()}


def _header_label_for_field(header_map: dict[str, str], field: str) -> str | None:
    raw_headers = _extract_raw_headers(header_map)
    column_index = header_map.get(field)
    if column_index is None:
        return None
    return raw_headers.get(column_index)


def _normalize_raw_header_label(value: object, index: int, used_labels: set[str]) -> str:
    base_label = str(value).strip() if value is not None else ""
    if not base_label:
        base_label = f"Column {index + 1}"

    label = base_label
    duplicate_number = 2
    while label in used_labels:
        label = f"{base_label} ({duplicate_number})"
        duplicate_number += 1
    used_labels.add(label)
    return label


def _build_raw_header_labels(raw_headers: list[object]) -> dict[str, str]:
    used_labels: set[str] = set()
    return {
        str(index): _normalize_raw_header_label(value, index, used_labels)
        for index, value in enumerate(raw_headers)
    }


def _build_raw_column_values(raw_data: dict[str, object], raw_headers: dict[str, str]) -> dict[str, object]:
    values: dict[str, object] = {}
    for raw_key, raw_value in raw_data.items():
        if raw_key in raw_headers:
            values[raw_headers[raw_key]] = raw_value
            continue
        if raw_key.isdigit():
            values[f"Column {int(raw_key) + 1}"] = raw_value
            continue
        values[raw_key] = raw_value
    return values


def _is_external_summary_row(raw_data: dict[str, object], header_map: dict[str, str]) -> bool:
    """Identify explicit client workbook totals without guessing about business rows."""

    sequence = _normalize_header_key(
        str(_extract_raw_value(raw_data, header_map, "seq_number") or "")
    )
    if sequence not in {"total", "subtotal", "grand total"}:
        return False
    identifying_values = (
        _extract_raw_value(raw_data, header_map, "interface_name"),
        _extract_raw_value(raw_data, header_map, "source_system"),
        _extract_raw_value(raw_data, header_map, "destination_system"),
    )
    return not any(value not in (None, "") for value in identifying_values)


def _extract_value_from_header_alias(raw_data: dict[str, object], field: str) -> object:
    aliases = HEADER_ALIASES.get(field, [])
    if not aliases:
        return _MISSING

    normalized_key_map = {
        _normalize_header_key(str(raw_key)): raw_key
        for raw_key in raw_data
        if isinstance(raw_key, str)
    }

    for alias in aliases:
        normalized_alias = _normalize_header_key(alias)
        direct_key = normalized_key_map.get(normalized_alias)
        if direct_key is not None:
            return raw_data[direct_key]
        for normalized_header, raw_key in normalized_key_map.items():
            if _header_alias_matches(normalized_alias, normalized_header):
                return raw_data[raw_key]

    return _MISSING


def _extract_raw_value(raw_data: dict[str, object], header_map: dict[str, str], field: str) -> object:
    if field in raw_data:
        return raw_data.get(field)

    contract_enforced = header_map.get(import_mapping_service.CONTRACT_ENFORCED_METADATA_KEY) == "true"
    raw_headers = _extract_raw_headers(header_map)
    column_index = header_map.get(field)
    if contract_enforced:
        if column_index is None or column_index == "-1":
            return None
        header_label = raw_headers.get(column_index)
        if header_label and header_label in raw_data:
            return raw_data.get(header_label)
        return raw_data.get(column_index)

    alias_value = _extract_value_from_header_alias(raw_data, field)
    if alias_value is not _MISSING:
        return alias_value

    if column_index is None:
        fallback_index = FALLBACK_COLUMN_INDEXES.get(field)
        if fallback_index is None:
            return None
        column_index = str(fallback_index)
    if column_index == "-1":
        return None
    header_label = raw_headers.get(column_index)
    if header_label and header_label in raw_data:
        return raw_data.get(header_label)
    if column_index in raw_data:
        return raw_data.get(column_index)
    if column_index.isdigit():
        fallback_label = f"Column {int(column_index) + 1}"
        if fallback_label in raw_data:
            return raw_data.get(fallback_label)
    return raw_data.get(column_index)


def _parse_payload_measurement(raw_value: object) -> tuple[float | None, str | None]:
    if raw_value in (None, ""):
        return None, None
    if isinstance(raw_value, (int, float)):
        return float(raw_value), None

    text = str(raw_value).strip()
    match = _PAYLOAD_WITH_UNIT_PATTERN.match(text)
    if match:
        numeric = match.group("value").replace(",", ".")
        return float(numeric), match.group("unit").upper()
    return parse_float(text), None


def _infer_payload_unit(raw_value: object, header_label: str | None) -> str:
    _, explicit_unit = _parse_payload_measurement(raw_value)
    if explicit_unit in {"KB", "MB"}:
        return explicit_unit

    normalized_header = _normalize_header_key(header_label or "")
    if "mb" in normalized_header or "mega" in normalized_header:
        return "MB"
    return "KB"


def _normalized_payload_value(
    raw_data: dict[str, object],
    header_map: dict[str, str],
) -> tuple[float | None, dict[str, object] | None]:
    raw_value = _extract_raw_value(raw_data, header_map, "payload_per_execution_kb")
    parsed_value, _ = _parse_payload_measurement(raw_value)
    if parsed_value is None:
        return None, None

    unit = _infer_payload_unit(raw_value, _header_label_for_field(header_map, "payload_per_execution_kb"))
    normalized = (
        normalize_payload_to_kb(parsed_value, "MB").value
        if unit == "MB"
        else normalize_payload_to_kb(parsed_value, "KB").value
    )
    if normalized is None:
        return None, None

    if unit == "MB":
        return normalized, {
            "field": "payload_per_execution_kb",
            "old_value": raw_value,
            "new_value": normalized,
            "rule": "payload_unit_mb_to_kb",
        }
    return normalized, None


def _split_destination_technologies(raw_value: str | None, secondary_value: str | None) -> tuple[str | None, str | None]:
    primary = raw_value.strip() if raw_value else None
    secondary = secondary_value.strip() if secondary_value else None
    if not primary or secondary:
        return primary or None, secondary or None

    separators = ["/", "|", ";", "\n", ","]
    for separator in separators:
        if separator not in primary:
            continue
        parts = [part.strip() for part in re.split(rf"\s*{re.escape(separator)}\s*", primary) if part.strip()]
        if len(parts) == 2:
            return parts[0], parts[1]

    return primary, None


def _normalized_frequency(raw_data: dict[str, object], events: list[dict[str, object]], header_map: dict[str, str]) -> str | None:
    for event in events:
        if event.get("field") == "frequency":
            return parse_text(event.get("new_value"))
    return parse_text(_extract_raw_value(raw_data, header_map, "frequency"))


def _build_catalog_integration(
    project_id: str,
    source_row_id: str,
    raw_data: dict[str, object],
    normalization_events: list[dict[str, object]],
    header_map: dict[str, str],
) -> CatalogIntegration:
    frequency_value = _normalized_frequency(raw_data, normalization_events, header_map)
    payload_value, _ = _normalized_payload_value(raw_data, header_map)
    destination_technology_1, destination_technology_2 = _split_destination_technologies(
        parse_text(_extract_raw_value(raw_data, header_map, "destination_technology_1")),
        parse_text(_extract_raw_value(raw_data, header_map, "destination_technology_2")),
    )
    execs_result = executions_per_day(frequency_value) if frequency_value else None
    execs_per_day = execs_result.value if execs_result is not None else None
    payload_hour_result = (
        payload_per_hour_kb(payload_value, execs_per_day)
        if payload_value is not None and execs_per_day is not None
        else None
    )
    qa_result = evaluate_qa(
        interface_id=parse_text(_extract_raw_value(raw_data, header_map, "interface_id")),
        trigger_type=parse_text(_extract_raw_value(raw_data, header_map, "trigger_type")),
        selected_pattern=parse_text(_extract_raw_value(raw_data, header_map, "selected_pattern")),
        pattern_rationale=parse_text(_extract_raw_value(raw_data, header_map, "pattern_rationale")),
        core_tools=parse_text(_extract_raw_value(raw_data, header_map, "core_tools")),
        payload_per_execution_kb=payload_value,
        is_fan_out=parse_bool(_extract_raw_value(raw_data, header_map, "is_fan_out")),
        fan_out_targets=parse_int(_extract_raw_value(raw_data, header_map, "fan_out_targets")),
        is_active_row=True,
        retry_policy=parse_text(_extract_raw_value(raw_data, header_map, "retry_policy")),
        idempotency=parse_text(_extract_raw_value(raw_data, header_map, "idempotency")),
        target_latency_sla=parse_text(_extract_raw_value(raw_data, header_map, "target_latency_sla")),
        data_security_classification=parse_text(
            _extract_raw_value(raw_data, header_map, "data_security_classification")
        ),
        retention_processing_window=parse_text(
            _extract_raw_value(raw_data, header_map, "retention_processing_window")
        ),
        business_criticality=parse_text(
            _extract_raw_value(raw_data, header_map, "business_criticality")
        ),
        additional_tools_overlays=parse_text(
            _extract_raw_value(raw_data, header_map, "additional_tools_overlays")
        ),
    )
    qa_reasons = list(qa_result.reasons)
    support_reason = support_reason_code(
        parse_text(_extract_raw_value(raw_data, header_map, "selected_pattern"))
    )
    if support_reason and support_reason not in qa_reasons:
        qa_reasons.append(support_reason)
    return CatalogIntegration(
        project_id=project_id,
        source_row_id=source_row_id,
        tbq=(
            "Y"
            if (parse_text(_extract_raw_value(raw_data, header_map, "tbq")) or "").strip().upper() == "Y"
            else "N"
        ),
        seq_number=parse_int(_extract_raw_value(raw_data, header_map, "seq_number")) or 0,
        interface_id=parse_text(_extract_raw_value(raw_data, header_map, "interface_id")),
        owner=parse_text(_extract_raw_value(raw_data, header_map, "owner")),
        brand=parse_text(_extract_raw_value(raw_data, header_map, "brand")),
        business_process=parse_text(_extract_raw_value(raw_data, header_map, "business_process")),
        interface_name=parse_text(_extract_raw_value(raw_data, header_map, "interface_name")),
        description=parse_text(_extract_raw_value(raw_data, header_map, "description")),
        business_criticality=parse_text(_extract_raw_value(raw_data, header_map, "business_criticality")),
        status=parse_text(_extract_raw_value(raw_data, header_map, "status")),
        mapping_status=parse_text(_extract_raw_value(raw_data, header_map, "mapping_status")),
        initial_scope=parse_text(_extract_raw_value(raw_data, header_map, "initial_scope")),
        complexity=parse_text(_extract_raw_value(raw_data, header_map, "complexity")),
        frequency=frequency_value,
        type=parse_text(_extract_raw_value(raw_data, header_map, "type")),
        base=parse_text(_extract_raw_value(raw_data, header_map, "base")),
        interface_status=parse_text(_extract_raw_value(raw_data, header_map, "interface_status")),
        is_real_time=parse_bool(_extract_raw_value(raw_data, header_map, "is_real_time")),
        target_latency_sla=parse_text(_extract_raw_value(raw_data, header_map, "target_latency_sla")),
        trigger_type=parse_text(_extract_raw_value(raw_data, header_map, "trigger_type")),
        response_size_kb=parse_float(_extract_raw_value(raw_data, header_map, "response_size_kb")),
        payload_per_execution_kb=payload_value,
        is_fan_out=parse_bool(_extract_raw_value(raw_data, header_map, "is_fan_out")),
        fan_out_targets=parse_int(_extract_raw_value(raw_data, header_map, "fan_out_targets")),
        source_system=parse_text(_extract_raw_value(raw_data, header_map, "source_system")),
        source_technology=parse_text(_extract_raw_value(raw_data, header_map, "source_technology")),
        source_api_reference=parse_text(_extract_raw_value(raw_data, header_map, "source_api_reference")),
        source_owner=parse_text(_extract_raw_value(raw_data, header_map, "source_owner")),
        destination_system=parse_text(_extract_raw_value(raw_data, header_map, "destination_system")),
        destination_technology_1=destination_technology_1,
        destination_technology_2=destination_technology_2,
        destination_owner=parse_text(_extract_raw_value(raw_data, header_map, "destination_owner")),
        data_security_classification=parse_text(
            _extract_raw_value(raw_data, header_map, "data_security_classification")
        ),
        executions_per_day=execs_per_day,
        payload_per_hour_kb=payload_hour_result.value if payload_hour_result else None,
        selected_pattern=parse_text(_extract_raw_value(raw_data, header_map, "selected_pattern")),
        pattern_rationale=parse_text(_extract_raw_value(raw_data, header_map, "pattern_rationale")),
        comments=parse_text(_extract_raw_value(raw_data, header_map, "comments")),
        retry_policy=parse_text(_extract_raw_value(raw_data, header_map, "retry_policy")),
        idempotency=parse_text(_extract_raw_value(raw_data, header_map, "idempotency")),
        core_tools=parse_text(_extract_raw_value(raw_data, header_map, "core_tools")),
        additional_tools_overlays=parse_text(
            _extract_raw_value(raw_data, header_map, "additional_tools_overlays")
        ),
        qa_status="OK" if not qa_reasons else "REVISAR",
        qa_reasons=qa_reasons,
        calendarization=parse_text(_extract_raw_value(raw_data, header_map, "calendarization")),
        retention_processing_window=parse_text(
            _extract_raw_value(raw_data, header_map, "retention_processing_window")
        ),
    )


async def create_import_batch(
    project_id: str,
    filename: str,
    db: AsyncSession,
    storage_reference: str | None = None,
) -> ImportBatch:
    """Create a pending import batch."""

    project = await db.get(Project, project_id)
    if project is None:
        raise HTTPException(
            status_code=404,
            detail={"detail": "Project not found", "error_code": "PROJECT_NOT_FOUND"},
        )
    batch = ImportBatch(
        project_id=project_id,
        filename=filename,
        storage_reference=storage_reference,
        parser_version=IMPORTER_MIN_VERSION,
        status=ImportStatus.PENDING,
    )
    db.add(batch)
    await db.flush()
    await db.refresh(batch)
    return batch


def _is_official_template(
    source_sheet_name: str,
    manifest: dict[str, str],
    header_map: dict[str, str],
    *,
    has_formulas: bool = False,
) -> bool:
    """Return whether a workbook matches a current or documented legacy template contract."""

    if source_sheet_name == SOURCE_SHEET_NAME and bool(manifest):
        return True
    # v1 templates had no manifest but used a legacy, localized sheet name and
    # a sufficiently complete governed header set. Modern unmanifested sheets
    # remain external evidence and must enter mapping review.
    return (
        source_sheet_name in LEGACY_SOURCE_SHEET_NAMES
        and not manifest
        and not has_formulas
        and len(header_map) >= 6
    )


async def _materialize_batch(
    batch: ImportBatch,
    db: AsyncSession,
    *,
    actor_id: str,
    correlation_id: str,
) -> int:
    """Create catalog records from persisted source rows exactly once."""

    source_rows = (
        await db.scalars(
            select(SourceIntegrationRow)
            .where(SourceIntegrationRow.import_batch_id == batch.id)
            .order_by(SourceIntegrationRow.source_row_number)
        )
    ).all()
    source_row_ids = [row.id for row in source_rows]
    if source_row_ids:
        existing_count = await db.scalar(
            select(func.count()).select_from(CatalogIntegration).where(
                CatalogIntegration.source_row_id.in_(source_row_ids)
            )
        )
        if existing_count:
            batch.loaded_count = int(existing_count)
            batch.status = ImportStatus.COMPLETED
            return int(existing_count)

    header_map = cast(dict[str, str], batch.header_map or {})
    contract = cast(dict[str, object], batch.mapping_contract or {})
    external = batch.intake_mode == "external_mapping"
    effective_header_map = (
        import_mapping_service.approved_header_map(header_map, contract)
        if external
        else header_map
    )
    loaded = 0
    for source_row in source_rows:
        if not source_row.included:
            continue
        raw_data = cast(dict[str, object], source_row.raw_data or {})
        working_data = (
            import_mapping_service.apply_contract_values(raw_data, contract)
            if external
            else raw_data
        )
        if external:
            formula_headers = {
                str(event.get("field", ""))
                for event in cast(list[dict[str, object]], source_row.normalization_events or [])
                if event.get("rule") == "external_formula_evidence_only"
            }
            for target in import_mapping_service.operational_targets_for_sources(
                contract,
                formula_headers,
            ):
                working_data[target] = None
        db.add(
            _build_catalog_integration(
                project_id=batch.project_id,
                source_row_id=source_row.id,
                raw_data=working_data,
                normalization_events=[] if external else cast(list[dict[str, Any]], source_row.normalization_events or []),
                header_map=effective_header_map,
            )
        )
        loaded += 1

    batch.loaded_count = loaded
    batch.status = ImportStatus.COMPLETED
    await audit_service.emit(
        event_type="import_materialized",
        entity_type="import_batch",
        entity_id=batch.id,
        actor_id=actor_id,
        old_value={"status": ImportStatus.PROCESSING.value, "loaded_count": 0},
        new_value={
            "status": ImportStatus.COMPLETED.value,
            "loaded_count": loaded,
            "intake_mode": batch.intake_mode,
        },
        project_id=batch.project_id,
        correlation_id=correlation_id,
        db=db,
    )
    return loaded


async def process_import(batch_id: str, source_reference: str, db: AsyncSession) -> ImportBatch:
    """Parse a workbook and stage external evidence before catalog materialization."""

    batch = await db.get(ImportBatch, batch_id)
    if batch is None:
        raise HTTPException(
            status_code=404,
            detail={"detail": "Import batch not found", "error_code": "IMPORT_BATCH_NOT_FOUND"},
        )

    batch.status = ImportStatus.PROCESSING
    await db.flush()

    workbook_bytes = storage_service.read_bytes(source_reference)
    workbook = load_workbook(
        filename=BytesIO(workbook_bytes),
        data_only=False,
        read_only=True,
    )
    manifest = _read_template_manifest(workbook)
    compatibility = _validate_template_version(manifest)
    source_sheet_name, all_rows = _select_import_sheet(workbook, manifest)
    sheet = workbook[source_sheet_name]
    _validate_current_headers(all_rows, manifest)
    import_result = parse_rows(all_rows)
    header_row_index = detect_header_row(all_rows)
    raw_header_labels = _build_raw_header_labels(all_rows[header_row_index] if all_rows else [])
    value_workbook = load_workbook(
        filename=BytesIO(workbook_bytes),
        data_only=True,
        read_only=True,
    )
    formula_evidence_by_row, formula_columns = _collect_external_formula_evidence(
        sheet,
        value_workbook[source_sheet_name],
        header_row_index,
        raw_header_labels,
    )
    stored_header_map = {
        **import_result.header_map,
        RAW_HEADERS_METADATA_KEY: json.dumps(raw_header_labels, ensure_ascii=True),
        TEMPLATE_VERSION_METADATA_KEY: manifest.get("template_version", "1.x-unversioned"),
        TEMPLATE_COMPATIBILITY_METADATA_KEY: compatibility,
        TEMPLATE_GENERATED_AT_METADATA_KEY: manifest.get("generated_at_utc", ""),
    }
    correlation_id = str(uuid.uuid4())
    staged_rows: list[tuple[SourceIntegrationRow, bool]] = []
    official_template = _is_official_template(
        source_sheet_name,
        manifest,
        import_result.header_map,
        has_formulas=bool(formula_columns),
    )
    if official_template:
        _assert_no_capture_formulas(sheet)

    for parsed_row in import_result.rows:
        normalization_events = (
            [sanitize_for_json(event.__dict__) for event in parsed_row.normalization_events]
            if official_template
            else []
        )
        if not official_template:
            normalization_events.extend(
                {
                    "field": str(item["source_header"]),
                    "old_value": item["formula"],
                    "new_value": item["cached_value"],
                    "rule": "external_formula_evidence_only",
                }
                for item in formula_evidence_by_row.get(parsed_row.source_row_number, [])
            )
        raw_column_values = _build_raw_column_values(parsed_row.raw_data, raw_header_labels)
        _, payload_event = _normalized_payload_value(raw_column_values, stored_header_map)
        if official_template and payload_event is not None:
            normalization_events.append(sanitize_for_json(payload_event))
        summary_row = not official_template and _is_external_summary_row(
            raw_column_values,
            stored_header_map,
        )
        source_row = SourceIntegrationRow(
            import_batch_id=batch.id,
            source_row_number=parsed_row.source_row_number,
            raw_data=sanitize_for_json(raw_column_values),
            included=parsed_row.included and not summary_row,
            exclusion_reason=(
                "External workbook summary row"
                if summary_row
                else parsed_row.exclusion_reason
            ),
            normalization_events=normalization_events,
        )
        db.add(source_row)
        await db.flush()
        staged_rows.append((source_row, source_row.included))

        for event in normalization_events:
            event_dict = cast(dict[str, Any], event)
            await audit_service.emit(
                event_type="normalization",
                entity_type="source_integration_row",
                entity_id=source_row.id,
                actor_id="system-import",
                old_value={"field": event_dict["field"], "value": event_dict["old_value"]},
                new_value={
                    "field": event_dict["field"],
                    "value": event_dict["new_value"],
                    "rule": event_dict["rule"],
                },
                project_id=batch.project_id,
                db=db,
                correlation_id=correlation_id,
            )

    batch.source_row_count = import_result.source_row_count
    staged_included = [row for row, included in staged_rows if included]
    if official_template:
        batch.candidate_count = import_result.loaded_count
        batch.tbq_y_count = import_result.tbq_y_count
        batch.tbq_n_count = import_result.tbq_n_count
        batch.excluded_count = import_result.excluded_count
    else:
        batch.candidate_count = len(staged_included)
        batch.tbq_y_count = sum(
            1
            for row in staged_included
            if str(
                _extract_raw_value(
                    cast(dict[str, object], row.raw_data or {}),
                    stored_header_map,
                    "tbq",
                )
                or ""
            ).strip().upper()
            == "Y"
        )
        batch.tbq_n_count = len(staged_included) - int(batch.tbq_y_count or 0)
        batch.excluded_count = len(staged_rows) - len(staged_included)
    batch.loaded_count = import_result.loaded_count if official_template else 0
    batch.header_map = stored_header_map
    batch.parser_version = import_result.parser_version
    if official_template:
        batch.intake_mode = "official_template"
        batch.mapping_contract = {
            "version": import_mapping_service.CONTRACT_VERSION,
            "status": "not_required",
            "source_kind": "official_template",
        }
        await _materialize_batch(
            batch,
            db,
            actor_id="system-import",
            correlation_id=correlation_id,
        )
    else:
        raw_rows = [cast(dict[str, object], row.raw_data or {}) for row, _ in staged_rows]
        contract = import_mapping_service.build_mapping_contract(
            raw_header_labels,
            raw_rows,
            formula_columns=formula_columns,
        )
        profile = await db.scalar(
            select(ImportMappingProfile).where(
                ImportMappingProfile.project_id == batch.project_id,
                ImportMappingProfile.header_fingerprint == contract["header_fingerprint"],
                ImportMappingProfile.is_active.is_(True),
            )
        )
        batch.intake_mode = "external_mapping"
        if profile is not None and not formula_columns:
            batch.mapping_contract = profile.contract
            batch.mapping_profile_id = profile.id
            batch.mapping_reviewed_by = profile.created_by
            batch.mapping_reviewed_at = datetime.now(UTC)
            batch.status = ImportStatus.PROCESSING
            await audit_service.emit(
                event_type="import_mapping_profile_applied",
                entity_type="import_batch",
                entity_id=batch.id,
                actor_id="system-import",
                old_value={"status": ImportStatus.PENDING.value},
                new_value={"profile_id": profile.id, "profile_name": profile.name},
                project_id=batch.project_id,
                correlation_id=correlation_id,
                db=db,
            )
            await _materialize_batch(
                batch,
                db,
                actor_id="system-import",
                correlation_id=correlation_id,
            )
        else:
            batch.mapping_contract = contract
            batch.status = ImportStatus.MAPPING_REVIEW
            await audit_service.emit(
                event_type="import_mapping_review_required",
                entity_type="import_batch",
                entity_id=batch.id,
                actor_id="system-import",
                old_value={"status": ImportStatus.PROCESSING.value},
                new_value={
                    "status": ImportStatus.MAPPING_REVIEW.value,
                    "header_count": len(raw_header_labels),
                    "candidate_count": import_result.loaded_count,
                    "question_count": len(import_mapping_service.contract_items(contract, "questions")),
                },
                project_id=batch.project_id,
                correlation_id=correlation_id,
                db=db,
            )
    await db.flush()
    await db.refresh(batch)
    return batch


def _serialize_mapping_profile(profile: ImportMappingProfile) -> ImportMappingProfileResponse:
    return ImportMappingProfileResponse(
        id=profile.id,
        project_id=profile.project_id,
        name=profile.name,
        header_fingerprint=profile.header_fingerprint,
        contract=profile.contract,
        created_by=profile.created_by,
        is_active=profile.is_active,
        created_at=profile.created_at,
        updated_at=profile.updated_at,
    )


async def list_import_mapping_profiles(
    project_id: str,
    db: AsyncSession,
) -> ImportMappingProfileListResponse:
    """List the approved, project-scoped contracts available to future external imports."""

    profiles = (
        await db.scalars(
            select(ImportMappingProfile)
            .where(ImportMappingProfile.project_id == project_id)
            .order_by(ImportMappingProfile.updated_at.desc())
        )
    ).all()
    return ImportMappingProfileListResponse(
        profiles=[_serialize_mapping_profile(profile) for profile in profiles]
    )


async def _load_mapping_review_batch(
    project_id: str,
    batch_id: str,
    db: AsyncSession,
) -> ImportBatch:
    batch = await db.scalar(
        select(ImportBatch).where(
            ImportBatch.project_id == project_id,
            ImportBatch.id == batch_id,
        )
    )
    if batch is None:
        raise HTTPException(
            status_code=404,
            detail={"detail": "Import batch not found", "error_code": "IMPORT_BATCH_NOT_FOUND"},
        )
    if batch.intake_mode != "external_mapping":
        raise HTTPException(
            status_code=409,
            detail={
                "detail": "The governed template does not require mapping review.",
                "error_code": "IMPORT_MAPPING_NOT_REQUIRED",
            },
        )
    return batch


async def update_import_mapping_review(
    project_id: str,
    batch_id: str,
    request: ImportMappingReviewUpdateRequest,
    actor_id: str,
    db: AsyncSession,
) -> ImportBatchResponse:
    """Save a draft mapping decision without releasing source rows downstream."""

    batch = await _load_mapping_review_batch(project_id, batch_id, db)
    if batch.status != ImportStatus.MAPPING_REVIEW:
        raise HTTPException(
            status_code=409,
            detail={"detail": "Only a staged mapping review can be edited.", "error_code": "IMPORT_MAPPING_NOT_STAGED"},
        )
    try:
        contract = import_mapping_service.validate_contract_update(
            cast(dict[str, object], batch.mapping_contract or {}),
            [item.model_dump() for item in request.fields],
            request.answers,
            require_complete=False,
        )
    except ValueError as exc:
        raise _bad_request(str(exc), "IMPORT_MAPPING_INVALID") from exc
    batch.mapping_contract = contract
    await audit_service.emit(
        event_type="import_mapping_review_saved",
        entity_type="import_batch",
        entity_id=batch.id,
        actor_id=actor_id,
        old_value=None,
        new_value={
            "mapped_header_count": len(request.fields),
            "answered_question_count": len(request.answers),
        },
        project_id=project_id,
        db=db,
    )
    await db.flush()
    await db.refresh(batch)
    return serialize_batch(batch)


async def approve_import_mapping_review(
    project_id: str,
    batch_id: str,
    request: ImportMappingReviewApproveRequest,
    actor_id: str,
    db: AsyncSession,
) -> ImportBatchResponse:
    """Approve a complete external mapping contract and queue safe materialization."""

    batch = await _load_mapping_review_batch(project_id, batch_id, db)
    if batch.status != ImportStatus.MAPPING_REVIEW:
        raise HTTPException(
            status_code=409,
            detail={"detail": "Only a staged mapping review can be approved.", "error_code": "IMPORT_MAPPING_NOT_STAGED"},
        )
    try:
        contract = import_mapping_service.validate_contract_update(
            cast(dict[str, object], batch.mapping_contract or {}),
            [item.model_dump() for item in request.fields],
            request.answers,
        )
    except ValueError as exc:
        raise _bad_request(str(exc), "IMPORT_MAPPING_INCOMPLETE") from exc

    profile_id: str | None = None
    if request.save_profile:
        profile = ImportMappingProfile(
            project_id=project_id,
            name=request.profile_name or f"Approved mapping {datetime.now(UTC).date().isoformat()}",
            header_fingerprint=str(contract["header_fingerprint"]),
            contract=contract,
            created_by=actor_id,
            is_active=True,
        )
        db.add(profile)
        await db.flush()
        profile_id = profile.id

    batch.mapping_contract = contract
    batch.mapping_profile_id = profile_id
    batch.mapping_reviewed_by = actor_id
    batch.mapping_reviewed_at = datetime.now(UTC)
    batch.status = ImportStatus.PENDING
    await audit_service.emit(
        event_type="import_mapping_approved",
        entity_type="import_batch",
        entity_id=batch.id,
        actor_id=actor_id,
        old_value={"status": ImportStatus.MAPPING_REVIEW.value},
        new_value={
            "status": ImportStatus.PENDING.value,
            "profile_id": profile_id,
            "mapped_header_count": len(request.fields),
        },
        project_id=project_id,
        db=db,
    )
    await db.flush()
    await db.refresh(batch)
    return serialize_batch(batch)


async def materialize_approved_import(batch_id: str, db: AsyncSession) -> ImportBatch:
    """Worker entrypoint for an explicitly approved external mapping contract."""

    batch = await db.get(ImportBatch, batch_id)
    if batch is None:
        raise HTTPException(
            status_code=404,
            detail={"detail": "Import batch not found", "error_code": "IMPORT_BATCH_NOT_FOUND"},
        )
    if batch.intake_mode != "external_mapping" or batch.status != ImportStatus.PENDING:
        raise HTTPException(
            status_code=409,
            detail={"detail": "Import batch is not ready for approved mapping materialization.", "error_code": "IMPORT_MAPPING_NOT_READY"},
        )
    batch.status = ImportStatus.PROCESSING
    await _materialize_batch(
        batch,
        db,
        actor_id=batch.mapping_reviewed_by or "system-import",
        correlation_id=str(uuid.uuid4()),
    )
    await db.flush()
    await db.refresh(batch)
    return batch


async def mark_import_failed(batch_id: str, error_details: dict[str, object], db: AsyncSession) -> None:
    """Mark an import batch as failed."""

    batch = await db.get(ImportBatch, batch_id)
    if batch is None:
        return
    batch.status = ImportStatus.FAILED
    batch.error_details = cast(dict[str, Any] | None, sanitize_for_json(error_details))
    await db.flush()


async def list_import_batches(project_id: str, db: AsyncSession) -> ImportBatchListResponse:
    """List all import batches for a project."""

    result = await db.scalars(
        select(ImportBatch)
        .where(ImportBatch.project_id == project_id)
        .order_by(ImportBatch.created_at.desc())
    )
    return ImportBatchListResponse(import_batches=[serialize_batch(batch) for batch in result.all()])


async def get_import_batch(project_id: str, batch_id: str, db: AsyncSession) -> ImportBatchResponse:
    """Load one import batch for a project."""

    batch = await db.scalar(
        select(ImportBatch).where(
            ImportBatch.project_id == project_id,
            ImportBatch.id == batch_id,
        )
    )
    if batch is None:
        raise HTTPException(
            status_code=404,
            detail={"detail": "Import batch not found", "error_code": "IMPORT_BATCH_NOT_FOUND"},
        )
    return serialize_batch(batch)


async def list_import_rows(
    project_id: str,
    batch_id: str,
    db: AsyncSession,
    page: int = 1,
    page_size: int = 100,
) -> SourceRowListResponse:
    """Return paginated source rows for one batch."""

    batch = await db.scalar(
        select(ImportBatch).where(
            ImportBatch.project_id == project_id,
            ImportBatch.id == batch_id,
        )
    )
    if batch is None:
        raise HTTPException(
            status_code=404,
            detail={"detail": "Import batch not found", "error_code": "IMPORT_BATCH_NOT_FOUND"},
        )

    query = select(SourceIntegrationRow).where(SourceIntegrationRow.import_batch_id == batch_id)
    total = await db.scalar(select(func.count()).select_from(query.subquery()))
    offset = (page - 1) * page_size
    result = await db.scalars(
        query.order_by(SourceIntegrationRow.source_row_number).offset(offset).limit(page_size)
    )
    return SourceRowListResponse(
        rows=[serialize_source_row(row) for row in result.all()],
        total=total or 0,
        page=page,
        page_size=page_size,
    )


def _quality_metric(label: str, value: str, detail: str) -> ImportQualityMetric:
    return ImportQualityMetric(label=label, value=value, detail=detail)


def _quality_finding(
    severity: str,
    title: str,
    summary: str,
    action_label: str,
    action_href: str,
) -> ImportQualityFinding:
    return ImportQualityFinding(
        severity=severity,
        title=title,
        summary=summary,
        action_label=action_label,
        action_href=action_href,
    )


def _quality_pct(complete: int, total: int) -> str:
    return f"{round((complete / total) * 100)}%" if total else "0%"


def _source_row_has_value(row: SourceIntegrationRow, header_map: dict[str, str], field: str) -> bool:
    value = _extract_raw_value(cast(dict[str, object], row.raw_data or {}), header_map, field)
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


async def get_import_quality_assistant(
    project_id: str,
    batch_id: str,
    db: AsyncSession,
) -> ImportQualityAssistantResponse:
    """Return deterministic data-quality guidance for one import batch."""

    batch = await db.scalar(
        select(ImportBatch)
        .options(selectinload(ImportBatch.source_rows))
        .where(
            ImportBatch.project_id == project_id,
            ImportBatch.id == batch_id,
        )
    )
    if batch is None:
        raise HTTPException(
            status_code=404,
            detail={"detail": "Import batch not found", "error_code": "IMPORT_BATCH_NOT_FOUND"},
        )

    if batch.status == ImportStatus.MAPPING_REVIEW:
        contract = cast(dict[str, object], batch.mapping_contract or {})
        fields = import_mapping_service.contract_items(contract, "fields")
        questions = import_mapping_service.contract_items(contract, "questions")
        answers = cast(dict[str, str], contract.get("answers", {}))
        unresolved_fields = sum(
            1 for item in fields if str(item.get("target_field", import_mapping_service.EVIDENCE_ONLY)) == import_mapping_service.EVIDENCE_ONLY
        )
        unresolved_questions = [item for item in questions if not answers.get(str(item.get("id", "")))]
        import_href = f"/projects/{project_id}/import?batch_id={batch_id}"
        staged_findings = [
            _quality_finding(
                "high",
                "External workbook is staged for mapping review",
                "No integration, QA, topology, or BOM record has been materialized. Classify each source column and answer the semantic questions before approval.",
                "Open mapping review",
                import_href,
            )
        ]
        for question in unresolved_questions[:5]:
            staged_findings.append(
                _quality_finding(
                    "medium",
                    "Client input required",
                    str(question.get("reason") or question.get("prompt") or "A source field needs clarification."),
                    "Answer in mapping review",
                    import_href,
                )
            )
        return ImportQualityAssistantResponse(
            project_id=project_id,
            batch_id=batch_id,
            status=batch.status.value,
            filename=batch.filename,
            row_count=batch.source_row_count or len(batch.source_rows),
            included_count=batch.candidate_count or 0,
            excluded_count=batch.excluded_count or 0,
            technical_only_count=batch.tbq_n_count or 0,
            normalization_event_count=0,
            recommended_next_action=(
                "Use the guided mapping review to classify every column and answer the outstanding business semantics. "
                "Approve only when the source-to-App contract is correct."
            ),
            metrics=[
                _quality_metric("Rows staged", str(batch.source_row_count or len(batch.source_rows)), "Raw workbook evidence is preserved but has not entered the catalog."),
                _quality_metric("Candidate rows", str(batch.candidate_count or 0), "Rows eligible for technical governance after approval."),
                _quality_metric("Columns to classify", str(len(fields)), "Every external column needs an explicit App field or evidence-only decision."),
                _quality_metric("Evidence only", str(unresolved_fields), "Columns retained for traceability instead of being used in calculations."),
                _quality_metric("Questions open", str(len(unresolved_questions)), "Semantic choices that require user confirmation before materialization."),
            ],
            findings=staged_findings,
        )

    source_rows = sorted(batch.source_rows, key=lambda row: row.source_row_number)
    included_rows = [row for row in source_rows if row.included]
    excluded_rows = [row for row in source_rows if not row.included]
    header_map = cast(dict[str, str], batch.header_map or {})
    technical_only_rows = [
        row
        for row in included_rows
        if str(
            _extract_raw_value(cast(dict[str, object], row.raw_data or {}), header_map, "tbq")
            or "N"
        ).strip().upper()
        != "Y"
    ]
    normalization_event_count = sum(len(row.normalization_events or []) for row in source_rows)
    included_total = len(included_rows)
    coverage_fields = {
        "Payload": "payload_per_execution_kb",
        "Pattern": "selected_pattern",
        "Trigger": "trigger_type",
        "Source": "source_system",
        "Destination": "destination_system",
    }
    coverage_counts = {
        label: sum(1 for row in included_rows if _source_row_has_value(row, header_map, field))
        for label, field in coverage_fields.items()
    }
    metrics = [
        _quality_metric("Rows parsed", str(len(source_rows)), "Source rows persisted from this workbook batch."),
        _quality_metric("Included rows", str(len(included_rows)), "Rows that entered the governed catalog flow."),
        _quality_metric(
            "Technical only",
            str(len(technical_only_rows)),
            "TBQ=N integrations included in technical governance and excluded from BOM and pricing.",
        ),
        _quality_metric("Excluded rows", str(len(excluded_rows)), "Rows excluded by import rules or workbook evidence."),
        _quality_metric(
            "Normalizations",
            str(normalization_event_count),
            "Frequency, payload, or header normalization events captured during ingest.",
        ),
        *[
            _quality_metric(label, _quality_pct(count, included_total), f"{count} of {included_total} included rows populated.")
            for label, count in coverage_counts.items()
        ],
    ]
    findings: list[ImportQualityFinding] = []
    import_href = f"/projects/{project_id}/import?batch_id={batch_id}"
    catalog_href = f"/projects/{project_id}/catalog"
    if batch.status.value == "failed":
        findings.append(
            _quality_finding(
                "critical",
                "Import failed before catalog governance",
                "The batch did not complete, so no downstream catalog quality can be trusted for this upload.",
                "Review import error",
                import_href,
            )
        )
    if excluded_rows:
        reasons = sorted({row.exclusion_reason or "No reason captured" for row in excluded_rows})
        findings.append(
            _quality_finding(
                "medium",
                "Source rows were excluded",
                f"{len(excluded_rows)} row(s) were excluded. Top reason: {reasons[0]}.",
                "Review excluded rows",
                import_href,
            )
        )
    if technical_only_rows:
        findings.append(
            _quality_finding(
                "positive",
                "Technical integrations retained outside the quote",
                f"{len(technical_only_rows)} TBQ=N integration(s) remain available for architecture analysis and are excluded from the economic exercise.",
                "Open governed catalog",
                catalog_href,
            )
        )
    for label, count in coverage_counts.items():
        if included_total and count < included_total:
            severity = "high" if label in {"Payload", "Pattern", "Trigger"} else "medium"
            findings.append(
                _quality_finding(
                    severity,
                    f"{label} coverage is incomplete",
                    f"{included_total - count} included row(s) are missing {label.lower()} evidence.",
                    "Open governed catalog",
                    catalog_href,
                )
            )
    if normalization_event_count:
        findings.append(
            _quality_finding(
                "positive",
                "Workbook normalization evidence captured",
                f"{normalization_event_count} normalization event(s) preserve the ingest decision trail.",
                "Review source rows",
                import_href,
            )
        )

    if any(finding.severity in {"critical", "high"} for finding in findings):
        next_action = "Resolve high-priority import evidence gaps before using forecasts or AI Review for sign-off."
    elif findings:
        next_action = "Review import evidence, then continue technical governance; move TBQ to Y only when an integration is ready for the economic exercise."
    else:
        next_action = "Import quality is clean; continue with catalog QA, canvas design, and recalculation."
    return ImportQualityAssistantResponse(
        project_id=project_id,
        batch_id=batch_id,
        status=batch.status.value,
        filename=batch.filename,
        row_count=len(source_rows),
        included_count=len(included_rows),
        excluded_count=len(excluded_rows),
        technical_only_count=len(technical_only_rows),
        normalization_event_count=normalization_event_count,
        recommended_next_action=next_action,
        metrics=metrics,
        findings=findings,
    )


async def save_upload_file(
    file_name: str,
    contents: bytes,
    *,
    project_id: str = "unassigned",
) -> str:
    """Persist an uploaded source workbook in authoritative Object Storage."""

    key = f"imports/{project_id}/{uuid.uuid4()}-{storage_service.safe_filename(file_name)}"
    return storage_service.put_bytes(
        key,
        contents,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        metadata={"project-id": project_id},
    )


async def delete_import_batch(
    project_id: str,
    batch_id: str,
    actor_id: str,
    db: AsyncSession,
) -> ImportBatchDeleteResponse:
    """Remove one import batch and its governed descendants, then recalculate the project."""

    from app.services.catalog_service import serialize_catalog_integration
    from app.services.justification_service import serialize_justification_record

    batch = await db.scalar(
        select(ImportBatch)
        .options(selectinload(ImportBatch.source_rows))
        .where(
            ImportBatch.project_id == project_id,
            ImportBatch.id == batch_id,
        )
    )
    if batch is None:
        raise HTTPException(
            status_code=404,
            detail={"detail": "Import batch not found", "error_code": "IMPORT_BATCH_NOT_FOUND"},
        )
    if batch.status in {ImportStatus.PENDING, ImportStatus.PROCESSING}:
        raise HTTPException(
            status_code=409,
            detail={
                "detail": "Cannot remove an import while it is still processing.",
                "error_code": "IMPORT_BATCH_ACTIVE",
            },
        )

    old_value = serialize_batch(batch).model_dump()
    storage_reference = batch.storage_reference
    source_rows = list(batch.source_rows)
    source_row_ids = [row.id for row in source_rows]
    integrations = (
        await db.scalars(
            select(CatalogIntegration).where(CatalogIntegration.source_row_id.in_(source_row_ids))
        )
    ).all() if source_row_ids else []
    integration_ids = [integration.id for integration in integrations]
    justifications = (
        await db.scalars(
            select(JustificationRecord).where(
                JustificationRecord.project_id == project_id,
                JustificationRecord.integration_id.in_(integration_ids),
            )
        )
    ).all() if integration_ids else []

    for record in justifications:
        await audit_service.emit(
            event_type="justification_deleted",
            entity_type="justification_record",
            entity_id=record.id,
            actor_id=actor_id,
            old_value=serialize_justification_record(record).model_dump(),
            new_value=None,
            project_id=project_id,
            db=db,
        )
        await db.delete(record)

    for integration in integrations:
        await audit_service.emit(
            event_type="catalog_deleted",
            entity_type="catalog_integration",
            entity_id=integration.id,
            actor_id=actor_id,
            old_value=serialize_catalog_integration(integration).model_dump(),
            new_value=None,
            project_id=project_id,
            db=db,
        )
        await db.delete(integration)

    for source_row in source_rows:
        await db.delete(source_row)
    await db.delete(batch)
    await db.flush()
    if storage_reference:
        storage_service.delete(storage_reference)

    snapshot = await recalc_service.recalculate_project(project_id=project_id, actor_id=actor_id, db=db)
    response = ImportBatchDeleteResponse(
        project_id=project_id,
        batch_id=batch_id,
        detail="Import removed and project recalculated.",
        deleted_source_rows=len(source_rows),
        deleted_integrations=len(integrations),
        deleted_justifications=len(justifications),
        recalculated_snapshot_id=snapshot.id,
    )
    await audit_service.emit(
        event_type="import_deleted",
        entity_type="import_batch",
        entity_id=batch_id,
        actor_id=actor_id,
        old_value=old_value,
        new_value=response.model_dump(),
        project_id=project_id,
        db=db,
    )
    return response
