"""End-to-end governance tests for official OCI commercial evidence."""

from __future__ import annotations

from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO
from typing import cast

import pytest
from fastapi import HTTPException
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncEngine, AsyncSession, async_sessionmaker

from app.models import (
    AuditEvent,
    CommercialDocumentSnapshot,
    CommercialException,
    CommercialMappingCandidate,
    CommercialRelease,
    CommercialRuleFamily,
    CommercialSku,
    GovernanceChangeSet,
    GovernanceSourceArtifact,
    PriceCatalogSnapshot,
    PriceItem,
    PriceSource,
    PriceSyncJob,
    ServiceProductSkuMapping,
    SkuCommercialConstraint,
    SkuCommercialRelationship,
    SkuCommercialTerm,
)
from app.services import commercial_catalog_service, storage_service
from app.services.commercial_catalog_service import (
    GENERATOR_VERSION,
    PARSER_VERSION,
    _commercial_decimal_availability,
    _commercial_fixture_result,
    _meaningful_document_text,
    _quantity_behavior,
    _stable_sku,
    commercial_agent_evidence,
)
from app.services.commercial_document_parser import PRICE_LIST_SHEET, SUPPLEMENT_SHEET


@pytest.mark.parametrize("value", [None, "", "   ", "-"])
def test_meaningful_document_text_rejects_empty_markers(value: str | None) -> None:
    assert _meaningful_document_text(value) is False


def test_meaningful_document_text_accepts_commercial_guidance() -> None:
    assert _meaningful_document_text("Includes management entitlement") is True


def _constraint(
    constraint_type: str,
    scope: str,
    value: str | None = None,
) -> dict[str, object]:
    return {
        "constraint_type": constraint_type,
        "scope": scope,
        "numeric_value": Decimal(value) if value is not None else None,
    }


@pytest.mark.parametrize(
    ("part_number", "price_type", "metric_name", "constraints"),
    [
        (
            "B93306",
            "HOUR",
            "Execution Hour",
            [
                _constraint("free_tier_allowance", "monthly_billed_quantity", "30"),
                _constraint("paid_tier_start", "pay_as_you_go", "30"),
            ],
        ),
        (
            "B95703",
            "HOUR",
            "ECPU Per Hour",
            [_constraint("license_eligibility", "byol")],
        ),
        (
            "B95754",
            "MONTH",
            "Gigabyte Storage Capacity Per Month",
            [
                _constraint("purchase_increment", "database_storage_quantity", "1024"),
                _constraint("purchase_increment", "backup_storage_quantity", "1"),
            ],
        ),
        (
            "B88206",
            "PER_ITEM",
            "Universal Credits",
            [
                _constraint("commitment_minimum", "annual_commitment", "2000"),
                _constraint("commitment_minimum", "annual_flex", "100000"),
            ],
        ),
        ("B92598", "HOUR", "DI Workspace Hours", []),
    ],
)
def test_official_sku_fixture_contracts_are_not_generic_placeholders(
    part_number: str,
    price_type: str,
    metric_name: str,
    constraints: list[dict[str, object]],
) -> None:
    passed, checks = _commercial_fixture_result(
        part_number=part_number,
        behavior="continuous",
        increment=Decimal("0.000001"),
        minimum=Decimal("0"),
        price_type=price_type,
        allow_decimal=True,
        metric_name=metric_name,
        constraints=constraints,
        api_items=[],
    )
    assert passed, checks
    assert len(checks) >= 2


def test_api_gateway_fixture_rejects_integer_package_semantics() -> None:
    passed, checks = _commercial_fixture_result(
        part_number="B92072",
        behavior="packaged",
        increment=Decimal("1"),
        minimum=Decimal("1"),
        price_type="MONTH",
        allow_decimal=False,
        metric_name="1,000,000 API Calls Per Month",
        constraints=[],
        api_items=[],
    )
    assert not passed
    assert next(item for item in checks if item["name"] == "api_gateway_decimal_scale")["passed"] is False


def test_api_gateway_billing_semantics_override_estimator_whole_unit_hint() -> None:
    assert _commercial_decimal_availability("B92072", False) is True
    assert _commercial_decimal_availability("B92598", False) is False


def _official_workbook() -> bytes:
    workbook = Workbook()
    price = workbook.active
    price.title = PRICE_LIST_SHEET
    supplement = workbook.create_sheet(SUPPLEMENT_SHEET)
    price.append(
        [
            "Service Category",
            "Service Name",
            "Pay as You Go",
            "Annual Commitment",
            "Metric",
            "Metric Minimum",
            "Additional Information",
            "Notes",
            "Part Number",
        ]
    )
    price.append(
        [
            "Oracle Data Management Cloud Services",
            "Oracle Autonomous AI Lakehouse - Exadata Storage",
            0.04,
            0.04,
            "Gigabyte Storage Capacity Per Month",
            1,
            "Storage is metered monthly in gigabytes.",
            1,
            "B95702",
        ]
    )
    price.append(
        [
            "Oracle Data Management Cloud Services",
            "Oracle Autonomous AI Lakehouse - ECPU",
            0.336,
            0.336,
            "ECPU Per Hour",
            2,
            "Partial ECPU hours consumed are billed per second with a one-minute minimum.",
            1,
            "B95701",
        ]
    )
    supplement.append(
        [
            "Part Number",
            "SUBSCRIPTION SERVICE",
            "Metric",
            "INCLUDED WITH SUBSCRIPTION SERVICE",
            "ADDITIONAL REQUIREMENTS AND PREREQUISITES",
        ]
    )
    supplement.append(
        [
            "B95701",
            "Oracle Autonomous AI Lakehouse - ECPU",
            "ECPU Per Hour",
            "Includes database feature entitlement.",
            "Exadata Storage for ECPU is required for Exadata deployments.",
        ]
    )
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


async def _seed_pricing_scope(test_engine: AsyncEngine) -> None:
    session_factory = async_sessionmaker(test_engine, expire_on_commit=False, class_=AsyncSession)
    async with session_factory() as db:
        now = datetime.now(UTC)
        source = PriceSource(
            id="source-public",
            name="Oracle public pricing API",
            source_type="public_api",
            currency="USD",
            status="active",
            created_by="test",
        )
        snapshot = PriceCatalogSnapshot(
            id="price-approved",
            source_id=source.id,
            currency="USD",
            retrieved_at=now,
            content_hash="api-price-hash",
            item_count=2,
            approval_status="approved",
            approved_by="reviewer",
            approved_at=now,
        )
        sync_job = PriceSyncJob(
            id="sync-official-set",
            source_id=source.id,
            requested_by="test",
            currency="USD",
            status="completed",
            started_at=now,
            completed_at=now,
            item_count=2,
            changes_detected=2,
            snapshot_id=snapshot.id,
        )
        snapshot.sync_job_id = sync_job.id
        db.add_all([source, sync_job, snapshot])
        await db.flush()
        change_set = GovernanceChangeSet(
            id="change-set-official",
            sync_job_id=sync_job.id,
            price_source_id=source.id,
            price_snapshot_id=snapshot.id,
            trigger_type="manual",
            currency="USD",
            status="promoted",
            validation_status="passed",
            approval_status="approved",
            approved_by="reviewer",
            approved_at=now,
            promoted_at=now,
        )
        db.add(change_set)
        await db.flush()
        artifact_payloads: dict[str, dict[str, object]] = {
            "products": {
                "items": [
                    {
                        "id": "autonomous-ai-lakehouse-ecpu",
                        "partNumber": "B95701",
                        "name": "Oracle Autonomous AI Lakehouse - ECPU",
                        "priceType": "HOUR",
                        "allowDecimalQty": True,
                        "metricId": "ecpu-hour",
                    },
                    {
                        "id": "autonomous-ai-lakehouse-storage",
                        "partNumber": "B95702",
                        "name": "Oracle Autonomous AI Lakehouse - Exadata Storage",
                        "priceType": "MONTH",
                        "allowDecimalQty": True,
                        "metricId": "storage-gb-month",
                    },
                ]
            },
            "metrics": {
                "items": [
                    {"id": "ecpu-hour", "name": "ECPU Per Hour"},
                    {"id": "storage-gb-month", "name": "Gigabyte Storage Capacity Per Month"},
                ]
            },
            "presets": {"items": []},
        }
        for source_kind, payload in artifact_payloads.items():
            items = cast(list[object], payload["items"])
            reference = storage_service.put_json(
                f"tests/commercial/{source_kind}.json", payload
            )
            db.add(
                GovernanceSourceArtifact(
                    change_set_id=change_set.id,
                    source_kind=source_kind,
                    source_url=f"https://www.oracle.com/{source_kind}.json",
                    content_hash=f"{source_kind}-hash",
                    record_count=len(items),
                    storage_reference=reference,
                    retrieval_status="verified",
                    retrieved_at=now,
                )
            )
        db.add_all(
            [
                PriceItem(
                    id="price-b95701",
                    snapshot_id=snapshot.id,
                    part_number="B95701",
                    display_name="Oracle Autonomous AI Lakehouse - ECPU",
                    metric_name="ECPU Per Hour",
                    service_category="Oracle Data Management Cloud Services",
                    price_type="HOUR",
                    currency="USD",
                    model="PAY_AS_YOU_GO",
                    value=0.336,
                ),
                PriceItem(
                    id="price-b95702",
                    snapshot_id=snapshot.id,
                    part_number="B95702",
                    display_name="Oracle Autonomous AI Lakehouse - Exadata Storage",
                    metric_name="Gigabyte Storage Capacity Per Month",
                    service_category="Oracle Data Management Cloud Services",
                    price_type="MONTH",
                    currency="USD",
                    model="PAY_AS_YOU_GO",
                    value=0.04,
                ),
                ServiceProductSkuMapping(
                    id="mapping-b95701",
                    service_id="AUTONOMOUS_AI_LAKEHOUSE",
                    tool_key="AUTONOMOUS_AI_LAKEHOUSE",
                    part_number="B95701",
                    billing_metric_key="ecpu_hours",
                    formula_key="metered_quantity",
                    quantity_behavior="continuous",
                    quantity_increment=0.000001,
                    minimum_quantity=2,
                    quantity_unit="ECPU-hours",
                    status="approved",
                    version="1.0.0",
                ),
            ]
        )
        await db.commit()


@pytest.mark.asyncio
async def test_official_workbook_requires_human_review_before_atomic_release(
    api_client: AsyncClient,
    test_engine: AsyncEngine,
    isolated_object_storage: dict[str, bytes],
) -> None:
    await _seed_pricing_scope(test_engine)
    session_factory = async_sessionmaker(test_engine, expire_on_commit=False, class_=AsyncSession)
    async with session_factory() as db:
        db.add(
            ServiceProductSkuMapping(
                id="mapping-outside-approved-document",
                service_id="UNRESOLVED_PRODUCT",
                tool_key="UNRESOLVED_PRODUCT",
                part_number="B99999",
                billing_metric_key="unresolved_quantity",
                formula_key="metered_quantity",
                quantity_behavior="continuous",
                quantity_increment=1,
                minimum_quantity=0,
                quantity_unit="units",
                status="approved",
                version="1.0.0",
            )
        )
        await db.commit()
    headers = {"X-Actor-Role": "Admin", "X-Actor-Id": "commercial-reviewer"}
    imported = await api_client.post(
        "/api/v1/pricing/commercial-documents",
        headers=headers,
        files={
            "file": (
                "ORACLE_LOCALIZABLE_PRICE_LIST.xlsx",
                _official_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert imported.status_code == 200, imported.text
    workspace = imported.json()
    document_id = workspace["document"]["id"]
    candidate = next(item for item in workspace["candidates"] if item["part_number"] == "B95701")
    prerequisite_candidate = next(
        item for item in workspace["candidates"] if item["part_number"] == "B95702"
    )
    dependency = next(item for item in workspace["exceptions"] if item["code"] == "DEPENDENCY_UNRESOLVED")
    assert workspace["document"]["status"] == "review_required"
    assert candidate["status"] == "pending_review"
    assert candidate["generator_version"] == GENERATOR_VERSION
    assert candidate["rule_status"] == "ready_for_review"
    assert candidate["rule_fixture_status"] == "passed"
    assert candidate["identity"]["display_name"] == "Oracle Autonomous AI Lakehouse - ECPU"
    assert candidate["identity"]["service_category"] == "Oracle Data Management Cloud Services"
    assert "product_hierarchy" not in candidate["identity"]
    assert candidate["commercial_term"] == {
        "service_name": "Oracle Autonomous AI Lakehouse - ECPU",
        "metric_name": "ECPU Per Hour",
        "price_type": "HOUR",
    }
    detail = await api_client.get(
        f"/api/v1/pricing/commercial-candidates/{candidate['id']}", headers=headers
    )
    assert detail.status_code == 200, detail.text
    detail_candidate = detail.json()
    assert detail_candidate["identity"]["product_hierarchy"] == ["Oracle Data Management Cloud Services"]
    assert detail_candidate["identity"]["product_paths"] == [["Oracle Data Management Cloud Services"]]
    assert detail_candidate["identity"]["official_location_count"] == 1
    assert detail_candidate["identity"]["structured_product"]["name"] == (
        "Oracle Autonomous AI Lakehouse - ECPU"
    )
    assert detail_candidate["commercial_term"]["metric_name"] == "ECPU Per Hour"
    assert workspace["field_authority"]["contract_rate"] == "authorized_customer_rate_card"
    assert any(reference.endswith("ORACLE_LOCALIZABLE_PRICE_LIST.xlsx") for reference in isolated_object_storage)

    first_page = await api_client.get(
        "/api/v1/pricing/commercial-catalog",
        headers=headers,
        params={"document_id": document_id, "page": 1, "page_size": 1, "status": "pending_review"},
    )
    assert first_page.status_code == 200, first_page.text
    assert first_page.json()["page"] == 1
    assert first_page.json()["page_size"] == 1
    assert first_page.json()["total"] == 2
    assert len(first_page.json()["candidates"]) == 1

    filtered = await api_client.get(
        "/api/v1/pricing/commercial-catalog",
        headers=headers,
        params={"document_id": document_id, "search": "B95701", "page": 1, "page_size": 50},
    )
    assert filtered.status_code == 200, filtered.text
    assert {item["part_number"] for item in filtered.json()["candidates"]} == {"B95701"}
    assert {item["part_number"] for item in filtered.json()["exceptions"]} == {"B95701"}

    filtered_by_name = await api_client.get(
        "/api/v1/pricing/commercial-catalog",
        headers=headers,
        params={"document_id": document_id, "search": "Autonomous AI Lakehouse", "page": 1, "page_size": 50},
    )
    assert filtered_by_name.status_code == 200, filtered_by_name.text
    assert {item["part_number"] for item in filtered_by_name.json()["candidates"]} == {
        "B95701",
        "B95702",
    }

    premature = await api_client.post(
        f"/api/v1/pricing/commercial-documents/{document_id}/releases", headers=headers
    )
    assert premature.status_code == 409

    approved_document = await api_client.post(
        f"/api/v1/pricing/commercial-documents/{document_id}/approve", headers=headers
    )
    assert approved_document.status_code == 200
    blocked_candidate_review = await api_client.post(
        f"/api/v1/pricing/commercial-candidates/{candidate['id']}/review",
        headers=headers,
        json={"decision": "approve", "rationale": "Matched exact official part number and metric."},
    )
    assert blocked_candidate_review.status_code == 409, blocked_candidate_review.text
    assert blocked_candidate_review.json()["detail"]["error_code"] == (
        "COMMERCIAL_CANDIDATE_APPROVAL_BLOCKED"
    )
    reviewed_prerequisite = await api_client.post(
        f"/api/v1/pricing/commercial-candidates/{prerequisite_candidate['id']}/review",
        headers=headers,
        json={"decision": "approve", "rationale": "Exact prerequisite SKU evidence is present."},
    )
    assert reviewed_prerequisite.status_code == 200, reviewed_prerequisite.text
    reviewed_exception = await api_client.post(
        f"/api/v1/pricing/commercial-exceptions/{dependency['id']}/review",
        headers=headers,
        json={
            "decision": "resolve",
            "rationale": "Linked to the exact governed Exadata Storage prerequisite SKU.",
            "target_part_number": "B95702",
        },
    )
    assert reviewed_exception.status_code == 200, reviewed_exception.text
    reviewed_candidate = await api_client.post(
        f"/api/v1/pricing/commercial-candidates/{candidate['id']}/review",
        headers=headers,
        json={"decision": "approve", "rationale": "Matched exact official part number and metric."},
    )
    assert reviewed_candidate.status_code == 200, reviewed_candidate.text
    promoted = await api_client.post(
        f"/api/v1/pricing/commercial-documents/{document_id}/releases", headers=headers
    )
    assert promoted.status_code == 200, promoted.text
    release_payload = promoted.json()["releases"][0]
    assert release_payload["status"] == "approved"
    assert release_payload["metadata"]["scope"] == "global_oci_catalog"
    assert release_payload["metadata"]["catalog_part_numbers"] == ["B95701", "B95702"]
    assert release_payload["metadata"]["quote_ready_part_numbers"] == ["B95701", "B95702"]
    assert release_payload["metadata"]["blocked_part_numbers"] == []
    assert release_payload["metadata"]["part_numbers"] == ["B95701"]
    assert release_payload["metadata"]["excluded_mapping_parts"] == ["B99999"]
    assert release_payload["metadata"]["included_mapping_count"] == 1
    assert release_payload["metadata"]["available_mapping_count"] == 2
    assert release_payload["metadata"]["excluded_mapping_count"] == 1
    outside_catalog_reasons = set(
        release_payload["metadata"]["excluded_mapping_reasons"]["B99999"]
    )
    assert outside_catalog_reasons
    assert "not_in_global_catalog" in outside_catalog_reasons or {
        "candidate_not_approved",
        "rule_not_approved",
    }.issubset(outside_catalog_reasons)

    async with session_factory() as db:
        term = await db.scalar(select(SkuCommercialTerm).where(SkuCommercialTerm.part_number == "B95701"))
        assert term is not None
        constraints = list(
            (await db.scalars(select(SkuCommercialConstraint).where(SkuCommercialConstraint.term_id == term.id))).all()
        )
        assert {(item.constraint_type, item.numeric_value) for item in constraints} >= {
            ("metric_minimum", 2),
            ("billing_granularity", 1),
            ("minimum_duration", 60),
        }
        assert await db.scalar(select(func.count()).select_from(CommercialRelease)) == 1
        relationship = await db.scalar(
            select(SkuCommercialRelationship).where(
                SkuCommercialRelationship.part_number == "B95701",
                SkuCommercialRelationship.relationship_type == "requires",
            )
        )
        assert relationship is not None
        assert relationship.resolution_status == "resolved"
        assert relationship.status == "approved"
        assert relationship.target_part_number == "B95702"
        agent_evidence = await commercial_agent_evidence(db)
        assert agent_evidence["commercial_release_scope"] == {
            "status": "partial",
            "included_part_count": 1,
            "available_mapping_part_count": 2,
            "excluded_mapping_part_count": 1,
            "excluded_mapping_parts": ["B99999"],
            "excluded_mapping_reasons": release_payload["metadata"][
                "excluded_mapping_reasons"
            ],
        }
        assert agent_evidence["candidate_revalidation"] == {
            "status": "clear",
            "count": 0,
            "items": [],
            "current_generator_version": GENERATOR_VERSION,
        }


@pytest.mark.asyncio
async def test_new_document_version_reuses_stable_sku_identity(
    api_client: AsyncClient, test_engine: AsyncEngine
) -> None:
    await _seed_pricing_scope(test_engine)
    headers = {"X-Actor-Role": "Admin", "X-Actor-Id": "commercial-reviewer"}
    first_contents = _official_workbook()
    first = await api_client.post(
        "/api/v1/pricing/commercial-documents",
        headers=headers,
        files={"file": ("price-list-v1.xlsx", first_contents, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert first.status_code == 200, first.text

    workbook = load_workbook(BytesIO(first_contents))
    workbook.properties.title = "Official price list version 2"
    workbook[PRICE_LIST_SHEET]["F3"] = 4
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    second = await api_client.post(
        "/api/v1/pricing/commercial-documents",
        headers=headers,
        files={"file": ("price-list-v2.xlsx", output.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert second.status_code == 200, second.text
    assert second.json()["document"]["id"] != first.json()["document"]["id"]

    session_factory = async_sessionmaker(test_engine, expire_on_commit=False, class_=AsyncSession)
    async with session_factory() as db:
        assert await db.scalar(select(func.count()).select_from(CommercialSku)) == 2
        assert await db.scalar(select(func.count()).select_from(CommercialDocumentSnapshot)) == 2
        assert await db.scalar(select(func.count()).select_from(SkuCommercialTerm)) == 4
        rules = list(
            (
                await db.scalars(
                    select(CommercialRuleFamily)
                    .where(CommercialRuleFamily.metric_pattern == "ECPU Per Hour")
                    .order_by(CommercialRuleFamily.created_at)
                )
            ).all()
        )
        assert [rule.version for rule in rules] == ["1.0.0", "1.0.1"], [
            (rule.version, rule.minimum_quantity, rule.evidence) for rule in rules
        ]
        assert [rule.minimum_quantity for rule in rules] == [2, 4]


@pytest.mark.asyncio
async def test_stable_sku_records_canonical_placement_drift_without_losing_history(
    test_engine: AsyncEngine,
) -> None:
    session_factory = async_sessionmaker(test_engine, expire_on_commit=False, class_=AsyncSession)
    async with session_factory() as db:
        first, first_drift = await _stable_sku(
            db,
            part_number="B95634",
            display_name="Oracle Cloud Infrastructure Log Analytics - Active Storage",
            service_category="Oracle Monitoring and Diagnostics Services",
            source_product_id=None,
            products_artifact_id="products-v1",
            product_hierarchy=("Management", "Log Analytics"),
            product_paths=(("Management", "Log Analytics"),),
            document_fingerprint="document-v1",
        )
        second, second_drift = await _stable_sku(
            db,
            part_number="B95634",
            display_name="Oracle Cloud Infrastructure Log Analytics - Active Storage",
            service_category="Oracle Monitoring and Diagnostics Services",
            source_product_id=None,
            products_artifact_id="products-v2",
            product_hierarchy=("Observability", "Log Analytics"),
            product_paths=(("Observability", "Log Analytics"),),
            document_fingerprint="document-v2",
        )
        await db.flush()

        assert second.id == first.id
        assert first_drift is None
        assert second_drift == {
            "part_number": "B95634",
            "previous_product_hierarchy": ["Management", "Log Analytics"],
            "current_product_hierarchy": ["Observability", "Log Analytics"],
            "document_fingerprint": "document-v2",
        }
        assert second.identity_metadata["canonical_placement_history"] == [
            {
                "product_hierarchy": ["Management", "Log Analytics"],
                "superseded_by_document_fingerprint": "document-v2",
            }
        ]


@pytest.mark.asyncio
async def test_import_rejects_material_sku_coverage_regression(
    api_client: AsyncClient,
    test_engine: AsyncEngine,
) -> None:
    await _seed_pricing_scope(test_engine)
    session_factory = async_sessionmaker(test_engine, expire_on_commit=False, class_=AsyncSession)
    async with session_factory() as db:
        db.add(
            CommercialDocumentSnapshot(
                id="commercial-full-coverage",
                document_kind="oracle_localizable_price_list",
                source_name="Oracle PaaS and IaaS Public Cloud Localizable Price List",
                original_filename="full-price-list.xlsx",
                storage_reference="minio://tests/full-price-list.xlsx",
                content_hash="full-price-list-hash",
                parser_version=PARSER_VERSION,
                currency="USD",
                status="approved",
                record_count=1163,
                retrieved_at=datetime.now(UTC),
                manifest={},
            )
        )
        await db.commit()

    response = await api_client.post(
        "/api/v1/pricing/commercial-documents",
        headers={"X-Actor-Role": "Admin", "X-Actor-Id": "commercial-reviewer"},
        files={
            "file": (
                "truncated-price-list.xlsx",
                _official_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 422, response.text
    assert response.json()["detail"]["error_code"] == "COMMERCIAL_DOCUMENT_CONTENT_INVALID"
    assert "SKU coverage dropped below 80%" in response.json()["detail"]["detail"]


@pytest.mark.asyncio
async def test_candidate_approval_requires_a_passing_rule_fixture(
    api_client: AsyncClient, test_engine: AsyncEngine
) -> None:
    await _seed_pricing_scope(test_engine)
    headers = {"X-Actor-Role": "Admin", "X-Actor-Id": "commercial-reviewer"}
    imported = await api_client.post(
        "/api/v1/pricing/commercial-documents",
        headers=headers,
        files={
            "file": (
                "price-list.xlsx",
                _official_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert imported.status_code == 200, imported.text
    document_id = imported.json()["document"]["id"]
    candidate_payload = next(
        item for item in imported.json()["candidates"] if item["part_number"] == "B95701"
    )
    dependency_target = next(
        item for item in imported.json()["candidates"] if item["part_number"] == "B95702"
    )
    dependency = next(
        item
        for item in imported.json()["exceptions"]
        if item["part_number"] == "B95701" and item["code"] == "DEPENDENCY_UNRESOLVED"
    )
    approved_document = await api_client.post(
        f"/api/v1/pricing/commercial-documents/{document_id}/approve", headers=headers
    )
    assert approved_document.status_code == 200, approved_document.text
    approved_dependency_target = await api_client.post(
        f"/api/v1/pricing/commercial-candidates/{dependency_target['id']}/review",
        headers=headers,
        json={
            "decision": "approve",
            "rationale": "Approve the exact prerequisite target before resolution.",
        },
    )
    assert approved_dependency_target.status_code == 200, approved_dependency_target.text
    resolved_dependency = await api_client.post(
        f"/api/v1/pricing/commercial-exceptions/{dependency['id']}/review",
        headers=headers,
        json={
            "decision": "resolve",
            "rationale": "Resolve dependency before testing the fixture gate.",
            "target_part_number": "B95702",
        },
    )
    assert resolved_dependency.status_code == 200, resolved_dependency.text

    session_factory = async_sessionmaker(test_engine, expire_on_commit=False, class_=AsyncSession)
    async with session_factory() as db:
        candidate = await db.get(CommercialMappingCandidate, candidate_payload["id"])
        assert candidate is not None
        rule_id = candidate.proposed_mapping["commercial_rule_family_id"]
        rule = await db.get(CommercialRuleFamily, rule_id)
        assert rule is not None
        rule.fixture_status = "failed"
        rule.status = "blocked"
        await db.commit()

    reviewed = await api_client.post(
        f"/api/v1/pricing/commercial-candidates/{candidate_payload['id']}/review",
        headers=headers,
        json={"decision": "approve", "rationale": "Attempt approval for fixture guard."},
    )
    assert reviewed.status_code == 409, reviewed.text
    assert reviewed.json()["detail"]["error_code"] == "COMMERCIAL_RULE_FIXTURE_REQUIRED"

    revalidated = await api_client.post(
        f"/api/v1/pricing/commercial-candidates/{candidate_payload['id']}/revalidate",
        headers=headers,
    )
    assert revalidated.status_code == 200, revalidated.text
    refreshed_candidate = next(
        item
        for item in revalidated.json()["candidates"]
        if item["id"] == candidate_payload["id"]
    )
    assert refreshed_candidate["status"] == "pending_review"
    refreshed_detail = await api_client.get(
        f"/api/v1/pricing/commercial-candidates/{candidate_payload['id']}",
        headers=headers,
    )
    assert refreshed_detail.status_code == 200, refreshed_detail.text
    assert refreshed_detail.json()["proposed_mapping"]["commercial_rule_family_id"] != rule_id

    approved = await api_client.post(
        f"/api/v1/pricing/commercial-candidates/{candidate_payload['id']}/review",
        headers=headers,
        json={"decision": "approve", "rationale": "Revalidated fixture now passes."},
    )
    assert approved.status_code == 200, approved.text

    revalidated_again = await api_client.post(
        f"/api/v1/pricing/commercial-candidates/{candidate_payload['id']}/revalidate",
        headers=headers,
    )
    assert revalidated_again.status_code == 200, revalidated_again.text
    stable_candidate = next(
        item
        for item in revalidated_again.json()["candidates"]
        if item["id"] == candidate_payload["id"]
    )
    assert stable_candidate["status"] == "approved"
    stable_detail = await api_client.get(
        f"/api/v1/pricing/commercial-candidates/{candidate_payload['id']}",
        headers=headers,
    )
    assert stable_detail.status_code == 200, stable_detail.text
    assert stable_detail.json()["proposed_mapping"]["commercial_rule_family_id"] == (
        refreshed_detail.json()["proposed_mapping"]["commercial_rule_family_id"]
    )
    assert stable_detail.json()["reasons"].count(
        f"Deterministically revalidated with {GENERATOR_VERSION}."
    ) == 1


@pytest.mark.parametrize(
    (
        "price_type",
        "allow_decimal",
        "expected_behavior",
        "expected_increment",
        "expected_rounding",
        "expected_proration",
    ),
    [
        ("HOUR", True, "continuous", Decimal("0.000001"), "metered", "prorated"),
        (
            "HOUR_UTILIZED",
            True,
            "continuous",
            Decimal("0.000001"),
            "metered",
            "prorated",
        ),
        ("MONTH", None, "fixed_capacity", Decimal("1"), "ceil_increment", "not_prorated"),
        ("DAY", True, "continuous", Decimal("0.000001"), "metered", "prorated"),
        ("PER_ITEM", False, "packaged", Decimal("1"), "ceil_increment", "not_prorated"),
        ("PER_ITEM", True, "continuous", Decimal("0.000001"), "metered", "prorated"),
    ],
)
def test_global_price_types_generate_explicit_quantity_contracts(
    price_type: str,
    allow_decimal: bool | None,
    expected_behavior: str,
    expected_increment: Decimal,
    expected_rounding: str,
    expected_proration: str,
) -> None:
    behavior, increment, rounding, proration = _quantity_behavior(price_type, allow_decimal)

    assert (behavior, increment, rounding, proration) == (
        expected_behavior,
        expected_increment,
        expected_rounding,
        expected_proration,
    )
    passed, checks = _commercial_fixture_result(
        part_number=f"TEST-{price_type}-{allow_decimal}",
        behavior=behavior,
        increment=increment,
        minimum=Decimal("1") if behavior in {"packaged", "fixed_capacity"} else Decimal("0"),
        price_type=price_type,
        allow_decimal=allow_decimal,
        metric_name=f"Synthetic {price_type} metric",
        constraints=[],
        api_items=[],
    )
    assert passed, checks
    if price_type == "DAY":
        assert next(item for item in checks if item["name"] == "day_metering")["passed"]


async def _import_approved_document(
    api_client: AsyncClient,
    test_engine: AsyncEngine,
) -> tuple[str, dict[str, str]]:
    await _seed_pricing_scope(test_engine)
    headers = {"X-Actor-Role": "Admin", "X-Actor-Id": "commercial-reviewer"}
    imported = await api_client.post(
        "/api/v1/pricing/commercial-documents",
        headers=headers,
        files={
            "file": (
                "ORACLE_LOCALIZABLE_PRICE_LIST.xlsx",
                _official_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert imported.status_code == 200, imported.text
    document_id = imported.json()["document"]["id"]
    approved = await api_client.post(
        f"/api/v1/pricing/commercial-documents/{document_id}/approve",
        headers=headers,
    )
    assert approved.status_code == 200, approved.text
    return document_id, headers


async def _add_catalog_contract_candidate(
    db: AsyncSession,
    *,
    document: CommercialDocumentSnapshot,
    part_number: str,
    classification: str,
    with_api_price: bool,
    exception_code: str | None = None,
    unresolved_dependency: bool = False,
) -> CommercialMappingCandidate:
    sku = CommercialSku(
        id=f"sku-{part_number.lower()}",
        part_number=part_number,
        display_name=f"Contract fixture {part_number}",
        service_category="M51 contract fixtures",
        lifecycle_status="active",
        identity_metadata={},
    )
    db.add(sku)
    await db.flush()

    price_snapshot_id = str(document.manifest["price_snapshot_id"])
    price_item = None
    if with_api_price:
        price_item = PriceItem(
            id=f"price-{part_number.lower()}",
            snapshot_id=price_snapshot_id,
            part_number=part_number,
            display_name=sku.display_name,
            metric_name="Contract fixture units",
            service_category=sku.service_category,
            price_type="PER_ITEM",
            currency="USD",
            model="PAY_AS_YOU_GO",
            value=1,
        )
        db.add(price_item)
        await db.flush()

    term = SkuCommercialTerm(
        id=f"term-{part_number.lower()}",
        document_snapshot_id=document.id,
        commercial_sku_id=sku.id,
        price_catalog_snapshot_id=price_snapshot_id,
        part_number=part_number,
        service_name=sku.display_name,
        service_category=sku.service_category,
        commercial_prices=[],
        currency="USD",
        metric_name="Contract fixture units",
        price_type="PER_ITEM",
        allow_decimal_quantity=True,
        availability=[],
        additional_information="Deterministic M51 contract fixture.",
        disposition=classification,
        family_key=f"per-item::{part_number.lower()}",
        status="review_required",
        confidence=1,
        source_sheet=PRICE_LIST_SHEET,
        source_row=100,
        source_cells={},
        extraction_metadata={},
    )
    rule = CommercialRuleFamily(
        id=f"rule-{part_number.lower()}",
        family_key=f"per-item::{part_number.lower()}",
        version="1.0.0",
        formula_key="metered_quantity",
        metric_pattern="Contract fixture units",
        price_types=["PER_ITEM"],
        quantity_behavior="continuous",
        quantity_increment=Decimal("0.000001"),
        minimum_quantity=Decimal("0"),
        aggregation_window="monthly",
        proration_policy="prorated",
        quote_rounding="metered",
        generator_version=GENERATOR_VERSION,
        status="ready_for_review",
        fixture_status="passed",
        evidence={"fixture_checks": [{"name": "quantity_normalization", "passed": True}]},
    )
    db.add_all([term, rule])
    await db.flush()
    candidate = CommercialMappingCandidate(
        id=f"candidate-{part_number.lower()}",
        document_snapshot_id=document.id,
        commercial_sku_id=sku.id,
        term_id=term.id,
        price_item_id=price_item.id if price_item is not None else None,
        part_number=part_number,
        proposed_service_id=None,
        family_key=rule.family_key,
        classification=classification,
        proposed_mapping={
            "commercial_rule_family_id": rule.id,
            "quantity_behavior": rule.quantity_behavior,
        },
        confidence=1,
        generator_version=GENERATOR_VERSION,
        status="pending_review",
        reasons=[],
    )
    db.add(candidate)
    await db.flush()

    if unresolved_dependency:
        relationship = SkuCommercialRelationship(
            id=f"relationship-{part_number.lower()}",
            document_snapshot_id=document.id,
            source_term_id=term.id,
            source_commercial_sku_id=sku.id,
            part_number=part_number,
            relationship_type="requires",
            target_name="Unresolved prerequisite",
            guidance="A prerequisite must be selected before publication.",
            resolution_status="unresolved",
            confidence=1,
            status="needs_review",
            source_sheet=SUPPLEMENT_SHEET,
            source_row=100,
            source_cell="E100",
        )
        db.add(relationship)
        await db.flush()
        exception_code = exception_code or "DEPENDENCY_UNRESOLVED"
        exception_details: dict[str, object] = {"relationship_id": relationship.id}
    else:
        exception_details = {}
    if exception_code:
        db.add(
            CommercialException(
                document_snapshot_id=document.id,
                candidate_id=candidate.id,
                part_number=part_number,
                exception_code=exception_code,
                severity="high",
                status="open",
                details=exception_details,
                proposed_resolution="Resolve the governed blocker before publication.",
            )
        )
    await db.flush()
    return candidate


async def _finalize_catalog_review_contract(
    document_id: str,
    actor_id: str,
    db: AsyncSession,
) -> dict[str, object]:
    finalizer = getattr(commercial_catalog_service, "finalize_catalog_review", None)
    assert callable(finalizer), (
        "M51 requires commercial_catalog_service.finalize_catalog_review(document_id, "
        "rationale, actor_id, db)"
    )
    result = await finalizer(
        document_id=document_id,
        rationale="Finalize every global OCI SKU with deterministic fixture-backed disposition.",
        actor_id=actor_id,
        db=db,
    )
    return cast(dict[str, object], result)


@pytest.mark.asyncio
async def test_finalize_catalog_review_disposes_every_candidate_once_with_aggregate_audit(
    api_client: AsyncClient,
    test_engine: AsyncEngine,
) -> None:
    document_id, _ = await _import_approved_document(api_client, test_engine)
    session_factory = async_sessionmaker(test_engine, expire_on_commit=False, class_=AsyncSession)
    async with session_factory() as db:
        document = await db.get(CommercialDocumentSnapshot, document_id)
        assert document is not None
        stale_direct = await _add_catalog_contract_candidate(
            db,
            document=document,
            part_number="BTESTDIRECT",
            classification="direct_metered",
            with_api_price=True,
        )
        stale_direct.generator_version = "commercial-product-factory-legacy"
        stale_rule_id = cast(
            str, stale_direct.proposed_mapping["commercial_rule_family_id"]
        )
        stale_rule = await db.get(CommercialRuleFamily, stale_rule_id)
        assert stale_rule is not None
        stale_rule.generator_version = "commercial-product-factory-legacy"
        await _add_catalog_contract_candidate(
            db,
            document=document,
            part_number="BTESTINCLUDED",
            classification="included_non_billable",
            with_api_price=False,
        )
        await _add_catalog_contract_candidate(
            db,
            document=document,
            part_number="BTESTAMBIG",
            classification="direct_metered",
            with_api_price=True,
            exception_code="PRODUCT_IDENTITY_VARIANCE",
        )
        await _add_catalog_contract_candidate(
            db,
            document=document,
            part_number="BTESTNOPRICE",
            classification="direct_metered",
            with_api_price=False,
        )
        await _add_catalog_contract_candidate(
            db,
            document=document,
            part_number="BTESTDEPEND",
            classification="dependent_entitlement",
            with_api_price=True,
            unresolved_dependency=True,
        )
        await db.commit()

        await _finalize_catalog_review_contract(document_id, "commercial-reviewer", db)
        await db.commit()
        candidates = {
            item.part_number: item
            for item in (
                await db.scalars(
                    select(CommercialMappingCandidate).where(
                        CommercialMappingCandidate.document_snapshot_id == document_id
                    )
                )
            ).all()
        }
        assert candidates["BTESTDIRECT"].status == "approved"
        assert candidates["BTESTINCLUDED"].status == "approved"
        assert candidates["BTESTAMBIG"].status == "blocked"
        assert "open_exception:PRODUCT_IDENTITY_VARIANCE" in cast(
            list[str], candidates["BTESTAMBIG"].proposed_mapping["catalog_disposition_reasons"]
        )
        assert candidates["BTESTNOPRICE"].status == "blocked"
        missing_price_reasons = set(
            cast(
                list[str],
                candidates["BTESTNOPRICE"].proposed_mapping["catalog_disposition_reasons"],
            )
        )
        assert missing_price_reasons & {"API_PRICE_MISSING", "approved_api_price_missing"}
        assert candidates["BTESTDEPEND"].status == "blocked"
        dependency_reasons = cast(
            list[str], candidates["BTESTDEPEND"].proposed_mapping["catalog_disposition_reasons"]
        )
        assert "classification:dependent_entitlement" in dependency_reasons
        assert "open_exception:DEPENDENCY_UNRESOLVED" in dependency_reasons
        assert "relationship_not_resolved:unresolved" in dependency_reasons
        assert not [item for item in candidates.values() if item.status == "pending_review"]

        events = list(
            (
                await db.scalars(
                    select(AuditEvent).where(
                        AuditEvent.event_type == "commercial_catalog_review_finalized",
                        AuditEvent.entity_id == document_id,
                    )
                )
            ).all()
        )
        assert len(events) == 1
        assert events[0].new_value is not None
        assert events[0].new_value["approved_count"] + events[0].new_value["blocked_count"] == len(
            candidates
        )

        await _finalize_catalog_review_contract(document_id, "commercial-reviewer", db)
        await db.commit()
        assert (
            await db.scalar(
                select(func.count()).select_from(AuditEvent).where(
                    AuditEvent.event_type == "commercial_catalog_review_finalized",
                    AuditEvent.entity_id == document_id,
                )
            )
        ) == 1


@pytest.mark.asyncio
async def test_promote_release_rejects_an_incomplete_global_catalog_review(
    api_client: AsyncClient,
    test_engine: AsyncEngine,
) -> None:
    document_id, _ = await _import_approved_document(api_client, test_engine)
    session_factory = async_sessionmaker(test_engine, expire_on_commit=False, class_=AsyncSession)
    async with session_factory() as db:
        with pytest.raises(HTTPException) as raised:
            await commercial_catalog_service.promote_release(
                document_id,
                "commercial-reviewer",
                db,
            )
        assert raised.value.status_code == 409
        detail = cast(dict[str, object], raised.value.detail)
        assert detail["error_code"] == "COMMERCIAL_CATALOG_REVIEW_INCOMPLETE"
        assert detail["pending_count"] == 2


@pytest.mark.asyncio
async def test_global_release_freezes_all_dispositions_and_limits_bom_to_quote_ready_subset(
    api_client: AsyncClient,
    test_engine: AsyncEngine,
) -> None:
    document_id, _ = await _import_approved_document(api_client, test_engine)
    session_factory = async_sessionmaker(test_engine, expire_on_commit=False, class_=AsyncSession)
    async with session_factory() as db:
        candidates = list(
            (
                await db.scalars(
                    select(CommercialMappingCandidate)
                    .where(CommercialMappingCandidate.document_snapshot_id == document_id)
                    .order_by(CommercialMappingCandidate.part_number)
                )
            ).all()
        )
        approved = next(item for item in candidates if item.part_number == "B95701")
        blocked = next(item for item in candidates if item.part_number == "B95702")
        approved.status = "approved"
        approved.reasons = ["Exact source identity, API price, and passing rule fixture."]
        blocked.status = "blocked"
        blocked.reasons = ["DEPENDENCY_UNRESOLVED"]
        blocked.proposed_mapping = {
            **blocked.proposed_mapping,
            "catalog_disposition": "blocked",
            "catalog_disposition_reasons": ["DEPENDENCY_UNRESOLVED"],
        }
        rule_id = str(approved.proposed_mapping["commercial_rule_family_id"])
        rule = await db.get(CommercialRuleFamily, rule_id)
        assert rule is not None
        rule.status = "approved"
        rule.fixture_status = "passed"
        term = await db.get(SkuCommercialTerm, approved.term_id)
        assert term is not None
        term.status = "approved"
        relationships = list(
            (
                await db.scalars(
                    select(SkuCommercialRelationship).where(
                        SkuCommercialRelationship.document_snapshot_id == document_id,
                        SkuCommercialRelationship.part_number == approved.part_number,
                    )
                )
            ).all()
        )
        for relationship in relationships:
            relationship.resolution_status = "resolved"
            relationship.status = "approved"
        exceptions = list(
            (
                await db.scalars(
                    select(CommercialException).where(
                        CommercialException.document_snapshot_id == document_id,
                        CommercialException.part_number == approved.part_number,
                    )
                )
            ).all()
        )
        for item in exceptions:
            item.status = "resolved"
        await db.commit()

        workspace = await commercial_catalog_service.promote_release(
            document_id,
            "commercial-reviewer",
            db,
        )
        await db.commit()
        release = cast(list[dict[str, object]], workspace["releases"])[0]
        metadata = cast(dict[str, object], release["metadata"])
        catalog_parts = {item.part_number for item in candidates}
        blocked_parts = set(cast(list[str], metadata["blocked_part_numbers"]))
        quote_ready_parts = set(cast(list[str], metadata["quote_ready_part_numbers"]))

        assert metadata["scope"] == "global_oci_catalog"
        assert set(cast(list[str], metadata["catalog_part_numbers"])) == catalog_parts
        assert blocked_parts == {"B95702"}
        assert quote_ready_parts == {"B95701"}
        assert blocked_parts | quote_ready_parts == catalog_parts
        assert not blocked_parts & quote_ready_parts
        assert set(cast(list[str], metadata["part_numbers"])) == {"B95701"}
        assert cast(dict[str, list[str]], metadata["blocked_catalog_reasons"])["B95702"] == [
            "DEPENDENCY_UNRESOLVED"
        ]
        assert metadata["catalog_candidate_count"] == 2
        assert metadata["quote_ready_count"] == 1
        assert metadata["blocked_count"] == 1
        assert set(cast(dict[str, list[str]], metadata["mapping_ids_by_part"])) == {"B95701"}
