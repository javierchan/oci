"""Focused fixtures for the pure OCI commercial workbook parser."""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

from app.services.commercial_document_parser import (
    PRICE_LIST_SHEET,
    SUPPLEMENT_SHEET,
    parse_oci_commercial_workbook,
)


def _workbook_bytes() -> BytesIO:
    workbook = Workbook()
    price = workbook.active
    price.title = PRICE_LIST_SHEET
    supplement = workbook.create_sheet(SUPPLEMENT_SHEET)

    price.append(["Official OCI commercial evidence"])
    price.append(
        [
            "Service Category",
            "Service Name",
            "Pay as You Go",
            "Annual\nCommitment",
            "Metric",
            "Metric\nMinimum",
            "Additional\nInformation",
            "Notes",
            "Part\nNumber",
        ]
    )
    price.append(
        [
            "Oracle Data Management Cloud Services",
            "Oracle Autonomous AI Lakehouse - ECPU",
            0.336,
            0.336,
            "ECPU Per Hour",
            2,
            "Partial ECPU hours consumed are billed per second with a one-minute minimum.\n"
            "ECPU metric is service specific.",
            1,
            "B95701",
        ]
    )
    price.append([])
    price.append(
        [
            "Category",
            "Service Name",
            "Pay as You Go",
            "Annual Commitment",
            "Metric",
            "Metric Minimum",
            "Additional Information",
            "Notes",
            "Part Number",
        ]
    )
    price.append(
        [
            "Oracle Data Management Cloud Services",
            "Oracle Autonomous AI Lakehouse - ECPU",
            0.336,
            None,
            "ECPU Per Hour",
            None,
            None,
            None,
            " B95701 ",
        ]
    )
    price.append(
        [
            "Oracle Free Services",
            "Oracle Example Always Free",
            "Always Free",
            True,
            "Instance Per Month",
            "-",
            None,
            False,
            "B99999",
        ]
    )
    price.append(
        [
            "Oracle Data Integration Cloud Services",
            "Oracle Cloud Infrastructure - Data Integration - Pipeline Operator Execution",
            None,
            None,
            None,
            "-",
            None,
            1,
            "B93306",
        ]
    )
    price.append(
        [None, "First 30 Execution Hours", "Free Tier", "Free Tier", "Execution Hour", "-"]
    )
    price.append(
        [None, "Greater than 30 Execution Hours", 0.3, 0.3, "Execution Hour", "-"]
    )
    price.append([])
    price.append(
        [
            "Service Name",
            "Price",
            "Universal Credits Minimums (per year)",
            "Minimum Universal Credits Details",
            "Notes",
            "Part Number",
            "Universal Credits Minimum (per month)",
        ]
    )
    price.append(
        [
            "Oracle PaaS and IaaS Universal Credits",
            1,
            2000,
            "Minimum Annual Commitment spend in Universal Credits to be eligible for Universal Credits model",
            "1, 9",
            "B88206",
            2000,
        ]
    )
    price.append(
        [
            None,
            None,
            100000,
            "Minimum Annual Commitment spend in Universal Credits to be eligible for discount",
            None,
            None,
            100000,
        ]
    )

    supplement.append(["Supplement evidence"])
    supplement.append(
        [
            "Part\nNumber",
            "SUBSCRIPTION SERVICE",
            "Metric",
            "INCLUDED WITH SUBSCRIPTION SERVICE",
            "ADDITIONAL REQUIREMENTS AND PREREQUISITES",
        ]
    )
    supplement.append(
        [
            "B93306",
            "Oracle Cloud Infrastructure - Data Integration - Pipeline Operator Execution",
            None,
            "-",
            "-",
        ]
    )
    supplement.append([None, "First 30 Execution Hours", "Execution Hour", None, None])
    supplement.append([None, "Greater than 30 Execution Hours", "Execution Hour", None, None])
    supplement.append(
        [
            "B88206",
            "Oracle PaaS and IaaS Universal Credits",
            "Currency Unit",
            "Increases Cloud balance by 1 Currency Unit.",
            "-",
        ]
    )
    supplement.append(
        [
            "B95701",
            "Oracle Autonomous AI Lakehouse - ECPU",
            "ECPU Per Hour",
            "Includes entitlement for all database functionalities made available by the service.",
            "Oracle Autonomous Data Warehouse - ECPU serverless deployment requires: "
            "Oracle Autonomous Data Warehouse Exadata Storage for ECPU.",
        ]
    )
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output


def _hierarchical_workbook_bytes() -> BytesIO:
    workbook = Workbook()
    price = workbook.active
    price.title = PRICE_LIST_SHEET
    supplement = workbook.create_sheet(SUPPLEMENT_SHEET)

    price.append(["Section 1 - Universal Credits"])
    price.append(["PaaS - Oracle Management Cloud Services:"])
    price.append(["Database Management; Fleet Management; and Log Analytics"])
    for row_number in (2, 3):
        price.cell(row_number, 1).font = Font(bold=True, size=14)
        price.cell(row_number, 1).alignment = Alignment(horizontal="center")
    price.append(
        [
            "Service Name",
            "Pay as You Go",
            "Annual Commitment",
            "Metric",
            "Metric Minimum",
            "Additional Information",
            "Notes",
            "Part Number",
        ]
    )
    price.append(["Oracle Management Cloud Services"])
    price.append(["Oracle Monitoring and Diagnostics Services"])
    price.append(
        [
            "Oracle Cloud Infrastructure Log Analytics - Active Storage",
            372,
            372,
            "Logging Analytics Storage Unit Per Month",
            "-",
            None,
            "1, 13",
            "B95634",
        ]
    )

    supplement.append(
        [
            "Part Number",
            "Subscription Service",
            "Metric",
            "Included With Subscription Service",
            "Additional Requirements and Prerequisites",
        ]
    )
    supplement.append(
        [
            "B95634",
            "Oracle Cloud Infrastructure Log Analytics - Active Storage",
            "Logging Analytics Storage Unit Per Month",
            "-",
            "-",
        ]
    )

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output


def _late_price_column_workbook_bytes() -> BytesIO:
    """Reproduce official tables whose second Price column follows the SKU."""

    workbook = Workbook()
    price = workbook.active
    price.title = PRICE_LIST_SHEET
    supplement = workbook.create_sheet(SUPPLEMENT_SHEET)

    price.cell(1, 1, "Section 1 - Universal Credits")
    price.cell(2, 4, "Oracle IaaS Public Cloud Services")
    price.cell(3, 12, "Price")
    price.cell(3, 16, "Minimum")
    price.cell(3, 18, "Includes")
    price.cell(3, 32, "Notes")
    price.cell(3, 34, "Part Number")
    price.cell(3, 39, "Price")
    price.cell(4, 4, "Oracle IaaS Public Cloud Services")
    price.cell(4, 12, 1)
    price.cell(4, 16, "-")
    price.cell(4, 32, 1)
    price.cell(4, 34, "B77077")
    price.cell(4, 37, 1)
    price.cell(4, 39, 1)

    supplement.append(
        [
            "Part Number",
            "Subscription Service",
            "Metric",
            "Included With Subscription Service",
            "Additional Requirements and Prerequisites",
        ]
    )
    supplement.append(
        [
            "B77077",
            "Oracle IaaS Public Cloud Services",
            "Unit Per Month",
            "-",
            "-",
        ]
    )

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output


def test_parses_and_joins_b95701_commercial_semantics() -> None:
    parsed = parse_oci_commercial_workbook(_workbook_bytes())
    record = parsed.by_part_number()["B95701"]

    assert record.service_name == "Oracle Autonomous AI Lakehouse - ECPU"
    assert record.service_category == "Oracle Data Management Cloud Services"
    assert record.pay_as_you_go == Decimal("0.336")
    assert record.annual_commitment == Decimal("0.336")
    assert [term.term_type for term in record.commercial_price_terms] == [
        "pay_as_you_go",
        "annual_commitment",
    ]
    assert record.metric == "ECPU Per Hour"
    assert record.metric_minimum == Decimal("2")
    assert "billed per second with a one-minute minimum" in (record.additional_information or "")
    assert "all database functionalities" in (record.included_entitlements or "")
    assert "Exadata Storage for ECPU" in (record.prerequisites or "")


def test_preserves_official_product_container_hierarchy() -> None:
    record = parse_oci_commercial_workbook(_hierarchical_workbook_bytes()).by_part_number()[
        "B95634"
    ]

    assert record.catalog_section == "Section 1 - Universal Credits"
    assert record.product_family == (
        "PaaS - Oracle Management Cloud Services: "
        "Database Management; Fleet Management; and Log Analytics"
    )
    assert record.product_group == "Oracle Management Cloud Services"
    assert record.service_category == "Oracle Monitoring and Diagnostics Services"
    assert record.product_hierarchy == (
        "Section 1 - Universal Credits",
        "PaaS - Oracle Management Cloud Services: "
        "Database Management; Fleet Management; and Log Analytics",
        "Oracle Management Cloud Services",
        "Oracle Monitoring and Diagnostics Services",
    )


def test_product_name_ignores_numeric_cells_before_a_late_price_column() -> None:
    parsed = parse_oci_commercial_workbook(_late_price_column_workbook_bytes())

    record = parsed.by_part_number()["B77077"]

    assert record.service_name == "Oracle IaaS Public Cloud Services"
    assert record.service_name != "1"


@pytest.mark.parametrize("sheet_name", [PRICE_LIST_SHEET, SUPPLEMENT_SHEET])
def test_rejects_unrecognized_official_sheet_layout(sheet_name: str) -> None:
    source = _workbook_bytes()
    workbook = load_workbook(source)
    sheet = workbook[sheet_name]
    sheet.delete_rows(1, sheet.max_row)
    sheet.append(["Unexpected", "Layout", "Without governed headers"])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)

    with pytest.raises(ValueError, match="layout is not recognized"):
        parse_oci_commercial_workbook(output)


def test_detects_repeated_headers_and_deduplicates_by_completeness() -> None:
    record = parse_oci_commercial_workbook(_workbook_bytes()).by_part_number()["B95701"]

    price_rows = [row for sheet, row in record.source_rows if sheet == PRICE_LIST_SHEET]
    assert price_rows == [3, 6]
    assert record.annual_commitment == Decimal("0.336")
    assert record.metric_minimum == Decimal("2")
    assert record.source_conflicts == ()


def test_repeated_sku_flags_material_identity_conflict_but_not_blank_overlap() -> None:
    source = _workbook_bytes()
    workbook = load_workbook(source)
    price = workbook[PRICE_LIST_SHEET]
    price.append(
        [
            "Service Category",
            "Service Name",
            "Pay as You Go",
            "Annual Commitment",
            "Metric",
            "Metric Minimum",
            "Additional Information",
            "Notes",
            "Part Number",
        ]
    )
    price.append(
        [
            "Oracle Data Management Cloud Services",
            "Oracle Autonomous AI Lakehouse - Different Product",
            0.336,
            None,
            "ECPU Per Hour",
            None,
            None,
            None,
            "B95701",
        ]
    )
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)

    record = parse_oci_commercial_workbook(output).by_part_number()["B95701"]

    assert record.source_conflicts == (
        {
            "field": "service_name",
            "values": [
                "Oracle Autonomous AI Lakehouse - Different Product",
                "Oracle Autonomous AI Lakehouse - ECPU",
            ],
        },
    )


def test_repeated_sku_flags_conflicting_price_within_the_same_commercial_scope() -> None:
    source = _workbook_bytes()
    workbook = load_workbook(source)
    price = workbook[PRICE_LIST_SHEET]
    price.append(
        [
            "Service Category",
            "Service Name",
            "Pay as You Go",
            "Annual Commitment",
            "Metric",
            "Metric Minimum",
            "Additional Information",
            "Notes",
            "Part Number",
        ]
    )
    price.append(
        [
            "Oracle Data Management Cloud Services",
            "Oracle Autonomous AI Lakehouse - ECPU",
            0.337,
            None,
            "ECPU Per Hour",
            2,
            None,
            None,
            "B95701",
        ]
    )
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)

    record = parse_oci_commercial_workbook(output).by_part_number()["B95701"]

    price_conflict = next(
        conflict
        for conflict in record.source_conflicts
        if conflict["field"] == "commercial_price_term"
    )
    assert price_conflict["values"] == ["0.336", "0.337"]


def test_canonical_product_placement_is_semantic_and_page_order_independent() -> None:
    source = _hierarchical_workbook_bytes()
    workbook = load_workbook(source)
    price = workbook[PRICE_LIST_SHEET]
    price.append(["Section 1 - Universal Credits"])
    price.append(["PaaS - Generic Cloud Services"])
    price.cell(price.max_row, 1).font = Font(bold=True, size=14)
    price.cell(price.max_row, 1).alignment = Alignment(horizontal="center")
    price.append(
        [
            "Service Name",
            "Pay as You Go",
            "Annual Commitment",
            "Metric",
            "Metric Minimum",
            "Additional Information",
            "Notes",
            "Part Number",
        ]
    )
    price.append(["Generic Product Group"])
    price.append(["Oracle Generic Diagnostics"])
    price.append(
        [
            "Oracle Cloud Infrastructure Log Analytics - Active Storage",
            372,
            372,
            "Logging Analytics Storage Unit Per Month",
            "-",
            None,
            "1, 13",
            "B95634",
        ]
    )
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)

    record = parse_oci_commercial_workbook(output).by_part_number()["B95634"]

    assert record.product_group == "Oracle Management Cloud Services"
    assert record.service_category == "Oracle Monitoring and Diagnostics Services"


def test_preserves_source_cells_for_both_official_sheets() -> None:
    record = parse_oci_commercial_workbook(_workbook_bytes()).by_part_number()["B95701"]

    assert record.source_sheets == (PRICE_LIST_SHEET, SUPPLEMENT_SHEET)
    evidence = {(item.sheet, item.row): item for item in record.source_evidence}
    assert any(cell.coordinate == "I3" and cell.value == "B95701" for cell in evidence[(PRICE_LIST_SHEET, 3)].cells)
    supplement_evidence = next(
        item
        for item in record.source_evidence
        if item.sheet == SUPPLEMENT_SHEET
        and any(cell.value == "B95701" for cell in item.cells)
    )
    assert any(
        cell.coordinate == f"A{supplement_evidence.row}" and cell.value == "B95701"
        for cell in supplement_evidence.cells
    )


def test_normalizes_always_free_booleans_and_blank_values() -> None:
    record = parse_oci_commercial_workbook(_workbook_bytes()).by_part_number()["B99999"]

    assert record.pay_as_you_go == "Always Free"
    assert record.annual_commitment is True
    assert record.metric_minimum == "-"
    assert record.additional_information is None
    assert record.notes == "False"


def test_inherits_commercial_continuation_rows_without_inventing_part_numbers() -> None:
    record = parse_oci_commercial_workbook(_workbook_bytes()).by_part_number()["B93306"]

    assert record.metric == "Execution Hour"
    assert [
        (term.term_type, term.value, term.source_label)
        for term in record.commercial_price_terms
    ] == [
        ("pay_as_you_go", "Free Tier", "First 30 Execution Hours"),
        ("annual_commitment", "Free Tier", "First 30 Execution Hours"),
        ("pay_as_you_go", Decimal("0.3"), "Greater than 30 Execution Hours"),
        ("annual_commitment", Decimal("0.3"), "Greater than 30 Execution Hours"),
    ]
    assert [row for sheet, row in record.source_rows if sheet == PRICE_LIST_SHEET] == [8, 9, 10]


def test_parses_universal_credit_threshold_continuation_as_terms() -> None:
    record = parse_oci_commercial_workbook(_workbook_bytes()).by_part_number()["B88206"]

    assert record.metric == "Currency Unit"
    assert record.pay_as_you_go == Decimal("1")
    assert [
        (term.term_type, term.value)
        for term in record.commercial_price_terms
    ] == [
        ("pay_as_you_go", Decimal("1")),
        ("annual_commitment_minimum", Decimal("2000")),
        ("monthly_commitment_minimum", Decimal("2000")),
        ("annual_commitment_minimum", Decimal("100000")),
        ("monthly_commitment_minimum", Decimal("100000")),
    ]


def test_persists_blank_state_for_governed_source_columns() -> None:
    record = parse_oci_commercial_workbook(_workbook_bytes()).by_part_number()["B95701"]
    repeated_row = next(
        evidence
        for evidence in record.source_evidence
        if evidence.sheet == PRICE_LIST_SHEET and evidence.row == 6
    )

    annual_commitment = next(cell for cell in repeated_row.cells if cell.coordinate == "D6")
    assert annual_commitment.value is None
    assert annual_commitment.state == "blank"


def test_requires_both_official_sheets() -> None:
    workbook = Workbook()
    workbook.active.title = PRICE_LIST_SHEET
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)

    try:
        parse_oci_commercial_workbook(output)
    except ValueError as exc:
        assert SUPPLEMENT_SHEET in str(exc)
    else:
        raise AssertionError("missing supplement sheet must be rejected")
