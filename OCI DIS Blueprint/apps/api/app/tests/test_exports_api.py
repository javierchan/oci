"""Integration coverage for export endpoints that do not require a seeded snapshot."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from fastapi import HTTPException
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncEngine, AsyncSession, async_sessionmaker

from app.models import CatalogIntegration, DashboardSnapshot, Project, VolumetrySnapshot
from app.services import capture_template_service, import_service

pytestmark = [
    pytest.mark.filterwarnings(
        "ignore:datetime.datetime.utcnow\\(\\) is deprecated and scheduled for removal.*:DeprecationWarning:openpyxl.packaging.core"
    ),
    pytest.mark.filterwarnings(
        "ignore:datetime.datetime.utcnow\\(\\) is deprecated and scheduled for removal.*:DeprecationWarning:openpyxl.writer.excel"
    ),
]


@pytest.mark.asyncio
async def test_capture_template_export_returns_valid_workbook(api_client: AsyncClient) -> None:
    """Verify the capture template endpoint returns a readable XLSX workbook."""

    response = await api_client.get("/api/v1/exports/template/xlsx")
    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(filename=BytesIO(response.content))
    assert workbook.active.title == "Start Here"
    sheet = workbook["Integration Catalog"]
    assert sheet["A1"].value == "#"
    assert sheet["B1"].value == "Interface ID"
    assert sheet["A2"].value is None
    assert sheet.freeze_panes == "F2"
    assert sheet.max_row <= 2
    assert all(cell.value is None for cell in sheet[2])
    assert sheet.auto_filter.ref == "A1:AQ501"
    assert workbook.sheetnames == [
        "Start Here",
        "Dashboard",
        "_Lists",
        "Client Catalogs",
        "Integration Catalog",
        "Preflight Validation",
        "Guided Examples",
        "Field Guide",
        "Patterns",
        "OCI Services",
        "OCI Limits",
        "Interoperability",
    ]
    assert workbook["_Lists"].sheet_state == "veryHidden"
    assert "LIST_FREQUENCY" in workbook.defined_names
    assert "LIST_PATTERNS" in workbook.defined_names
    assert len(sheet.data_validations.dataValidation) >= 6
    pattern_sheet = workbook["Patterns"]
    assert pattern_sheet["D4"].value == "Certification"
    assert pattern_sheet["E4"].value == "Certification Version"
    assert pattern_sheet["G4"].value == "Required Evidence"
    assert pattern_sheet["L4"].value == "Validation Controls"
    assert pattern_sheet.auto_filter.ref.endswith(f"V{pattern_sheet.max_row}")
    assert "v3.1.0" in response.headers["content-disposition"]


@pytest.mark.asyncio
async def test_capture_template_metadata_matches_download_contract(api_client: AsyncClient) -> None:
    """Expose one backend-owned version and column contract to the web app."""

    response = await api_client.get("/api/v1/exports/template/metadata")
    assert response.status_code == 200
    payload = response.json()
    assert payload["template_version"] == "3.1.0"
    assert payload["filename"] == "oci-dis-import-template-v3.1.0.xlsx"
    assert payload["capture_sheet"] == "Integration Catalog"
    assert payload["capture_row_limit"] == 500
    assert len(payload["columns"]) == 43
    required = {item["field"] for item in payload["columns"] if item["requirement"] == "Required"}
    assert required == {"brand", "business_process", "interface_name", "frequency", "source_system", "destination_system", "tbq"}


@pytest.mark.asyncio
async def test_capture_template_round_trip_imports_tbq_y_and_n_rows(
    test_engine: AsyncEngine,
    tmp_path: Path,
) -> None:
    """Current en-US template keeps TBQ Y and N rows in the technical catalog."""

    session_factory = async_sessionmaker(test_engine, expire_on_commit=False, class_=AsyncSession)
    async with session_factory() as session:
        project = Project(id="template-roundtrip", name="Template Roundtrip", owner_id="architect", status="active")
        session.add(project)
        await session.flush()
        workbook_bytes, _ = await capture_template_service.generate_capture_template(session)
        workbook = load_workbook(BytesIO(workbook_bytes))
        sheet = workbook["Integration Catalog"]
        header_columns = {cell.value: cell.column for cell in sheet[1]}
        values = {
            "#": 1,
            "Interface ID": "INT-ROUNDTRIP-001",
            "Brand": "Retail",
            "Business Process": "Order to Cash",
            "Interface Name": "Publish confirmed order",
            "Business Criticality": "High",
            "Frequency": "Every hour",
            "Trigger Type": "Event Trigger",
            "Target Latency SLA": "p95 < 5 seconds",
            "Payload per Execution (KB)": 150,
            "Fan-out (Yes/No)": "Yes",
            "# Destinations": 3,
            "Source System": "Oracle ERP Cloud",
            "Destination System": "Order Fulfillment",
            "Data / Security Classification": "Confidential",
            "Retention / Processing Window": "Retain 7 days",
            "TBQ": "Y",
            "Selected Pattern (Manual)": "#02",
            "Pattern Rationale (Manual)": "Decouples the producer from three consumers.",
            "Retry Policy": "3 attempts; exponential backoff; DLQ",
            "Idempotency": "Use orderId for deduplication",
            "Quantifiable Core Tools": "OCI Streaming | OIC Gen3",
            "Architectural Overlays": "OCI API Gateway | OCI APM",
        }
        for header, value in values.items():
            sheet.cell(2, header_columns[header], value)
        technical_only_values = {
            **values,
            "#": 2,
            "Interface ID": "INT-ROUNDTRIP-002",
            "Interface Name": "Publish unquoted order exception",
            "TBQ": "N",
        }
        for header, value in technical_only_values.items():
            sheet.cell(3, header_columns[header], value)
        file_path = tmp_path / "capture-v3.xlsx"
        workbook.save(file_path)

        batch = await import_service.create_import_batch(project.id, file_path.name, session)
        processed = await import_service.process_import(batch.id, str(file_path), session)
        await session.flush()
        integrations = list((await session.scalars(select(CatalogIntegration).where(CatalogIntegration.project_id == project.id))).all())

        assert processed.source_row_count == 2
        assert processed.loaded_count == 2
        assert processed.tbq_y_count == 1
        assert processed.tbq_n_count == 1
        assert processed.header_map is not None
        assert processed.header_map["__template_version__"] == "3.1.0"
        assert processed.header_map["__template_compatibility__"] == "current"
        assert len(integrations) == 2
        integration = next(item for item in integrations if item.interface_id == "INT-ROUNDTRIP-001")
        assert integration.interface_id == "INT-ROUNDTRIP-001"
        assert integration.trigger_type == "Event Trigger"
        assert integration.selected_pattern == "#02"
        assert integration.fan_out_targets == 3
        assert integration.core_tools == "OCI Streaming | OIC Gen3"
        assert integration.business_criticality == "High"
        assert integration.target_latency_sla == "p95 < 5 seconds"
        assert integration.data_security_classification == "Confidential"
        assert integration.retention_processing_window == "Retain 7 days"
        assert integration.idempotency == "Use orderId for deduplication"
        technical_only = next(item for item in integrations if item.interface_id == "INT-ROUNDTRIP-002")
        assert technical_only.tbq == "N"


@pytest.mark.asyncio
async def test_active_project_exports_its_catalog_in_the_official_template(
    api_client: AsyncClient,
    test_engine: AsyncEngine,
    tmp_path: Path,
) -> None:
    """An active project can continue offline in the same governed workbook contract."""

    session_factory = async_sessionmaker(test_engine, expire_on_commit=False, class_=AsyncSession)
    async with session_factory() as session:
        project = Project(
            id="template-project-export",
            name="Northwind Retail 2026",
            owner_id="architect",
            status="active",
        )
        integrations = [
            CatalogIntegration(
                id="template-project-export-1",
                project_id=project.id,
                seq_number=1,
                interface_id="INT-NW-001",
                owner="Retail Architecture",
                brand="Northwind Retail",
                business_process="Order to Cash",
                interface_name="Publish confirmed order",
                description="Publishes confirmed orders to fulfillment.",
                business_criticality="High",
                complexity="High",
                frequency="Every hour",
                is_real_time=True,
                target_latency_sla="p95 < 5 seconds",
                trigger_type="Event Trigger",
                payload_per_execution_kb=150.0,
                is_fan_out=True,
                fan_out_targets=3,
                source_system="Oracle Retail Merchandising",
                source_technology="REST API",
                destination_system="Order Fulfillment",
                destination_technology_1="OCI Streaming",
                data_security_classification="Confidential",
                tbq="Y",
                selected_pattern="#02",
                pattern_rationale="Decouples one producer from three consumers.",
                retry_policy="3 attempts; DLQ",
                idempotency="orderId",
                core_tools="OCI Streaming | OIC Gen3",
                additional_tools_overlays="OCI API Gateway",
            ),
            CatalogIntegration(
                id="template-project-export-2",
                project_id=project.id,
                seq_number=2,
                interface_id="INT-NW-002",
                brand="Northwind Retail",
                business_process="Inventory Synchronization",
                interface_name="Publish unquoted inventory exception",
                frequency="Once daily",
                source_system="Inventory Hub",
                destination_system="Store Operations",
                tbq="N",
                is_fan_out=False,
            ),
        ]
        session.add(project)
        session.add_all(integrations)
        await session.commit()

    response = await api_client.get(f"/api/v1/exports/{project.id}/template/xlsx")
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "northwind-retail-2026-integration-catalog-v3.1.0.xlsx" in response.headers["content-disposition"]

    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["Integration Catalog"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.cell(2, headers["Interface ID"]).value == "INT-NW-001"
    assert sheet.cell(2, headers["Fan-out (Yes/No)"]).value == "Yes"
    assert sheet.cell(2, headers["TBQ"]).value == "Y"
    assert sheet.cell(3, headers["Interface ID"]).value == "INT-NW-002"
    assert sheet.cell(3, headers["Fan-out (Yes/No)"]).value == "No"
    assert sheet.cell(3, headers["TBQ"]).value == "N"
    assert "Northwind Retail 2026" in workbook["Start Here"]["A2"].value
    brand_suggestions = [workbook["Client Catalogs"].cell(row, 1).value for row in range(5, 15)]
    assert "Northwind Retail" in brand_suggestions

    async with session_factory() as session:
        target = Project(
            id="template-project-export-roundtrip",
            name="Northwind Re-import",
            owner_id="architect",
            status="active",
        )
        session.add(target)
        await session.flush()
        export_path = tmp_path / "northwind-project-export.xlsx"
        export_path.write_bytes(response.content)
        batch = await import_service.create_import_batch(target.id, export_path.name, session)
        processed = await import_service.process_import(batch.id, str(export_path), session)
        await session.flush()
        imported = list(
            (
                await session.scalars(
                    select(CatalogIntegration).where(CatalogIntegration.project_id == target.id)
                )
            ).all()
        )

        assert processed.loaded_count == 2
        assert {item.interface_id for item in imported} == {"INT-NW-001", "INT-NW-002"}
        assert next(item for item in imported if item.interface_id == "INT-NW-002").tbq == "N"


@pytest.mark.asyncio
async def test_archived_project_cannot_export_the_official_capture_template(
    api_client: AsyncClient,
    test_engine: AsyncEngine,
) -> None:
    """Archived workspaces cannot create an editable offline continuation."""

    session_factory = async_sessionmaker(test_engine, expire_on_commit=False, class_=AsyncSession)
    async with session_factory() as session:
        project = Project(
            id="template-project-archived",
            name="Archived Template Export",
            owner_id="architect",
            status="archived",
        )
        session.add(project)
        await session.commit()

    response = await api_client.get(f"/api/v1/exports/{project.id}/template/xlsx")
    assert response.status_code == 409
    assert response.json()["detail"]["error_code"] == "PROJECT_NOT_ACTIVE"


@pytest.mark.asyncio
async def test_capture_template_rejects_formulas(
    test_engine: AsyncEngine,
    tmp_path: Path,
) -> None:
    """Capture formulas cannot hide logic or executable spreadsheet payloads."""

    session_factory = async_sessionmaker(test_engine, expire_on_commit=False, class_=AsyncSession)
    async with session_factory() as session:
        project = Project(id="template-formula", name="Template Formula", owner_id="architect", status="active")
        session.add(project)
        await session.flush()
        workbook_bytes, _ = await capture_template_service.generate_capture_template(session)
        workbook = load_workbook(BytesIO(workbook_bytes))
        workbook["Integration Catalog"]["E2"] = '=HYPERLINK("https://example.com","Open")'
        file_path = tmp_path / "capture-formula.xlsx"
        workbook.save(file_path)
        batch = await import_service.create_import_batch(project.id, file_path.name, session)
        with pytest.raises(HTTPException) as exc_info:
            await import_service.process_import(batch.id, str(file_path), session)
        detail = exc_info.value.detail
        assert isinstance(detail, dict)
        assert detail["error_code"] == "IMPORT_FORMULA_NOT_ALLOWED"


@pytest.mark.asyncio
async def test_unversioned_v1_workbook_remains_supported(
    test_engine: AsyncEngine,
    tmp_path: Path,
) -> None:
    """Existing unversioned capture workbooks import with an explicit legacy label."""

    session_factory = async_sessionmaker(test_engine, expire_on_commit=False, class_=AsyncSession)
    async with session_factory() as session:
        project = Project(id="template-v1", name="Template v1", owner_id="architect", status="active")
        session.add(project)
        await session.flush()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Catálogo de Integraciones"
        sheet.append([
            "#",
            "ID de Interfaz",
            "Marca",
            "Proceso de Negocio",
            "Interfaz",
            "Sistema de Origen",
            "Sistema de Destino",
            "Frecuencia",
            "TBQ",
        ])
        sheet.append([1, "INT-V1-001", "Retail", "Order to Cash", "Legacy capture", "ERP", "OMS", "Mensual", "Y"])
        file_path = tmp_path / "capture-v1.xlsx"
        workbook.save(file_path)
        batch = await import_service.create_import_batch(project.id, file_path.name, session)
        processed = await import_service.process_import(batch.id, str(file_path), session)
        assert processed.loaded_count == 1
        assert processed.header_map is not None
        assert processed.header_map["__template_compatibility__"] == "legacy_v1_accepted"


@pytest.mark.asyncio
async def test_v3_workbook_rejects_changed_headers(
    test_engine: AsyncEngine,
    tmp_path: Path,
) -> None:
    """A governed v3 manifest cannot be paired with silently changed capture headers."""

    session_factory = async_sessionmaker(test_engine, expire_on_commit=False, class_=AsyncSession)
    async with session_factory() as session:
        project = Project(id="template-headers", name="Template Headers", owner_id="architect", status="active")
        session.add(project)
        await session.flush()
        workbook_bytes, _ = await capture_template_service.generate_capture_template(session)
        workbook = load_workbook(BytesIO(workbook_bytes))
        workbook["Integration Catalog"]["E1"] = "Invented Name"
        file_path = tmp_path / "capture-renamed.xlsx"
        workbook.save(file_path)
        batch = await import_service.create_import_batch(project.id, file_path.name, session)
        with pytest.raises(HTTPException) as exc_info:
            await import_service.process_import(batch.id, str(file_path), session)
        detail = exc_info.value.detail
        assert isinstance(detail, dict)
        assert detail["error_code"] == "IMPORT_TEMPLATE_HEADERS_CHANGED"


@pytest.mark.asyncio
async def test_project_brief_export_returns_markdown(
    api_client: AsyncClient,
    test_engine: AsyncEngine,
) -> None:
    """Verify the executive brief endpoint creates a governed Markdown artifact."""

    session_factory = async_sessionmaker(test_engine, expire_on_commit=False, class_=AsyncSession)
    async with session_factory() as session:
        project = Project(
            id="project-brief-1",
            name="Brief Fixture",
            owner_id="architect",
            status="active",
            description=None,
            project_metadata=None,
        )
        integration = CatalogIntegration(
            id="integration-brief-1",
            project_id=project.id,
            seq_number=1,
            interface_name="Order sync",
            interface_id="INT-001",
            source_system="POS",
            destination_system="OMS",
            selected_pattern="#01",
            core_tools="OIC Gen3",
            payload_per_execution_kb=25.0,
            qa_status="OK",
            qa_reasons=[],
        )
        snapshot = VolumetrySnapshot(
            id="snapshot-brief-1",
            project_id=project.id,
            assumption_set_version="1.0.0",
            triggered_by="test",
            row_results={},
            consolidated={
                "oic": {"total_billing_msgs_month": 10, "peak_packs_hour": 1},
                "data_integration": {"workspace_active": False, "data_processed_gb_month": 0},
                "functions": {"total_execution_units_gb_s": 0},
                "streaming": {},
                "queue": {},
            },
            snapshot_metadata={"integration_count": 1},
        )
        dashboard = DashboardSnapshot(
            id="dashboard-brief-1",
            project_id=project.id,
            volumetry_snapshot_id=snapshot.id,
            mode="technical",
            kpi_strip={
                "oic_msgs_month": 10,
                "peak_packs_hour": 1,
                "di_workspace_active": False,
                "di_data_processed_gb_month": 0,
                "functions_execution_units_gb_s": 0,
            },
            charts={
                "coverage": {
                    "total_integrations": 1,
                    "formal_id": {"complete": 1, "total": 1, "ratio": 1.0},
                    "pattern": {"complete": 1, "total": 1, "ratio": 1.0},
                    "payload": {"complete": 1, "total": 1, "ratio": 1.0},
                    "trigger": {"complete": 0, "total": 1, "ratio": 0.0},
                    "source_destination": {"complete": 1, "total": 1, "ratio": 1.0},
                    "fan_out": {"complete": 1, "total": 1, "ratio": 1.0},
                },
                "completeness": {
                    "qa_ok": 1,
                    "qa_revisar": 0,
                    "qa_pending": 0,
                    "rationale_informed": 0,
                    "core_tools_informed": 1,
                    "comments_informed": 0,
                    "retry_policy_informed": 0,
                },
                "pattern_mix": [{"pattern_id": "#01", "name": "Synchronous API", "count": 1}],
                "payload_distribution": [{"label": "0-50 KB", "count": 1}],
                "forecast_confidence": {
                    "level": "high",
                    "title": "High confidence",
                    "message": "Payload coverage is complete.",
                    "payload_coverage_ratio": 1.0,
                },
            },
            risks=[],
            maturity={
                "qa_ok_pct": 100.0,
                "pattern_assigned_pct": 100.0,
                "payload_informed_pct": 100.0,
                "governed_pct": 100.0,
            },
        )
        session.add_all([project, integration, snapshot, dashboard])
        await session.commit()

    response = await api_client.get("/api/v1/exports/project-brief-1/brief")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/markdown")
    assert "# OCI DIS Blueprint Executive Brief - Brief Fixture" in response.text
    assert "Catalog integrations: 1" in response.text
